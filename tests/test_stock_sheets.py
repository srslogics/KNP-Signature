from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from app import models
from app.stock import SOURCE_ITEMS, StockHistory


DAY = date(2026, 9, 2)


def outlet(db):
    result = models.Outlet(name=str(uuid4()), code=str(uuid4()))
    db.add(result)
    db.commit()
    return result


def scope(outlet):
    return {"mode": "single", "selected": outlet}


def opening(db, outlet, kg=100, nag=50, day=DAY - timedelta(days=1)):
    for item in SOURCE_ITEMS:
        db.add(models.DailyItemStock(outlet_id=outlet.id, date=day, item_type=item,
            actual_closing_weight=kg if item == "BB" else 0,
            actual_closing_quantity=nag if item == "BB" else 0))
    db.commit()


def txn(db, outlet, kind, kg=0, rate=0, **kwargs):
    record = models.Transaction(outlet_id=outlet.id, date=kwargs.pop("date", DAY), type=kind,
        item_type=kwargs.pop("item_type", "BB"), weight=kg, rate=rate,
        quantity=kwargs.pop("quantity", 0), amount=kwargs.pop("amount", Decimal(str(kg)) * Decimal(str(rate))), **kwargs)
    db.add(record)
    db.commit()
    return record


def sheet(endpoints, db, outlet, day=DAY, kind="stock"):
    return endpoints["daily_sheet"](str(day), kind, db, None, scope(outlet))


def process(endpoints, db, outlet, kg=80, day=DAY, item="BB"):
    result = endpoints["process_day_items"](str(day), [{"item_type": item, "actual_weight": kg, "actual_quantity": 40}], db, outlet)
    assert result.get("status") == "success", result
    return result


def report(endpoints, db, outlet, kind, day=DAY):
    return endpoints["export_report"](kind, file_format="json", date=str(day), db=db, scope=scope(outlet))


def test_mortality_matches_process_sheet_inventory_and_dashboard(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "PURCHASE", 50, 100, quantity=25)
    txn(db, o, "SALE", 60, 120, quantity=30)
    txn(db, o, "MORTALITY", 10, category="SHOP MORTALITY", quantity=5)
    saved = process(endpoints, db, o)
    assert saved["total_expected_stock"] == 80
    assert saved["total_expected_nag"] == 40
    assert saved["total_leakage"] == 0
    snapshot = db.query(models.DailyStock).filter_by(outlet_id=o.id, date=DAY).one()
    assert snapshot.expected_closing_weight == 80
    assert snapshot.leakage == 0
    inventory = report(endpoints, db, o, "inventory")["rows"][0]
    assert inventory["Expected Kg"] == 80
    dashboard = endpoints["get_dashboard"](str(DAY), db, scope(o))
    assert dashboard["expected_stock"] == 80, dashboard
    assert dashboard["leakage"] == 0


def test_unentered_actual_is_not_zero_or_a_shortage(db, endpoints):
    o = outlet(db)
    opening(db, o)
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["actual_stock"]["weight"] is None
    assert data["final_stock"]["short_by"]["weight"] is None
    assert next(card for card in data["metric_cards"] if card["label"] == "Leakage %")["value"] is None
    process(endpoints, db, o, 0)
    assert sheet(endpoints, db, o)["final_stock"]["short_by"]["weight"] == 100


def test_no_purchase_day_values_opening_stock_and_cost_of_sales(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "PURCHASE", 100, 100, date=DAY - timedelta(days=1))
    txn(db, o, "SALE", 20, 120, quantity=10)
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["closing_stock"]["total"] == 8000
    assert data["final_stock"]["gross_profit"]["total"] == 400
    assert data["business_controls"]["item_performance"][0]["gross_profit"] == 400
    rows = report(endpoints, db, o, "summary")["rows"]
    assert next(row for row in rows if row["Date"] == "02/09/2026")["Profit"] == 400


def test_missing_cost_never_becomes_zero_cost_profit(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "SALE", 20, 120)
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["gross_profit"]["total"] is None
    assert data["business_controls"]["item_performance"][0]["gross_profit"] is None


def test_backdated_actual_correction_refreshes_following_snapshots(db, endpoints):
    o = outlet(db)
    opening(db, o)
    process(endpoints, db, o, 80)
    process(endpoints, db, o, 70, DAY + timedelta(days=1))
    process(endpoints, db, o, 60)
    next_day = db.query(models.DailyStock).filter_by(outlet_id=o.id, date=DAY + timedelta(days=1)).one()
    assert next_day.opening_weight == 60
    assert next_day.leakage == -10
    assert next_day.actual_closing_weight == 70


def test_backdated_sales_change_is_recomputed_on_read(db, endpoints):
    o = outlet(db)
    opening(db, o)
    process(endpoints, db, o, 80)
    txn(db, o, "SALE", 20, 120)
    assert sheet(endpoints, db, o)["final_stock"]["closing_stock"]["weight"] == 80
    assert report(endpoints, db, o, "inventory")["rows"][0]["Expected Kg"] == 80
    assert endpoints["get_dashboard"](str(DAY), db, scope(o))["expected_stock"] == 80


def test_no_setup_or_older_day_fallback(db, endpoints):
    o = outlet(db)
    db.add(models.ItemOpeningStock(outlet_id=o.id, date=DAY - timedelta(days=30), item_type="BB", opening_weight=999))
    db.commit()
    opening(db, o, 250, day=DAY - timedelta(days=2))
    assert sheet(endpoints, db, o)["opening_stock"]["total"]["weight"] is None
    saved = process(endpoints, db, o, 50)
    assert saved["total_expected_stock"] is None
    assert report(endpoints, db, o, "inventory")["rows"][0]["Opening Kg"] is None
    assert sheet(endpoints, db, o, DAY + timedelta(days=1))["opening_stock"]["total"]["weight"] == 50


def test_lowercase_and_subtypes_share_live_buckets(db, endpoints):
    o = outlet(db)
    opening(db, o, 0)
    txn(db, o, "PURCHASE", 100, 100, item_type="bb")
    txn(db, o, "SALE", 20, 120, item_type=" BB HOTEL ")
    saved = process(endpoints, db, o, 80, item="bb")
    assert saved["total_actual_stock"] == 80
    assert saved["total_expected_stock"] == 80
    assert db.query(models.DailyItemStock).filter_by(outlet_id=o.id, date=DAY).count() == 5


@pytest.mark.parametrize("value", ["NaN", "Infinity", "-1"])
def test_invalid_actuals_are_rejected_without_deleting_old_counts(db, endpoints, value):
    o = outlet(db)
    process(endpoints, db, o, 30)
    result = endpoints["process_day_items"](str(DAY), [{"item_type": "BB", "actual_weight": value}], db, o)
    assert "error" in result
    assert sheet(endpoints, db, o)["final_stock"]["actual_stock"]["weight"] == 30


@pytest.mark.parametrize("paid", [0, 200, 500])
def test_retail_cash_credit_and_partial_match_ledger_in_reports(db, endpoints, paid):
    o = outlet(db)
    p = models.Party(name="Customer", type="VENDOR")
    db.add(p)
    db.flush()
    bill = models.RetailBill(id=uuid4(), outlet_id=o.id, party_id=p.id, date=DAY, bill_number="47",
        total_amount=500, paid_amount=paid, outstanding_amount=500-paid, payment_mode="Cash")
    db.add(bill)
    db.commit()
    txn(db, o, "SALE", 5, 100, party_id=p.id, category="RETAIL", bill_number="47", source_ref=f"retail-bill:{bill.id}:1")
    if 0 < paid < 500:
        txn(db, o, "PAYMENT", amount=paid, party_id=p.id, category="RECEIVED", source_ref=f"retail-payment:{bill.id}")
    vendor = sheet(endpoints, db, o, kind="vendor")["totals"]
    assert vendor["purchases"] == 500
    assert vendor["payment"] == paid
    assert vendor["balance"] == 500-paid
    summary = report(endpoints, db, o, "summary")["rows"][0]
    assert summary["Sales"] == 500
    assert summary["Payment Received"] == paid
    assert endpoints["get_dashboard"](str(DAY), db, scope(o))["payments_received"] == paid


def test_dealer_balance_history_and_outlet_scope(db, endpoints):
    o, other = outlet(db), outlet(db)
    party = models.Party(name="Dealer", type="BOTH")
    db.add(party)
    db.commit()
    txn(db, o, "OPENING", category="PAYABLE", party_id=party.id, amount=1000, date=DAY-timedelta(days=1))
    txn(db, o, "PURCHASE", 5, 100, party_id=party.id)
    txn(db, o, "PAYMENT", category="PAID", party_id=party.id, amount=200)
    txn(db, other, "PURCHASE", 100, 100, party_id=party.id)
    data = sheet(endpoints, db, o, kind="dealer")["totals"]
    assert data["old_balance"] == 1000
    assert data["balance"] == 1300


def test_credit_excel_has_correct_columns_and_payment_mode(db, endpoints):
    response = endpoints["build_daily_sheet_export_report"]({"retail_credit_sheet": {"rows": [
        {"customer_name": "Customer", "bill_number": "47", "total_amount": 1000,
         "paid_amount": 200, "outstanding_amount": 800, "payment_mode": "Online"}
    ]}}, "stock", DAY)
    workbook = load_workbook(BytesIO(response.body))
    rows = list(workbook.active.values)
    header = next(row for row in rows if row[0] == "Section")
    record = dict(zip(header, next(row for row in rows if row[0] == "Retail Credit")))
    assert record["Bill No"] == "47"
    assert record["Bill Amount"] == 1000
    assert record["Mode"] == "Online"
    assert record["NAG"] is None and record["Weight"] is None


def test_dressing_deducts_live_input_once_and_carries_dressed_cost(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "PURCHASE", 100, 100, date=DAY-timedelta(days=1))
    db.add(models.DressedStockEntry(outlet_id=o.id, date=DAY, item_name="BB", live_quantity=5,
                                   live_weight=10, dressed_weight=8))
    db.commit()
    txn(db, o, "SALE", 4, 200, category="RETAIL DRESSED")
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["closing_stock"]["weight"] == 90
    assert data["final_stock"]["closing_stock"]["nag"] == 45
    assert data["final_stock"]["gross_profit"]["total"] == 300
    process(endpoints, db, o, 90)
    txn(db, o, "SALE", 4, 200, category="RETAIL DRESSED", date=DAY+timedelta(days=1))
    next_day = sheet(endpoints, db, o, DAY+timedelta(days=1))
    assert next_day["final_stock"]["closing_stock"]["weight"] == 90
    assert next_day["final_stock"]["gross_profit"]["total"] == 300


def test_unmapped_cut_does_not_guess_bb(db, endpoints):
    o = outlet(db)
    opening(db, o)
    db.add(models.DressedStockEntry(outlet_id=o.id, date=DAY, item_name="Dressed Chicken", live_quantity=5, live_weight=10))
    db.commit()
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["closing_stock"]["weight"] is None
    assert "live hen type" in data["stock_warning"]
    assert all(row["Expected Kg"] is None for row in report(endpoints, db, o, "inventory")["rows"])


def test_weighted_cost_includes_opening_stock_and_todays_purchases(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "PURCHASE", 100, 100, date=DAY-timedelta(days=1))
    txn(db, o, "PURCHASE", 100, 140)
    txn(db, o, "SALE", 50, 160)
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["closing_stock"]["total"] == 18000
    assert data["final_stock"]["gross_profit"]["total"] == 2000


def test_analytics_use_same_cost_and_missing_actual_rules(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "PURCHASE", 100, 100, date=DAY-timedelta(days=1))
    txn(db, o, "SALE", 20, 120)
    summary = endpoints["analytics_summary"](str(DAY), str(DAY), db, scope(o))
    assert summary["profit"] == 400
    assert summary["leakage"] is None
    assert endpoints["profit_by_item"](str(DAY), str(DAY), db, scope(o))[0]["profit"] == 400
    assert endpoints["get_trend"](str(DAY), str(DAY), db, scope(o))[0]["profit"] == 400
    process(endpoints, db, o, 80)
    assert endpoints["leakage_trend"](str(DAY), str(DAY), db, scope(o))[0]["leakage"] == 0


def test_all_outlet_stock_does_not_hide_missing_counts(db, endpoints):
    first, second = outlet(db), outlet(db)
    opening(db, first)
    opening(db, second)
    process(endpoints, db, first, 80)
    all_scope = {"mode": "all", "selected": None, "outlets": [first, second]}
    dashboard = endpoints["get_dashboard"](str(DAY), db, all_scope)
    assert dashboard["expected_stock"] == 200
    assert dashboard["leakage"] is None


def test_duplicate_alias_counts_are_rejected_without_replacing_actuals(db, endpoints):
    o = outlet(db)
    process(endpoints, db, o, 80)
    data = endpoints["process_day_items"](str(DAY), [
        {"item_type": "BB", "actual_weight": 10},
        {"item_type": "BB HOTEL", "actual_weight": 20},
    ], db, o)
    assert "Duplicate" in data["error"]
    assert sheet(endpoints, db, o)["final_stock"]["actual_stock"]["weight"] == 80


def test_missing_nag_does_not_claim_complete_bird_reconciliation(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "SALE", 10, 120, quantity=None)
    txn(db, o, "SALE", 10, 120, quantity=5)
    data = sheet(endpoints, db, o)
    assert data["final_stock"]["closing_stock"]["weight"] == 80
    assert data["final_stock"]["closing_stock"]["nag"] == ""
    assert StockHistory(db, o.id, DAY).snapshot(DAY)["totals"]["expected_closing_quantity"] is None


def test_invalid_dressed_batch_preserves_form_retry_without_partial_insert(db, endpoints):
    o = outlet(db)
    payload = {"rows": [
        {"item_name": "BB", "live_quantity": 5, "live_weight": 10, "dressed_weight": 8},
        {"item_name": "CB", "live_quantity": 2, "live_weight": 4, "dressed_weight": 5},
    ]}
    result = endpoints["create_dressed_stock_entries"](payload, str(DAY), db, o)
    assert "error" in result
    assert db.query(models.DressedStockEntry).count() == 0
    payload["rows"][1]["dressed_weight"] = 3
    assert endpoints["create_dressed_stock_entries"](payload, str(DAY), db, o)["rows_inserted"] == 2


def test_legacy_malformed_bill_reference_does_not_crash_balance_sheet(db, endpoints):
    o = outlet(db)
    party = models.Party(name="Legacy Customer", type="VENDOR")
    db.add(party)
    db.commit()
    txn(db, o, "SALE", 5, 100, party_id=party.id, source_ref="retail-bill:invalid:1")
    assert sheet(endpoints, db, o, kind="vendor")["totals"]["balance"] == 500


def test_inventory_exports_preserve_missing_counts_and_kg_precision(db, endpoints):
    o = outlet(db)
    opening(db, o)
    txn(db, o, "MORTALITY", Decimal("1.237"), category="TRANSPORTATION", quantity=1)
    excel = endpoints["export_report"]("inventory", file_format="excel", date=str(DAY), db=db, scope=scope(o))
    rows = list(load_workbook(BytesIO(excel.body)).active.values)
    header = next(row for row in rows if row[0] == "Date")
    bb = dict(zip(header, next(row for row in rows if len(row) > 1 and row[1] == "BB")))
    assert bb["Expected Kg"] == 98.763
    assert bb["Actual Kg"] is None
    assert bb["Leakage Kg"] is None
    pdf = endpoints["export_report"]("inventory", file_format="pdf", date=str(DAY), db=db, scope=scope(o))
    assert pdf.body.startswith(b"%PDF-")
    assert b"98.763" in pdf.body
    assert endpoints["pdf_format_value"]("Mortality Kg", 1.237) == "1.237"
