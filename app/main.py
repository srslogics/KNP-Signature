from uuid import UUID, uuid4
from io import BytesIO
from urllib.parse import quote
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import hashlib
import hmac
import os
import secrets

from fastapi import FastAPI, HTTPException, Header
from app.db import engine, Base
from fastapi import UploadFile, File, Depends, Body
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.db import SessionLocal
from app import models
from app.stock import (
    SOURCE_ITEMS, SOURCE_ALIASES, StockHistory, normalize_source,
    complete_sum, stock_value, refresh_stock_snapshots,
)
from app.finance import (
    ACCOUNT_LEDGER,
    LEGACY_LEDGER,
    LEDGER_CUTOVER_DATE,
    build_account_ledger,
    build_legacy_ledger,
    empty_balances,
    ledger_mode_for_date,
    normalize_balances,
    retail_bill_id_from_source_ref,
    summarize_transactions,
    summarize_legacy_transactions,
)
from sqlalchemy import case, func, text, or_, and_, exists, cast, String
from sqlalchemy.exc import OperationalError
from decimal import Decimal
import uvicorn
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, JSONResponse

app = FastAPI()
BUSINESS_TIMEZONE = ZoneInfo("Asia/Kolkata")


def ledger_today():
    return datetime.now(BUSINESS_TIMEZONE).date()


def ledger_range_error(start, end):
    effective_end = end or ledger_today()
    if start and start < LEDGER_CUTOVER_DATE <= effective_end:
        return "Choose a date range either up to 03/09/2026 or from 04/09/2026"
    return None


def safe_create_index(conn, sql: str):
    try:
        conn.execute(text(sql))
    except OperationalError as e:
        message = str(e).lower()
        if "deadlock detected" in message:
            print(f"Skipping index creation during startup due to transient deadlock: {sql}")
            return
        raise


def current_shared_document_number(target_date, db: Session, outlet_id) -> str:
    next_number = db.execute(
        text("""
            SELECT GREATEST(
                COALESCE((
                    SELECT MAX(CAST(NULLIF(regexp_replace(COALESCE(bill_number, ''), '[^0-9]', '', 'g'), '') AS INTEGER))
                    FROM retail_bills
                    WHERE date = :target_date AND outlet_id = :outlet_id
                ), 0),
                COALESCE((
                    SELECT MAX(CAST(NULLIF(regexp_replace(COALESCE(receipt_number, ''), '[^0-9]', '', 'g'), '') AS INTEGER))
                    FROM payment_receipts
                    WHERE date = :target_date AND outlet_id = :outlet_id
                ), 0),
                COALESCE((
                    SELECT next_number - 1
                    FROM document_number_counters
                    WHERE target_date = :target_date AND outlet_id = :outlet_id
                ), 0)
            ) + 1
        """),
        {
            "target_date": target_date,
            "outlet_id": str(outlet_id),
        }
    ).scalar_one()
    return str(next_number)


def reserve_shared_document_number(target_date, db: Session, outlet_id) -> str:
    reserved_number = db.execute(
        text("""
            WITH current_max AS (
                SELECT COALESCE(MAX(number_value), 0) AS max_number
                FROM (
                    SELECT CAST(regexp_replace(COALESCE(bill_number, ''), '[^0-9]', '', 'g') AS INTEGER) AS number_value
                    FROM retail_bills
                    WHERE date = :target_date
                      AND outlet_id = :outlet_id
                      AND COALESCE(regexp_replace(COALESCE(bill_number, ''), '[^0-9]', '', 'g'), '') <> ''
                    UNION ALL
                    SELECT CAST(regexp_replace(COALESCE(receipt_number, ''), '[^0-9]', '', 'g') AS INTEGER) AS number_value
                    FROM payment_receipts
                    WHERE date = :target_date
                      AND outlet_id = :outlet_id
                      AND COALESCE(regexp_replace(COALESCE(receipt_number, ''), '[^0-9]', '', 'g'), '') <> ''
                ) existing_numbers
            ),
            reserved AS (
                INSERT INTO document_number_counters (id, outlet_id, target_date, next_number)
                VALUES (
                    :id,
                    :outlet_id,
                    :target_date,
                    (SELECT max_number + 2 FROM current_max)
                )
            ON CONFLICT (outlet_id, target_date)
            DO UPDATE SET
                next_number = GREATEST(
                    document_number_counters.next_number,
                    (SELECT max_number + 1 FROM current_max)
                ) + 1,
                updated_at = now()
                RETURNING next_number - 1 AS reserved_number
            )
            SELECT reserved_number FROM reserved
        """),
        {
            "id": str(uuid4()),
            "outlet_id": str(outlet_id),
            "target_date": target_date,
        }
    ).scalar_one()
    return str(reserved_number)


def ensure_database_schema():
    with engine.begin() as conn:
        # Serialize boot-time schema work so parallel app instances don't deadlock
        # while creating tables and adding columns on the same relations.
        conn.execute(text("SELECT pg_advisory_xact_lock(4815162342)"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS parties (
                id UUID PRIMARY KEY,
                name VARCHAR NOT NULL,
                normalized_name VARCHAR,
                type VARCHAR,
                phone VARCHAR,
                address VARCHAR,
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS outlets (
                id UUID PRIMARY KEY,
                name VARCHAR NOT NULL UNIQUE,
                code VARCHAR UNIQUE,
                is_active VARCHAR NOT NULL DEFAULT 'true',
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS party_aliases (
                id UUID PRIMARY KEY,
                alias VARCHAR,
                normalized_alias VARCHAR,
                party_id UUID REFERENCES parties(id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS transactions (
                id UUID PRIMARY KEY,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                party_id UUID REFERENCES parties(id),
                type VARCHAR,
                category VARCHAR,
                item_type VARCHAR,
                quantity NUMERIC,
                weight NUMERIC,
                rate NUMERIC,
                amount NUMERIC,
                payment_mode VARCHAR,
                bill_number VARCHAR,
                source_ref VARCHAR NOT NULL DEFAULT '',
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS uploaded_files (
                id UUID PRIMARY KEY,
                file_hash VARCHAR UNIQUE,
                outlet_id UUID REFERENCES outlets(id),
                file_type VARCHAR,
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS daily_stock (
                id UUID PRIMARY KEY,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                opening_weight NUMERIC,
                purchase_weight NUMERIC,
                sales_weight NUMERIC,
                expected_closing_weight NUMERIC,
                actual_closing_weight NUMERIC,
                leakage NUMERIC
            )
        """))
        conn.execute(text("ALTER TABLE parties ADD COLUMN IF NOT EXISTS phone VARCHAR"))
        conn.execute(text("ALTER TABLE parties ADD COLUMN IF NOT EXISTS address VARCHAR"))
        conn.execute(text("ALTER TABLE transactions ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE transactions ADD COLUMN IF NOT EXISTS item_type VARCHAR"))
        conn.execute(text("ALTER TABLE transactions ADD COLUMN IF NOT EXISTS source_ref VARCHAR NOT NULL DEFAULT ''"))
        conn.execute(text("ALTER TABLE transactions ADD COLUMN IF NOT EXISTS quantity NUMERIC"))
        conn.execute(text("ALTER TABLE transactions ADD COLUMN IF NOT EXISTS bill_number VARCHAR"))
        conn.execute(text("UPDATE transactions SET source_ref = '' WHERE source_ref IS NULL"))
        conn.execute(text("ALTER TABLE retail_bill_items ADD COLUMN IF NOT EXISTS line_type VARCHAR NOT NULL DEFAULT 'STANDARD'"))
        conn.execute(text("ALTER TABLE retail_bill_items ADD COLUMN IF NOT EXISTS source_item_type VARCHAR"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                id UUID PRIMARY KEY,
                username VARCHAR NOT NULL UNIQUE,
                password_hash VARCHAR NOT NULL,
                role VARCHAR NOT NULL DEFAULT 'STAFF',
                display_name VARCHAR,
                is_active VARCHAR NOT NULL DEFAULT 'true',
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS user_outlet_access (
                id UUID PRIMARY KEY,
                user_id UUID NOT NULL REFERENCES users(id),
                outlet_id UUID NOT NULL REFERENCES outlets(id),
                created_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_user_outlet_access UNIQUE (user_id, outlet_id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS user_sessions (
                id UUID PRIMARY KEY,
                user_id UUID NOT NULL REFERENCES users(id),
                token VARCHAR NOT NULL UNIQUE,
                expires_at TIMESTAMP NOT NULL,
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS document_number_counters (
                id UUID PRIMARY KEY,
                outlet_id UUID NOT NULL REFERENCES outlets(id),
                target_date DATE NOT NULL,
                next_number INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT now(),
                updated_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_document_counter_per_day UNIQUE (outlet_id, target_date)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS retail_shortcuts (
                id UUID PRIMARY KEY,
                outlet_id UUID NOT NULL REFERENCES outlets(id),
                name VARCHAR NOT NULL,
                normalized_name VARCHAR NOT NULL,
                line_type VARCHAR NOT NULL DEFAULT 'STANDARD',
                source_item_type VARCHAR,
                unit VARCHAR NOT NULL DEFAULT 'KGS',
                rate NUMERIC,
                created_at TIMESTAMP DEFAULT now(),
                updated_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_retail_shortcut_per_outlet UNIQUE (outlet_id, normalized_name, line_type)
            )
        """))
        conn.execute(text("ALTER TABLE retail_shortcuts ADD COLUMN IF NOT EXISTS source_item_type VARCHAR"))
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_user_sessions_token ON user_sessions (token)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_user_sessions_user_id ON user_sessions (user_id)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_document_number_counters_outlet_date ON document_number_counters (outlet_id, target_date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_retail_shortcuts_outlet_line_type ON retail_shortcuts (outlet_id, line_type)")
        conn.execute(text("ALTER TABLE transactions DROP CONSTRAINT IF EXISTS unique_txn"))
        conn.execute(text("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1
                    FROM pg_constraint
                    WHERE conname = 'unique_txn'
                ) THEN
                    ALTER TABLE transactions
                    ADD CONSTRAINT unique_txn UNIQUE (date, outlet_id, party_id, weight, rate, type, category, item_type, bill_number, source_ref);
                END IF;
            END
            $$;
        """))
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_transactions_date_type ON transactions (date, type)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_transactions_party_date ON transactions (party_id, date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_transactions_item_date ON transactions (item_type, date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_transactions_outlet_date ON transactions (outlet_id, date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_party_alias_normalized ON party_aliases (normalized_alias)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_daily_stock_date ON daily_stock (date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_retail_bills_date ON retail_bills (date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_retail_bill_items_bill_id ON retail_bill_items (bill_id)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_payment_receipts_date ON payment_receipts (date)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_user_outlet_access_user ON user_outlet_access (user_id)")
        safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_user_outlet_access_outlet ON user_outlet_access (outlet_id)")
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS item_opening_stock (
                id UUID PRIMARY KEY,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                item_type VARCHAR NOT NULL,
                opening_quantity NUMERIC,
                opening_weight NUMERIC,
                created_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_item_opening_stock UNIQUE (date, outlet_id, item_type)
            )
        """))
        conn.execute(text("ALTER TABLE item_opening_stock ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE item_opening_stock ADD COLUMN IF NOT EXISTS opening_quantity NUMERIC"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS daily_item_stock (
                id UUID PRIMARY KEY,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                item_type VARCHAR NOT NULL,
                opening_quantity NUMERIC,
                opening_weight NUMERIC,
                purchase_quantity NUMERIC,
                purchase_weight NUMERIC,
                sales_quantity NUMERIC,
                sales_weight NUMERIC,
                expected_closing_quantity NUMERIC,
                expected_closing_weight NUMERIC,
                actual_closing_quantity NUMERIC,
                actual_closing_weight NUMERIC,
                quantity_leakage NUMERIC,
                leakage NUMERIC,
                created_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_daily_item_stock UNIQUE (date, outlet_id, item_type)
            )
        """))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS opening_quantity NUMERIC"))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS purchase_quantity NUMERIC"))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS sales_quantity NUMERIC"))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS expected_closing_quantity NUMERIC"))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS actual_closing_quantity NUMERIC"))
        conn.execute(text("ALTER TABLE daily_item_stock ADD COLUMN IF NOT EXISTS quantity_leakage NUMERIC"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS retail_bills (
                id UUID PRIMARY KEY,
                bill_number VARCHAR NOT NULL,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                party_id UUID REFERENCES parties(id),
                customer_name VARCHAR,
                customer_phone VARCHAR,
                customer_address VARCHAR,
                cashier_name VARCHAR,
                payment_mode VARCHAR,
                total_quantity NUMERIC,
                total_weight NUMERIC,
                ice_amount NUMERIC,
                total_amount NUMERIC,
                paid_amount NUMERIC,
                outstanding_amount NUMERIC,
                notes VARCHAR,
                created_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_retail_bill_number_per_day UNIQUE (date, outlet_id, bill_number)
            )
        """))
        conn.execute(text("ALTER TABLE retail_bills ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE retail_bills ADD COLUMN IF NOT EXISTS ice_amount NUMERIC"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS retail_bill_items (
                id UUID PRIMARY KEY,
                bill_id UUID NOT NULL REFERENCES retail_bills(id),
                line_order INTEGER NOT NULL DEFAULT 1,
                item_name VARCHAR NOT NULL,
                line_type VARCHAR NOT NULL DEFAULT 'STANDARD',
                quantity NUMERIC,
                unit VARCHAR,
                weight NUMERIC,
                rate NUMERIC,
                amount NUMERIC,
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS dressed_stock_entries (
                id UUID PRIMARY KEY,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                item_name VARCHAR NOT NULL,
                live_quantity NUMERIC,
                live_weight NUMERIC,
                dressed_weight NUMERIC,
                remaining_dressed_weight NUMERIC,
                default_rate NUMERIC,
                notes VARCHAR,
                created_at TIMESTAMP DEFAULT now()
            )
        """))
        conn.execute(text("ALTER TABLE dressed_stock_entries ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS payment_receipts (
                id UUID PRIMARY KEY,
                receipt_number VARCHAR NOT NULL,
                date DATE NOT NULL,
                outlet_id UUID REFERENCES outlets(id),
                party_id UUID REFERENCES parties(id),
                party_name VARCHAR,
                party_phone VARCHAR,
                party_address VARCHAR,
                cashier_name VARCHAR,
                direction VARCHAR NOT NULL,
                payment_mode VARCHAR,
                amount NUMERIC,
                notes VARCHAR,
                created_at TIMESTAMP DEFAULT now(),
                CONSTRAINT unique_payment_receipt_number_per_day UNIQUE (date, outlet_id, receipt_number)
            )
        """))
        conn.execute(text("ALTER TABLE payment_receipts ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE uploaded_files ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE daily_stock ADD COLUMN IF NOT EXISTS outlet_id UUID"))
        conn.execute(text("ALTER TABLE daily_stock DROP CONSTRAINT IF EXISTS daily_stock_date_key"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_daily_stock_outlet_date_unique ON daily_stock (outlet_id, date)"))
        conn.execute(text("ALTER TABLE retail_bills DROP CONSTRAINT IF EXISTS unique_retail_bill_number_per_day"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_retail_bill_number_per_day_outlet ON retail_bills (date, outlet_id, bill_number)"))
        conn.execute(text("ALTER TABLE payment_receipts DROP CONSTRAINT IF EXISTS unique_payment_receipt_number_per_day"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_payment_receipt_number_per_day_outlet ON payment_receipts (date, outlet_id, receipt_number)"))
        conn.execute(text("ALTER TABLE item_opening_stock DROP CONSTRAINT IF EXISTS unique_item_opening_stock"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_item_opening_stock_outlet_date_type ON item_opening_stock (date, outlet_id, item_type)"))
        conn.execute(text("ALTER TABLE daily_item_stock DROP CONSTRAINT IF EXISTS unique_daily_item_stock"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_daily_item_stock_outlet_date_type ON daily_item_stock (date, outlet_id, item_type)"))

        default_outlet = conn.execute(text("SELECT id FROM outlets ORDER BY created_at ASC LIMIT 1")).scalar()
        if not default_outlet:
            default_outlet = str(uuid4())
            conn.execute(
                text("INSERT INTO outlets (id, name, code, is_active) VALUES (:id, :name, :code, 'true')"),
                {"id": default_outlet, "name": "Main Outlet", "code": "MAIN"}
            )

        conn.execute(text("UPDATE transactions SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE uploaded_files SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE item_opening_stock SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE daily_stock SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE daily_item_stock SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE retail_bills SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE dressed_stock_entries SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})
        conn.execute(text("UPDATE payment_receipts SET outlet_id = :outlet_id WHERE outlet_id IS NULL"), {"outlet_id": default_outlet})

def ensure_outlet_access_defaults():
    db = SessionLocal()
    try:
        default_outlet = get_default_outlet(db)
        all_outlets = db.query(models.Outlet).filter(
            func.lower(models.Outlet.is_active) == "true"
        ).all()
        users = db.query(models.User).all()
        for user in users:
            role = (user.role or ROLE_STAFF).upper()
            existing_ids = {
                str(row.outlet_id)
                for row in db.query(models.UserOutletAccess).filter(
                    models.UserOutletAccess.user_id == user.id
                ).all()
            }
            required_outlets = all_outlets if role == ROLE_OWNER else [default_outlet]
            for outlet in required_outlets:
                if str(outlet.id) in existing_ids:
                    continue
                db.add(models.UserOutletAccess(id=uuid4(), user_id=user.id, outlet_id=outlet.id))
        db.commit()
    finally:
        db.close()
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


SESSION_DAYS = 14
ROLE_OWNER = "OWNER"
ROLE_STAFF = "STAFF"
ALL_OUTLETS_TOKEN = "ALL"


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    derived = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000)
    return f"{salt}${derived.hex()}"


def verify_password(password: str, password_hash: str) -> bool:
    try:
        salt, stored = str(password_hash or "").split("$", 1)
    except ValueError:
        return False
    derived = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000).hex()
    return hmac.compare_digest(derived, stored)


def serialize_outlet(outlet: models.Outlet):
    return {
        "id": str(outlet.id),
        "name": outlet.name,
        "code": outlet.code or ""
    }


def get_user_accessible_outlets(db: Session, user: models.User):
    role = (user.role or ROLE_STAFF).upper()
    if role == ROLE_OWNER:
        return db.query(models.Outlet).filter(
            func.lower(models.Outlet.is_active) == "true"
        ).order_by(models.Outlet.name.asc()).all()

    outlet_ids = [
        row.outlet_id
        for row in db.query(models.UserOutletAccess).filter(
            models.UserOutletAccess.user_id == user.id
        ).all()
    ]
    if not outlet_ids:
        return []
    return db.query(models.Outlet).filter(
        models.Outlet.id.in_(outlet_ids),
        func.lower(models.Outlet.is_active) == "true"
    ).order_by(models.Outlet.name.asc()).all()


def get_default_outlet(db: Session):
    outlet = db.query(models.Outlet).order_by(models.Outlet.created_at.asc()).first()
    if outlet:
        return outlet
    outlet = models.Outlet(id=uuid4(), name="Main Outlet", code="MAIN", is_active="true")
    db.add(outlet)
    db.commit()
    db.refresh(outlet)
    return outlet


def serialize_user(user: models.User, db: Session | None = None):
    outlets = []
    if db is not None:
        outlets = [serialize_outlet(outlet) for outlet in get_user_accessible_outlets(db, user)]
    return {
        "id": str(user.id),
        "username": user.username,
        "role": (user.role or ROLE_STAFF).upper(),
        "display_name": user.display_name or user.username,
        "outlets": outlets,
        "can_view_all_outlets": (user.role or ROLE_STAFF).upper() == ROLE_OWNER
    }


def create_user_session(db: Session, user: models.User):
    token = secrets.token_urlsafe(32)
    session = models.UserSession(
        id=uuid4(),
        user_id=user.id,
        token=token,
        expires_at=datetime.utcnow() + timedelta(days=SESSION_DAYS)
    )
    db.add(session)
    db.flush()
    return session


def get_active_session(db: Session, token: str | None):
    if not token:
        return None
    session = db.query(models.UserSession).filter(
        models.UserSession.token == token
    ).first()
    if not session:
        return None
    if session.expires_at and session.expires_at < datetime.utcnow():
        db.delete(session)
        db.commit()
        return None
    return session


def get_current_user(
    db: Session = Depends(get_db),
    x_auth_token: str | None = Header(default=None)
):
    session = get_active_session(db, x_auth_token)
    if not session:
        raise HTTPException(status_code=401, detail="Authentication required")
    user = db.query(models.User).filter(models.User.id == session.user_id).first()
    if not user or str(user.is_active or "true").lower() != "true":
        raise HTTPException(status_code=401, detail="Authentication required")
    return user


def require_owner(user: models.User = Depends(get_current_user)):
    if (user.role or ROLE_STAFF).upper() != ROLE_OWNER:
        raise HTTPException(status_code=403, detail="Owner access required")
    return user


def get_outlet_scope(
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
    x_outlet_id: str | None = Header(default=None, alias="X-Outlet-Id")
):
    accessible_outlets = get_user_accessible_outlets(db, user)
    if not accessible_outlets:
        raise HTTPException(status_code=403, detail="No outlet access assigned")

    role = (user.role or ROLE_STAFF).upper()
    requested = str(x_outlet_id or "").strip()

    if role == ROLE_OWNER and requested.upper() == ALL_OUTLETS_TOKEN:
        return {"mode": "all", "outlets": accessible_outlets, "selected": None}

    if requested:
        for outlet in accessible_outlets:
            if str(outlet.id) == requested:
                return {"mode": "single", "outlets": accessible_outlets, "selected": outlet}
        raise HTTPException(status_code=403, detail="Outlet access denied")

    return {"mode": "single", "outlets": accessible_outlets, "selected": accessible_outlets[0]}


def get_current_outlet(scope=Depends(get_outlet_scope)):
    outlet = scope.get("selected")
    if not outlet:
        raise HTTPException(status_code=400, detail="Select one outlet for this action")
    return outlet


def resolve_scope_outlet_id(scope):
    selected = scope.get("selected")
    if selected:
        return selected.id
    outlet = scope.get("outlet")
    if outlet:
        return outlet.id
    outlet_id = scope.get("outlet_id")
    if outlet_id:
        return outlet_id
    raise KeyError("selected")


def apply_outlet_scope(query, model, scope):
    if scope["mode"] == "all":
        outlet_ids = [outlet.id for outlet in scope["outlets"]]
        return query.filter(model.outlet_id.in_(outlet_ids))
    return query.filter(model.outlet_id == resolve_scope_outlet_id(scope))


def outlet_scope_filter(model, scope):
    if scope["mode"] == "all":
        outlet_ids = [outlet.id for outlet in scope["outlets"]]
        return model.outlet_id.in_(outlet_ids)
    return model.outlet_id == resolve_scope_outlet_id(scope)


ensure_database_schema()
ensure_outlet_access_defaults()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # for now
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


PUBLIC_PATH_PREFIXES = (
    "/auth/",
    "/healthz",
)
PUBLIC_PATHS = {"/"}


@app.middleware("http")
async def auth_middleware(request, call_next):
    path = request.url.path
    if request.method == "OPTIONS":
        return await call_next(request)
    if path in PUBLIC_PATHS or any(path.startswith(prefix) for prefix in PUBLIC_PATH_PREFIXES):
        return await call_next(request)

    token = request.headers.get("X-Auth-Token")
    db = SessionLocal()
    try:
        session = get_active_session(db, token)
        if not session:
            return JSONResponse(status_code=401, content={"error": "Authentication required"})
        user = db.query(models.User).filter(models.User.id == session.user_id).first()
        if not user or str(user.is_active or "true").lower() != "true":
            return JSONResponse(status_code=401, content={"error": "Authentication required"})
    finally:
        db.close()

    return await call_next(request)

@app.get("/")
def root():
    return {"message": "Backend running"}


@app.get("/healthz")
def health_check():
    return {"status": "ok"}


@app.head("/healthz")
def health_check_head():
    return Response(status_code=200)


@app.get("/auth/setup-status")
def auth_setup_status(db: Session = Depends(get_db)):
    count = db.query(models.User).count()
    return {"has_users": count > 0}


@app.post("/auth/setup-owner")
def auth_setup_owner(payload: dict = Body(...), db: Session = Depends(get_db)):
    if db.query(models.User).count() > 0:
        return {"error": "Owner setup already completed"}

    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    display_name = str(payload.get("display_name") or username).strip()
    if len(username) < 3:
        return {"error": "Username must be at least 3 characters"}
    if len(password) < 4:
        return {"error": "Password must be at least 4 characters"}

    user = models.User(
        id=uuid4(),
        username=username,
        password_hash=hash_password(password),
        role=ROLE_OWNER,
        display_name=display_name,
        is_active="true"
    )
    db.add(user)
    db.flush()
    default_outlet = get_default_outlet(db)
    db.add(models.UserOutletAccess(id=uuid4(), user_id=user.id, outlet_id=default_outlet.id))
    session = create_user_session(db, user)
    db.commit()
    return {"user": serialize_user(user, db), "token": session.token}


@app.post("/auth/login")
def auth_login(payload: dict = Body(...), db: Session = Depends(get_db)):
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    user = db.query(models.User).filter(func.lower(models.User.username) == username.lower()).first()
    if not user or not verify_password(password, user.password_hash):
        return {"error": "Invalid username or password"}
    if str(user.is_active or "true").lower() != "true":
        return {"error": "User is inactive"}

    session = create_user_session(db, user)
    db.commit()
    return {"user": serialize_user(user, db), "token": session.token}


@app.get("/auth/me")
def auth_me(db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    return {"user": serialize_user(user, db)}


@app.post("/auth/logout")
def auth_logout(
    db: Session = Depends(get_db),
    x_auth_token: str | None = Header(default=None),
    user: models.User = Depends(get_current_user)
):
    session = get_active_session(db, x_auth_token)
    if session:
        db.delete(session)
        db.commit()
    return {"status": "ok"}


@app.get("/auth/users")
def auth_list_users(db: Session = Depends(get_db), user: models.User = Depends(require_owner)):
    users = db.query(models.User).order_by(models.User.username.asc()).all()
    return {"results": [serialize_user(item, db) for item in users]}


@app.post("/auth/users")
def auth_create_user(payload: dict = Body(...), db: Session = Depends(get_db), user: models.User = Depends(require_owner)):
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    display_name = str(payload.get("display_name") or username).strip()
    role = str(payload.get("role") or ROLE_STAFF).strip().upper()
    outlet_ids = payload.get("outlet_ids") or []

    if len(username) < 3:
        return {"error": "Username must be at least 3 characters"}
    if len(password) < 4:
        return {"error": "Password must be at least 4 characters"}
    if role not in [ROLE_OWNER, ROLE_STAFF]:
        return {"error": "Role must be OWNER or STAFF"}
    existing = db.query(models.User).filter(func.lower(models.User.username) == username.lower()).first()
    if existing:
        return {"error": "Username already exists"}

    new_user = models.User(
        id=uuid4(),
        username=username,
        password_hash=hash_password(password),
        role=role,
        display_name=display_name,
        is_active="true"
    )
    db.add(new_user)
    db.flush()

    accessible_outlets = db.query(models.Outlet).filter(
        func.lower(models.Outlet.is_active) == "true"
    ).order_by(models.Outlet.name.asc()).all()

    if role == ROLE_OWNER:
        assigned = accessible_outlets
    else:
        if outlet_ids:
            assigned = [outlet for outlet in accessible_outlets if str(outlet.id) in set(str(v) for v in outlet_ids)]
        else:
            default_outlet = get_default_outlet(db)
            assigned = [default_outlet]

    for outlet in assigned:
        db.add(models.UserOutletAccess(id=uuid4(), user_id=new_user.id, outlet_id=outlet.id))

    db.commit()
    db.refresh(new_user)
    return {"user": serialize_user(new_user, db)}


@app.get("/outlets")
def list_outlets(db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    outlets = get_user_accessible_outlets(db, user)
    return {
        "results": [serialize_outlet(outlet) for outlet in outlets],
        "can_view_all": (user.role or ROLE_STAFF).upper() == ROLE_OWNER
    }


@app.get("/retail-shortcuts")
def list_retail_shortcuts(
    db: Session = Depends(get_db),
    current_outlet: models.Outlet = Depends(get_current_outlet)
):
    rows = db.query(models.RetailShortcut).filter(
        models.RetailShortcut.outlet_id == current_outlet.id
    ).order_by(
        models.RetailShortcut.line_type.asc(),
        models.RetailShortcut.name.asc()
    ).all()

    return {
        "results": [
            {
                "id": str(row.id),
                "name": row.name or "",
                "rate": float(row.rate or 0),
                "line_type": row.line_type or "STANDARD",
                "source_item_type": row.source_item_type or "",
                "unit": row.unit or "KGS"
            }
            for row in rows
        ]
    }


@app.post("/retail-shortcuts")
def save_retail_shortcut(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_outlet: models.Outlet = Depends(get_current_outlet)
):
    shortcut_id_raw = str(payload.get("id") or "").strip()
    name = str(payload.get("name") or "").strip()
    line_type = str(payload.get("line_type") or "STANDARD").strip().upper() or "STANDARD"
    source_item_type = resolve_stock_source_type(name, payload.get("source_item_type") or "")
    unit = "KGS" if line_type == "DRESSED" else (str(payload.get("unit") or "KGS").strip().upper() or "KGS")
    rate = Decimal(str(payload.get("rate") or 0))

    if not name:
        return {"error": "Enter shortcut item name"}

    if line_type not in ["STANDARD", "DRESSED"]:
        return {"error": "Invalid shortcut type"}

    normalized_name = normalize_party_name(name)
    shortcut = None
    if shortcut_id_raw:
        try:
            shortcut = db.query(models.RetailShortcut).filter(
                models.RetailShortcut.id == UUID(shortcut_id_raw),
                models.RetailShortcut.outlet_id == current_outlet.id
            ).first()
        except ValueError:
            shortcut = None

    duplicate_query = db.query(models.RetailShortcut).filter(
        models.RetailShortcut.outlet_id == current_outlet.id,
        models.RetailShortcut.normalized_name == normalized_name,
        models.RetailShortcut.line_type == line_type
    )
    if shortcut:
        duplicate_query = duplicate_query.filter(models.RetailShortcut.id != shortcut.id)

    if duplicate_query.first():
        return {"error": "Shortcut already exists"}

    if shortcut is None:
        shortcut = models.RetailShortcut(
            outlet_id=current_outlet.id,
            name=name,
            normalized_name=normalized_name,
            line_type=line_type,
            source_item_type=source_item_type or None,
            unit=unit,
            rate=rate
        )
        db.add(shortcut)
    else:
        shortcut.name = name
        shortcut.normalized_name = normalized_name
        shortcut.line_type = line_type
        shortcut.source_item_type = source_item_type or None
        shortcut.unit = unit
        shortcut.rate = rate

    try:
        db.commit()
        db.refresh(shortcut)
    except Exception as e:
        db.rollback()
        return {"error": "Saving shortcut failed", "details": str(e)}

    return {
        "status": "success",
        "shortcut": {
            "id": str(shortcut.id),
            "name": shortcut.name or "",
            "rate": float(shortcut.rate or 0),
            "line_type": shortcut.line_type or "STANDARD",
            "source_item_type": shortcut.source_item_type or "",
            "unit": shortcut.unit or "KGS"
        }
    }


@app.delete("/retail-shortcuts/{shortcut_id}")
def delete_retail_shortcut(
    shortcut_id: UUID,
    db: Session = Depends(get_db),
    current_outlet: models.Outlet = Depends(get_current_outlet)
):
    shortcut = db.query(models.RetailShortcut).filter(
        models.RetailShortcut.id == shortcut_id,
        models.RetailShortcut.outlet_id == current_outlet.id
    ).first()
    if not shortcut:
        return {"error": "Shortcut not found"}

    try:
        db.delete(shortcut)
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Deleting shortcut failed", "details": str(e)}

    return {"status": "success"}


@app.post("/outlets")
def create_outlet(payload: dict = Body(...), db: Session = Depends(get_db), user: models.User = Depends(require_owner)):
    name = str(payload.get("name") or "").strip()
    code = str(payload.get("code") or "").strip() or None
    if len(name) < 2:
        return {"error": "Outlet name must be at least 2 characters"}

    existing = db.query(models.Outlet).filter(func.lower(models.Outlet.name) == name.lower()).first()
    if existing:
        return {"error": "Outlet already exists"}

    outlet = models.Outlet(id=uuid4(), name=name, code=code, is_active="true")
    db.add(outlet)
    db.flush()

    owners = db.query(models.User).filter(
        func.upper(func.coalesce(models.User.role, ROLE_STAFF)) == ROLE_OWNER
    ).all()
    for owner in owners:
        db.add(models.UserOutletAccess(id=uuid4(), user_id=owner.id, outlet_id=outlet.id))

    db.commit()
    db.refresh(outlet)
    return {"outlet": serialize_outlet(outlet)}


TEMPLATES = {
    "dealer": "DEALER,BILL_NO,HEN_TYPE,NAG,KGS,RATE_PER_KG\nABC Supplier,INV-101,Broiler,52,100,120\n",
    "vendor": "VENDOR,HEN_TYPE,NAG,KGS,RATE_PER_KG\nXYZ Hotel,Broiler,24,40,150\n",
    "payment": "DATE,PARTY,AMOUNT,PAYMENT_MODE,DIRECTION\n2026-04-21,XYZ Hotel,5000,Online,RECEIVED\n",
    "opening-balance": "DATE,PARTY,OPENING_BALANCE,BALANCE_TYPE\n2026-04-01,XYZ Hotel,25000,RECEIVABLE\n",
    "opening-stock": "DATE,HEN_TYPE,OPENING_NAG,OPENING_KGS\n2026-04-01,Broiler,220,500\n"
}


@app.get("/templates/{template_type}")
def download_template(template_type: str):
    template = TEMPLATES.get(template_type)
    if not template:
        return {"error": "Template not found"}

    return Response(
        content=template,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={template_type}_template.csv"}
    )


def normalize_party_name(name: str) -> str:
    return name.lower().replace(" ", "").replace(".", "")


def parse_input_date(value: str):
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def settled_retail_bill_keys(db: Session, txns):
    if not txns:
        return set()

    retail_keys = {
        (txn.outlet_id, txn.date, txn.party_id, str(txn.bill_number or "").strip())
        for txn in txns
        if txn.party_id
        and str(txn.bill_number or "").strip()
        and txn.type == "SALE"
        and (txn.category or "").upper() in ["RETAIL", "RETAIL DRESSED"]
    }

    if not retail_keys:
        return set()

    retail_outlets = {key[0] for key in retail_keys}
    retail_dates = {key[1] for key in retail_keys}
    retail_parties = {key[2] for key in retail_keys}
    retail_bill_numbers = {key[3] for key in retail_keys}

    return {
        (bill.outlet_id, bill.date, bill.party_id, str(bill.bill_number or "").strip())
        for bill in db.query(models.RetailBill).filter(
            models.RetailBill.outlet_id.in_(list(retail_outlets)),
            models.RetailBill.party_id.in_(list(retail_parties)),
            models.RetailBill.date.in_(list(retail_dates)),
            models.RetailBill.bill_number.in_(list(retail_bill_numbers)),
            models.RetailBill.outstanding_amount <= 0
        ).all()
    }


def summarize_ledger_transactions(db: Session, txns):
    bill_ids = set()
    for txn in txns:
        source_ref = str(getattr(txn, "source_ref", "") or "")
        bill_id = retail_bill_id_from_source_ref(source_ref)
        if not bill_id:
            continue
        try:
            bill_ids.add(UUID(bill_id))
        except ValueError:
            continue

    retail_bills = {
        str(bill.id): bill
        for bill in db.query(models.RetailBill).filter(
            models.RetailBill.id.in_(list(bill_ids))
        ).all()
    } if bill_ids else {}
    return summarize_transactions(txns, retail_bills)


def scoped_stock_histories(db, scope, end):
    outlet_ids = [outlet.id for outlet in scope["outlets"]] if scope["mode"] == "all" else [resolve_scope_outlet_id(scope)]
    return [StockHistory(db, outlet_id, end) for outlet_id in outlet_ids]


def stock_total_for_day(histories, day, key):
    return complete_sum(history.snapshot(day)["totals"][key] for history in histories)


def financial_events(db, scope, start=None, end=None):
    query = apply_outlet_scope(db.query(models.Transaction), models.Transaction, scope)
    if start:
        query = query.filter(models.Transaction.date >= start)
    if end:
        query = query.filter(models.Transaction.date <= end)
    return summarize_ledger_transactions(db, query.order_by(
        models.Transaction.date, models.Transaction.created_at, models.Transaction.id
    ).all())


def party_account_balances_as_of(db, scope, as_of):
    """Account balances from the cutover opening and later postings only."""
    query = apply_outlet_scope(db.query(models.Transaction).filter(
        models.Transaction.party_id.isnot(None),
        models.Transaction.date >= LEDGER_CUTOVER_DATE,
        models.Transaction.date <= as_of,
    ), models.Transaction, scope)
    txns = query.order_by(
        models.Transaction.party_id,
        models.Transaction.date,
        models.Transaction.created_at,
        models.Transaction.id,
    ).all()
    if not txns:
        return {}

    bill_ids = set()
    for txn in txns:
        bill_id = retail_bill_id_from_source_ref(txn.source_ref)
        if bill_id:
            try:
                bill_ids.add(UUID(bill_id))
            except ValueError:
                continue
    bills = {
        str(bill.id): bill
        for bill in db.query(models.RetailBill).filter(models.RetailBill.id.in_(bill_ids)).all()
    } if bill_ids else {}

    grouped = {}
    for txn in txns:
        grouped.setdefault(txn.party_id, []).append(txn)
    return {
        party_id: build_account_ledger(summarize_transactions(rows, bills))[0]
        for party_id, rows in grouped.items()
    }


def serialize_ledger_balances(balances):
    normalized = normalize_balances(balances)
    return {
        **{key: float(value) for key, value in normalized.items()},
        "balance": float(Decimal(balances.get("balance", normalized["net"]))),
    }


def payment_direction_balance(balances, direction):
    if balances.get("ledger_mode") == ACCOUNT_LEDGER:
        normalized = normalize_balances(balances)
        return normalized["payable"] if str(direction or "").upper() == "PAID" else normalized["receivable"]
    return Decimal(balances.get("balance", 0))


def build_ledger(db: Session, txns, opening_balances=None, mode=None, as_of=None):
    as_of = as_of or ledger_today()
    mode = mode or ledger_mode_for_date(as_of)
    txns = sorted(txns, key=lambda txn: (txn.date, txn.created_at or datetime.min, str(txn.id)))
    txns = [txn for txn in txns if txn.date <= as_of]

    if mode == ACCOUNT_LEDGER:
        txns = [txn for txn in txns if txn.date >= LEDGER_CUTOVER_DATE]
        balances, postings = build_account_ledger(
            summarize_ledger_transactions(db, txns),
            opening_balances,
        )
        balances["balance"] = balances["net"]
    else:
        txns = [txn for txn in txns if txn.date < LEDGER_CUTOVER_DATE]
        balances, _ = build_account_ledger(
            summarize_ledger_transactions(db, txns),
            opening_balances,
        )
        balances["balance"], postings = build_legacy_ledger(
            summarize_legacy_transactions(txns, settled_retail_bill_keys(db, txns)),
            (opening_balances or {}).get("balance", 0),
        )

    balances["ledger_mode"] = mode
    ledger = []
    for posting in postings:
        row = {
            "date": str(posting["date"]),
            "type": posting["type"],
            "bill_number": posting["bill_number"] or "",
            "category": posting["category"] or "",
            "item": posting["item"] or "",
            "quantity": float(Decimal(posting["quantity"] or 0)),
            "weight": float(Decimal(posting["weight"] or 0)),
            "rate": float(Decimal(posting.get("rate") or 0)),
            "payment_mode": posting["payment_mode"] or "NA",
            "amount": float(Decimal(posting["amount"] or 0)),
            "delta": float(posting["delta"]),
            "ledger_mode": mode,
        }
        if mode == ACCOUNT_LEDGER:
            row.update({
                "account": posting["account"],
                "debit": float(posting["debit"]),
                "credit": float(posting["credit"]),
                "account_balance": float(posting["account_balance"]),
                "net_balance": float(posting["net_balance"]),
                "balance": float(posting["net_balance"]),
            })
        else:
            row["balance"] = float(posting["balance"])
        ledger.append(row)
    return balances, ledger


def build_party_summary_for_period(ledger, opening_balances, closing_balances):
    opening = normalize_balances(opening_balances)
    closing = normalize_balances(closing_balances)
    total_sales = sum(Decimal(str(row["amount"] or 0)) for row in ledger if row["type"] == "SALE")
    total_purchase = sum(Decimal(str(row["amount"] or 0)) for row in ledger if row["type"] == "PURCHASE")
    total_received = sum(Decimal(str(row["amount"] or 0)) for row in ledger if row["type"] == "PAYMENT RECEIVED")
    total_paid = sum(Decimal(str(row["amount"] or 0)) for row in ledger if row["type"] == "PAYMENT PAID")

    return {
        "opening_balance": float(opening_balances.get("balance", opening["net"])),
        "opening_receivable": float(opening["receivable"]),
        "opening_payable": float(opening["payable"]),
        "total_sales": float(total_sales),
        "total_purchase": float(total_purchase),
        "total_received": float(total_received),
        "total_paid": float(total_paid),
        "receivable_balance": float(closing["receivable"]),
        "payable_balance": float(closing["payable"]),
        "current_balance": float(closing_balances.get("balance", closing["net"])),
        "ledger_mode": closing_balances.get("ledger_mode", LEGACY_LEDGER),
        "last_transaction_date": ledger[-1]["date"] if ledger else None,
    }


def build_party_ledger_window(db: Session, query, start=None, end=None):
    as_of = end or ledger_today()
    mode = ledger_mode_for_date(as_of)
    opening_balances = empty_balances()
    opening_balances["balance"] = Decimal("0")
    opening_balances["ledger_mode"] = mode
    if start:
        opening_txns = query.filter(
            models.Transaction.date < start
        )
        if mode == ACCOUNT_LEDGER:
            opening_txns = opening_txns.filter(models.Transaction.date >= LEDGER_CUTOVER_DATE)
        else:
            opening_txns = opening_txns.filter(models.Transaction.date < LEDGER_CUTOVER_DATE)
        opening_txns = opening_txns.order_by(
            models.Transaction.date.asc(),
            models.Transaction.created_at.asc(),
            models.Transaction.id.asc(),
        ).all()
        opening_balances, _ = build_ledger(db, opening_txns, mode=mode, as_of=as_of)

    range_query = query
    if start:
        range_query = range_query.filter(models.Transaction.date >= start)
    if mode == ACCOUNT_LEDGER:
        range_query = range_query.filter(models.Transaction.date >= LEDGER_CUTOVER_DATE)
    else:
        range_query = range_query.filter(models.Transaction.date < LEDGER_CUTOVER_DATE)
    if end:
        range_query = range_query.filter(models.Transaction.date <= end)
    else:
        range_query = range_query.filter(models.Transaction.date <= as_of)
    range_txns = range_query.order_by(
        models.Transaction.date.asc(),
        models.Transaction.created_at.asc(),
        models.Transaction.id.asc(),
    ).all()
    range_txns = filter_party_ledger_transactions(db, range_txns)

    closing_balances = opening_balances
    ledger = []
    if range_txns:
        closing_balances, ledger = build_ledger(
            db, range_txns, opening_balances=opening_balances, mode=mode, as_of=as_of
        )

    return opening_balances, closing_balances, range_txns, ledger, mode


def build_party_summary(db: Session, txns, balance):
    _, ledger = build_ledger(db, txns)
    return build_party_summary_for_period(ledger, empty_balances(), balance)


def receivable_delta(txn, settled_keys=None):
    amount = Decimal(txn.amount or 0)
    settled_keys = settled_keys or set()

    bill_key = (txn.outlet_id, txn.date, txn.party_id, str(txn.bill_number or "").strip())
    if (
        txn.type == "SALE"
        and (txn.category or "").upper() in ["RETAIL", "RETAIL DRESSED"]
        and bill_key in settled_keys
    ):
        return Decimal("0")

    if txn.type == "SALE" or (txn.type == "OPENING" and txn.category == "RECEIVABLE"):
        return amount

    if txn.type == "PAYMENT" and txn.category == "RECEIVED":
        return -amount

    return Decimal("0")


def payable_delta(txn):
    amount = Decimal(txn.amount or 0)

    if txn.type == "PURCHASE" or (txn.type == "OPENING" and txn.category == "PAYABLE"):
        return amount

    if txn.type == "PAYMENT" and txn.category == "PAID":
        return -amount

    return Decimal("0")


def build_balance_sheet_rows_from_ledger(
    db: Session,
    grouped_parties: dict,
    target_date,
    include_txn,
    running_delta,
    day_purchase_amount,
    day_payment_amount,
    opening_belongs_to_old
):
    result_rows = []
    total_old = Decimal("0")
    total_purchases = Decimal("0")
    total_payment = Decimal("0")
    total_balance = Decimal("0")

    # Use the same retail sale/payment events as the party ledger, including
    # fully paid bills. Fetch bill records once for the complete sheet.
    all_txns = [txn for party_data in grouped_parties.values() for txn in party_data["txns"]]
    bill_ids = set()
    for txn in all_txns:
        value = retail_bill_id_from_source_ref(txn.source_ref)
        if value:
            try:
                bill_ids.add(UUID(value))
            except ValueError:
                continue
    bills = {str(bill.id): bill for bill in db.query(models.RetailBill).filter(
        models.RetailBill.id.in_(bill_ids)
    ).all()} if bill_ids else {}

    for party_data in grouped_parties.values():
        txns = [txn for txn in party_data["txns"] if include_txn(txn)]
        if not txns:
            continue
        old_balance = Decimal("0")
        purchases = Decimal("0")
        payment = Decimal("0")
        for event in summarize_transactions(txns, bills):
            amount = Decimal(event["amount"])
            is_payment = event["entry_type"] == "PAYMENT"
            if event["date"] < target_date or event["entry_type"] == "OPENING":
                old_balance += -amount if is_payment else amount
            elif is_payment:
                payment += amount
            else:
                purchases += amount

        balance = old_balance + purchases - payment

        if old_balance == 0 and purchases == 0 and payment == 0 and balance == 0:
            continue

        result_rows.append(format_balance_row(party_data["party_name"], old_balance, purchases, payment, balance))
        total_old += old_balance
        total_purchases += purchases
        total_payment += payment
        total_balance += balance

    return {
        "rows": result_rows,
        "totals": format_balance_row("TOTAL", total_old, total_purchases, total_payment, total_balance)
    }


def receivable_case():
    return case(
        (
            (models.Transaction.type == "SALE") &
            ~(
                models.Transaction.source_ref.like("retail-bill:%") &
                models.Transaction.party_id.isnot(None) &
                exists().where(
                    and_(
                        models.RetailBill.party_id == models.Transaction.party_id,
                        models.RetailBill.outlet_id == models.Transaction.outlet_id,
                        models.RetailBill.date == models.Transaction.date,
                        cast(models.RetailBill.bill_number, String) == cast(models.Transaction.bill_number, String),
                        models.RetailBill.outstanding_amount <= 0
                    )
                )
            ),
            models.Transaction.amount
        ),
        (
            (models.Transaction.type == "OPENING") & (models.Transaction.category == "RECEIVABLE"),
            models.Transaction.amount
        ),
        (
            (models.Transaction.type == "PAYMENT") & (models.Transaction.category == "RECEIVED"),
            -models.Transaction.amount
        ),
        else_=0
    )


def payable_case():
    return case(
        (models.Transaction.type == "PURCHASE", models.Transaction.amount),
        (
            (models.Transaction.type == "OPENING") & (models.Transaction.category == "PAYABLE"),
            models.Transaction.amount
        ),
        (
            (models.Transaction.type == "PAYMENT") & (models.Transaction.category == "PAID"),
            -models.Transaction.amount
        ),
        else_=0
    )


def row_error(errors, row_number, message):
    errors.append({"row": row_number, "error": message})


def upload_result(inserted, skipped, errors, extra=None):
    result = {
        "status": "success",
        "rows_inserted": inserted,
        "rows_skipped": skipped,
        "errors": errors[:25],
        "preview": {
            "inserted": inserted,
            "skipped": skipped,
            "errors": len(errors)
        }
    }

    if extra:
        result.update(extra)

    return result


def report_response(rows, columns, filename, file_format, title, meta_rows=None):
    if file_format == "json":
        return {
            "title": title,
            "filename": filename,
            "columns": columns,
            "rows": rows,
            "meta_rows": meta_rows or []
        }

    if file_format == "excel":
        output = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"

        header_fill = PatternFill("solid", fgColor="EDE7DB")
        title_fill = PatternFill("solid", fgColor="F7F2E8")
        meta_fill = PatternFill("solid", fgColor="FBF8F1")
        summary_fill = PatternFill("solid", fgColor="F3E2BE")
        alt_fill = PatternFill("solid", fgColor="FCFAF5")
        thin_border = Border(
            left=Side(style="thin", color="D6C7AE"),
            right=Side(style="thin", color="D6C7AE"),
            top=Side(style="thin", color="D6C7AE"),
            bottom=Side(style="thin", color="D6C7AE"),
        )
        summary_border = Border(
            left=Side(style="thin", color="C9A86A"),
            right=Side(style="thin", color="C9A86A"),
            top=Side(style="medium", color="C9A86A"),
            bottom=Side(style="medium", color="C9A86A"),
        )

        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 24
        if columns:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

        current_row = 2
        for meta in meta_rows or []:
            meta_cell = sheet.cell(row=current_row, column=1, value=meta)
            meta_cell.font = Font(bold=True)
            meta_cell.fill = meta_fill
            meta_cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[current_row].height = 20
            if columns:
                sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            current_row += 1

        if meta_rows:
            current_row += 1

        header_row = current_row
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=header_row, column=index, value=column)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        sheet.row_dimensions[header_row].height = 22

        for row_index, row in enumerate(rows, start=header_row + 1):
            row_label = str(row.get("Type", "") or row.get(columns[0], "")).strip().lower() if columns else ""
            is_summary_row = row_label in ["closing balance", "total"]
            is_alt_row = (row_index - header_row) % 2 == 0
            for col_index, column in enumerate(columns, start=1):
                raw_value = row.get(column, "")
                cell = sheet.cell(row=row_index, column=col_index, value=raw_value)
                cell.border = summary_border if is_summary_row else thin_border
                if is_summary_row:
                    cell.font = Font(bold=True)
                    cell.fill = summary_fill
                elif is_alt_row:
                    cell.fill = alt_fill
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(raw_value, (int, float, Decimal)) else "left",
                    vertical="center"
                )
                if isinstance(raw_value, (int, float, Decimal)):
                    if column in ["Amount", "Balance", "Rate", "Sales", "Purchase", "Profit", "Payment Received", "Payment Paid", "Opening", "Receivable", "Payable", "Net Outstanding", "Old Bal", "Purchases", "Payment", "Total", "Bill Amount", "Paid", "Outstanding", "Debit", "Credit", "Account Balance", "Net"]:
                        cell.number_format = '₹#,##0.00'
                    elif column in ["KGS", "Kg", "Opening Kg", "Purchase Kg", "Sales Kg", "Expected Kg", "Actual Kg", "Leakage Kg", "Weight", "Live Cut Kg", "Mortality Kg"]:
                        cell.number_format = '#,##0.000'
                    elif column in ["NAG", "Nag"] or column.endswith(" NAG"):
                        cell.number_format = '#,##0'
                    if column in ["Amount", "Balance"] and float(raw_value or 0) < 0:
                        cell.font = Font(bold=is_summary_row, color="A94442")
                    elif is_summary_row:
                        cell.font = Font(bold=True)

            sheet.row_dimensions[row_index].height = 20

        center_columns = {"Date", "Type", "Bill No", "Mode", "Category"}
        for index, column in enumerate(columns, start=1):
            if column in center_columns:
                for row_index in range(header_row + 1, sheet.max_row + 1):
                    sheet.cell(row=row_index, column=index).alignment = Alignment(horizontal="center", vertical="center")

        for index, column in enumerate(columns, start=1):
            values = [str(column)]
            for row in rows:
                values.append(str(row.get(column, "")))
            width_cap = 24 if column in ["Category", "Item"] else 16
            if column in ["Date", "Type", "Bill No", "Mode"]:
                width_cap = 14
            if column in ["Amount", "Balance"]:
                width_cap = 16
            width = min(max(len(value) for value in values) + 2, width_cap)
            sheet.column_dimensions[get_column_letter(index)].width = max(width, 11)

        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.sheet_view.showGridLines = False
        if columns:
            sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{sheet.max_row}"
        workbook.save(output)
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )

    if file_format == "pdf":
        return Response(
            content=build_simple_pdf(title, columns, rows, meta_rows=meta_rows),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )

    return {"error": "Invalid format"}


def format_export_date(value):
    if not value:
        return ""
    try:
        return pd.to_datetime(value).strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def ledger_row_details(row):
    details = [row.get("category"), row.get("item")]
    if Decimal(str(row.get("quantity") or 0)):
        details.append(f"{row['quantity']:g} NAG")
    if Decimal(str(row.get("weight") or 0)):
        details.append(f"{row['weight']:g} kg")
    if Decimal(str(row.get("rate") or 0)):
        details.append(f"Rs {row['rate']:g}/kg")
    if row.get("payment_mode") and row["payment_mode"] != "NA":
        details.append(row["payment_mode"])
    return " | ".join(str(value) for value in details if value)


def pdf_format_value(column, value):
    if value in [None, ""]:
        return ""
    if isinstance(value, Decimal):
        value = float(value)
    if isinstance(value, (int, float)):
        if column in ["Amount", "Balance", "Rate", "Sales", "Purchase", "Profit", "Payment Received", "Payment Paid", "Opening", "Receivable", "Payable", "Net Outstanding", "Old Bal", "Purchases", "Payment", "Total", "Bill Amount", "Paid", "Outstanding", "Debit", "Credit", "Account Balance", "Net"]:
            return f"Rs {value:,.2f}"
        if column in ["KGS", "Kg", "Opening Kg", "Purchase Kg", "Sales Kg", "Expected Kg", "Actual Kg", "Leakage Kg", "Weight", "Live Cut Kg", "Mortality Kg"]:
            return f"{value:,.3f}"
        if column in ["NAG", "Nag"] or column.endswith(" NAG"):
            return f"{value:,.0f}"
    return str(value)


def pdf_fit_text(value, width, font_size):
    text = str(value or "")
    approx_char_width = max(font_size * 0.52, 1)
    max_chars = max(int(width / approx_char_width), 1)
    if len(text) <= max_chars:
        return text
    if max_chars <= 3:
        return text[:max_chars]
    return text[:max_chars - 3] + "..."


def pdf_column_widths(columns, usable_width):
    weights = {
        "Date": 1.2,
        "Type": 1.3,
        "Bill No": 0.9,
        "Category": 1.5,
        "Item": 1.2,
        "NAG": 0.7,
        "KGS": 0.9,
        "Rate": 0.9,
        "Mode": 0.9,
        "Amount": 1.1,
        "Balance": 1.1,
    }
    raw = [weights.get(column, 1.0) for column in columns]
    total = sum(raw) or 1
    return [(weight / total) * usable_width for weight in raw]


def pdf_text_line(x, y, text, font="F1", size=9, color=(0.18, 0.14, 0.10)):
    return f"BT /{font} {size} Tf {color[0]:.3f} {color[1]:.3f} {color[2]:.3f} rg 1 0 0 1 {x:.2f} {y:.2f} Tm ({pdf_escape(text)}) Tj ET"


def pdf_rect(x, y, width, height, fill=None, stroke=None, line_width=0.6):
    ops = []
    if fill:
        ops.append(f"{fill[0]:.3f} {fill[1]:.3f} {fill[2]:.3f} rg")
    if stroke:
        ops.append(f"{stroke[0]:.3f} {stroke[1]:.3f} {stroke[2]:.3f} RG")
        ops.append(f"{line_width:.2f} w")
    paint = "B" if fill and stroke else ("f" if fill else "S")
    ops.append(f"{x:.2f} {y:.2f} {width:.2f} {height:.2f} re {paint}")
    return "\n".join(ops)


def build_simple_pdf(title, columns, rows, meta_rows=None):
    landscape = len(columns) > 8
    page_width, page_height = (842, 595) if landscape else (595, 842)
    margin_x = 24
    margin_top = 26
    margin_bottom = 24
    usable_width = page_width - (margin_x * 2)
    row_height = 18
    header_height = 20
    title_gap = 26
    meta_gap = 18
    font_size = 7 if landscape else 8
    col_widths = pdf_column_widths(columns, usable_width)
    number_columns = {"Amount", "Balance", "Rate", "Sales", "Purchase", "Profit", "Payment Received", "Payment Paid", "Opening", "Receivable", "Payable", "Net Outstanding", "Old Bal", "Purchases", "Payment", "Total", "KGS", "Kg", "Opening Kg", "Purchase Kg", "Sales Kg", "Expected Kg", "Actual Kg", "Leakage Kg", "Weight", "NAG", "Nag"}
    number_columns.update({"Bill Amount", "Paid", "Outstanding", "Live Cut Kg", "Mortality Kg", "Opening NAG", "Actual NAG"})

    objects = {
        1: "<< /Type /Catalog /Pages 2 0 R >>",
        3: "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        4: "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    }
    page_refs = []
    next_ref = 5

    def draw_header(commands, y):
        x = margin_x
        for index, column in enumerate(columns):
            width = col_widths[index]
            commands.append(pdf_rect(x, y - header_height, width, header_height, fill=(0.929, 0.906, 0.859), stroke=(0.839, 0.780, 0.682)))
            commands.append(pdf_text_line(x + 4, y - 13, pdf_fit_text(column, width - 8, 8), font="F2", size=8, color=(0.29, 0.23, 0.15)))
            x += width
        return y - header_height

    def new_page(first_page=False):
        commands = []
        y = page_height - margin_top
        commands.append(pdf_rect(margin_x, y - 18, usable_width, 18, fill=(0.969, 0.949, 0.910), stroke=(0.839, 0.780, 0.682), line_width=0.8))
        commands.append(pdf_text_line(margin_x + 6, y - 13, title, font="F2", size=12, color=(0.20, 0.15, 0.10)))
        y -= title_gap

        if first_page:
            for meta in meta_rows or []:
                commands.append(pdf_rect(margin_x, y - 14, usable_width, 14, fill=(0.984, 0.973, 0.945), stroke=None))
                commands.append(pdf_text_line(margin_x + 4, y - 10, pdf_fit_text(meta, usable_width - 8, 8), font="F2", size=8, color=(0.28, 0.22, 0.14)))
                y -= meta_gap
            if meta_rows:
                y -= 6

        y = draw_header(commands, y)
        return commands, y

    if not rows:
        rows = [{column: ("No records found" if index == 0 else "") for index, column in enumerate(columns)}]

    page_index = 0
    commands, current_y = new_page(first_page=True)
    page_index += 1

    for row in rows:
        if current_y - row_height < margin_bottom:
            page_ref = next_ref + 1
            stream = "\n".join(commands)
            content_ref = next_ref
            objects[content_ref] = f"<< /Length {len(stream.encode('latin-1', errors='replace'))} >>\nstream\n{stream}\nendstream"
            next_ref += 1
            objects[page_ref] = (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
                f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_ref} 0 R >>"
            )
            next_ref += 1
            page_refs.append(page_ref)
            commands, current_y = new_page(first_page=False)
            page_index += 1

        row_label = str(row.get("Type", "") or row.get(columns[0], "")).strip().lower() if columns else ""
        is_summary_row = row_label in ["closing balance", "total"]
        is_alt_row = (len(page_refs) + len(commands)) % 2 == 0
        fill = (0.953, 0.886, 0.745) if is_summary_row else ((0.988, 0.980, 0.961) if is_alt_row else None)
        stroke = (0.788, 0.659, 0.416) if is_summary_row else (0.839, 0.780, 0.682)
        x = margin_x
        for index, column in enumerate(columns):
            width = col_widths[index]
            commands.append(pdf_rect(x, current_y - row_height, width, row_height, fill=fill, stroke=stroke, line_width=0.8 if is_summary_row else 0.5))
            display = pdf_format_value(column, row.get(column, ""))
            fitted = pdf_fit_text(display, width - 8, font_size)
            if column in number_columns:
                approx_width = len(fitted) * font_size * 0.50
                text_x = max(x + 4, x + width - approx_width - 4)
            else:
                text_x = x + 4
            color = (0.66, 0.27, 0.26) if column in ["Amount", "Balance"] and str(display).startswith("Rs -") else (0.20, 0.15, 0.10)
            commands.append(pdf_text_line(text_x, current_y - 12, fitted, font="F2" if is_summary_row else "F1", size=font_size, color=color))
            x += width
        current_y -= row_height

    stream = "\n".join(commands)
    content_ref = next_ref
    objects[content_ref] = f"<< /Length {len(stream.encode('latin-1', errors='replace'))} >>\nstream\n{stream}\nendstream"
    next_ref += 1
    page_ref = next_ref
    objects[page_ref] = (
        f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
        f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_ref} 0 R >>"
    )
    next_ref += 1
    page_refs.append(page_ref)

    objects[2] = (
        "<< /Type /Pages /Kids ["
        + " ".join(f"{ref} 0 R" for ref in page_refs)
        + f"] /Count {len(page_refs)} >>"
    )

    pdf = "%PDF-1.4\n"
    offsets = [0]
    for index in sorted(objects):
        value = objects[index]
        offsets.append(len(pdf.encode("latin-1")))
        pdf += f"{index} 0 obj\n{value}\nendobj\n"

    xref_offset = len(pdf.encode("latin-1"))
    max_ref = max(objects)
    pdf += f"xref\n0 {max_ref + 1}\n0000000000 65535 f \n"
    for offset in offsets[1:]:
        pdf += f"{offset:010d} 00000 n \n"
    pdf += f"trailer\n<< /Size {max_ref + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
    return pdf.encode("latin-1", errors="replace")


def pdf_escape(value):
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def parse_report_dates(start_date, end_date):
    start = parse_input_date(start_date) if start_date else None
    end = parse_input_date(end_date) if end_date else None
    return start, end


def safe_filename(value):
    cleaned = "".join(char if char.isalnum() or char in ["-", "_"] else "_" for char in value)
    return quote(cleaned[:120] or "report")


def append_daily_sheet_goods_rows(export_rows, section_name, section):
    if not section:
        return
    for row in section.get("rows", []) or []:
        export_rows.append({
            "Section": section_name,
            "Label": "Row",
            "Name": row.get("goods", ""),
            "Nag": row.get("nag", ""),
            "Weight": row.get("weight", ""),
            "Rate": row.get("rate", ""),
            "Total": row.get("total", ""),
            "Mode": "",
            "Paid": "",
            "Outstanding": ""
        })
    total = section.get("total")
    if total:
        export_rows.append({
            "Section": section_name,
            "Label": "Total",
            "Name": total.get("goods", "TOTAL"),
            "Nag": total.get("nag", ""),
            "Weight": total.get("weight", ""),
            "Rate": total.get("rate", ""),
            "Total": total.get("total", ""),
            "Mode": "",
            "Paid": "",
            "Outstanding": ""
        })


def build_daily_sheet_export_report(sheet_payload, sheet_type, target_date):
    if sheet_type in ["vendor", "dealer"]:
        business_label = "Sales" if sheet_type == "vendor" else "Purchases"
        rows = [{
            "Party Name": row.get("party_name", ""), "Old Bal": row.get("old_balance", 0),
            business_label: row.get("purchases", 0), "Payment": row.get("payment", 0),
            "Balance": row.get("balance", 0),
        } for row in [*(sheet_payload.get("rows") or []), sheet_payload.get("totals") or {}]]
        return report_response(
            rows, ["Party Name", "Old Bal", business_label, "Payment", "Balance"],
            safe_filename(f"{sheet_type}_balance_sheet_{target_date}"), "excel",
            sheet_payload.get("title") or "Balance Sheet",
            meta_rows=[f"Date: {format_export_date(target_date)}"],
        )

    rows = []
    def stock_rows(section, data):
        for row in [*(data.get("rows") or []), *([data["total"]] if data.get("total") else [])]:
            rows.append({
                "Section": section, "Goods": row.get("goods", ""), "NAG": row.get("nag", ""),
                "Weight": row.get("weight"), "Rate": row.get("rate"), "Total": row.get("total"),
            })

    for name, key in [("Opening Live Stock", "opening_stock"), ("Purchases", "purchase_stock"),
                      ("Transport Mortality", "transport_mortality_stock"), ("Shop Mortality", "shop_mortality_stock")]:
        stock_rows(name, sheet_payload.get(key) or {})
    for section in sheet_payload.get("sales_sections") or []:
        stock_rows(section.get("title", "Sales"), section)
    final = sheet_payload.get("final_stock") or {}
    stock_rows("Final Live Stock", {"rows": [
        final[key] for key in ("total_purchases", "transport_mortality", "shop_mortality",
                              "sales", "live_cut", "closing_stock", "actual_stock", "short_by") if key in final
    ]})
    for row in (sheet_payload.get("retail_credit_sheet") or {}).get("rows", []):
        rows.append({
            "Section": "Retail Credit", "Party": row.get("customer_name", ""),
            "Bill No": row.get("bill_number", ""), "Bill Amount": row.get("total_amount"),
            "Paid": row.get("paid_amount"), "Outstanding": row.get("outstanding_amount"),
            "Mode": row.get("payment_mode", ""),
        })
    credit_total = (sheet_payload.get("retail_credit_sheet") or {}).get("total")
    if credit_total:
        rows.append({"Section": "Retail Credit", "Party": "TOTAL",
                     "Bill Amount": credit_total.get("total_amount"), "Paid": credit_total.get("paid_amount"),
                     "Outstanding": credit_total.get("outstanding_amount")})
    columns = ["Section", "Goods", "NAG", "Weight", "Rate", "Total", "Party", "Bill No",
               "Bill Amount", "Paid", "Outstanding", "Mode"]
    meta = [f"Date: {format_export_date(target_date)}", "Stock basis: Live birds"]
    if sheet_payload.get("stock_warning"):
        meta.append(sheet_payload["stock_warning"])
    return report_response(rows, columns, f"stock_sheet_{target_date}", "excel", "Stock Sheet", meta_rows=meta)


def latest_item_rates(db: Session, target_date, scope=None):
    query = db.query(
        models.Transaction.item_type,
        models.Transaction.rate,
        models.Transaction.date
    ).filter(
        models.Transaction.item_type.isnot(None),
        models.Transaction.rate.isnot(None),
        models.Transaction.date <= target_date,
        models.Transaction.type == "PURCHASE"
    )
    if scope:
        query = apply_outlet_scope(query, models.Transaction, scope)
    rows = query.order_by(
        models.Transaction.item_type.asc(),
        models.Transaction.date.desc(),
        models.Transaction.created_at.desc()
    ).all()

    rates = {}
    for row in rows:
        if row.item_type and row.item_type not in rates:
            rates[row.item_type] = Decimal(row.rate or 0)
    return rates


def stock_item_names_query(db: Session, scope=None):
    purchase_query = db.query(models.Transaction.item_type).filter(
        models.Transaction.item_type.isnot(None),
        models.Transaction.type == "PURCHASE"
    )
    if scope:
        purchase_query = apply_outlet_scope(purchase_query, models.Transaction, scope)
    purchase_items = purchase_query.distinct().all()

    opening_query = db.query(models.ItemOpeningStock.item_type)
    if scope:
        opening_query = apply_outlet_scope(opening_query, models.ItemOpeningStock, scope)
    opening_items = opening_query.distinct().all()

    processed_query = db.query(models.DailyItemStock.item_type)
    if scope:
        processed_query = apply_outlet_scope(processed_query, models.DailyItemStock, scope)
    processed_items = processed_query.distinct().all()

    items = set()
    for row in purchase_items:
        if row.item_type:
            items.add(row.item_type)
    for row in opening_items:
        if row.item_type:
            items.add(row.item_type)
    for row in processed_items:
        if row.item_type:
            items.add(row.item_type)
    return items


PROCESS_DAY_SOURCE_ITEMS = list(SOURCE_ITEMS)
RETAIL_SOURCE_TYPE_ALIASES = SOURCE_ALIASES


def normalize_stock_source_key(value):
    return " ".join(str(value or "").strip().upper().split())


def resolve_stock_source_type(item_name="", source_item_type=""):
    explicit = normalize_stock_source_key(source_item_type)
    if explicit in PROCESS_DAY_SOURCE_ITEMS:
        return explicit

    normalized_item = normalize_stock_source_key(item_name)
    mapped = RETAIL_SOURCE_TYPE_ALIASES.get(normalized_item)
    if mapped:
        return mapped

    return explicit or normalized_item or ""


def format_sheet_row(label, weight=0, rate=0, amount=0, nag=None):
    return {
        "goods": label,
        "nag": float(nag) if nag is not None else "",
        "weight": optional_float(weight),
        "rate": optional_float(rate),
        "total": optional_float(amount)
    }


def format_balance_row(party_name, old_balance=0, purchases=0, payment=0, balance=0):
    return {
        "party_name": party_name,
        "old_balance": float(old_balance or 0),
        "purchases": float(purchases or 0),
        "payment": float(payment or 0),
        "balance": float(balance or 0)
    }


def format_retail_credit_row(customer_name, bill_number, total_amount=0, paid_amount=0, outstanding_amount=0, payment_mode="Credit"):
    return {
        "customer_name": customer_name or "Walk-in Customer",
        "bill_number": bill_number or "",
        "total_amount": float(total_amount or 0),
        "paid_amount": float(paid_amount or 0),
        "outstanding_amount": float(outstanding_amount or 0),
        "payment_mode": payment_mode or "Credit"
    }


def decimal_ratio(amount, weight):
    amount_value = Decimal(amount or 0)
    weight_value = Decimal(weight or 0)
    if weight_value <= 0:
        return Decimal("0")
    return amount_value / weight_value


def optional_decimal_sum(values):
    found = False
    total = Decimal("0")
    for value in values:
        if value not in [None, ""]:
            found = True
            total += Decimal(value or 0)
    return total if found else None


def optional_float(value):
    if value in [None, ""]:
        return None
    return float(value)


def format_rate_analysis_row(label, avg_rate=0, weight=0, amount=0, category=None, goods=None):
    row = {
        "label": label,
        "avg_rate": float(avg_rate or 0),
        "weight": float(weight or 0),
        "amount": float(amount or 0)
    }
    if category is not None:
        row["category"] = category
    if goods is not None:
        row["goods"] = goods
    return row


def format_performance_row(item, purchase_kg=0, sales_kg=0, buy_rate=0, sell_rate=0, spread=0, gross_profit=0):
    return {
        "item": item,
        "purchase_kg": float(purchase_kg or 0),
        "sales_kg": float(sales_kg or 0),
        "buy_rate": float(buy_rate or 0),
        "sell_rate": float(sell_rate or 0),
        "spread": float(spread or 0),
        "gross_profit": float(gross_profit or 0)
    }


def parse_decimal(value, default="0"):
    if value in [None, ""]:
        return Decimal(default)

    try:
        return Decimal(str(value).strip())
    except Exception:
        return Decimal(default)


def serialize_dressed_stock_entry(entry):
    return {
        "id": str(entry.id),
        "date": str(entry.date),
        "item_name": entry.item_name,
        "live_quantity": float(entry.live_quantity or 0),
        "live_weight": float(entry.live_weight or 0),
        "dressed_weight": float(entry.dressed_weight or 0),
        "remaining_dressed_weight": float(entry.remaining_dressed_weight or 0),
        "default_rate": float(entry.default_rate or 0),
        "notes": entry.notes or ""
    }


def serialize_retail_bill(bill, items):
    created_at = bill.created_at or datetime.utcnow()
    has_dressed = any((item.line_type or "STANDARD").upper() == "DRESSED" for item in items)
    has_regular = any((item.line_type or "STANDARD").upper() != "DRESSED" for item in items)
    if has_regular and has_dressed:
        bill_mode = "both"
    elif has_dressed:
        bill_mode = "dressed"
    else:
        bill_mode = "regular"

    return {
        "id": str(bill.id),
        "outlet_id": str(bill.outlet_id) if getattr(bill, "outlet_id", None) else "",
        "bill_number": bill.bill_number,
        "bill_mode": bill_mode,
        "date": str(bill.date),
        "time": created_at.strftime("%H:%M:%S"),
        "customer_name": bill.customer_name or "",
        "customer_phone": bill.customer_phone or "",
        "customer_address": bill.customer_address or "",
        "cashier_name": bill.cashier_name or "admin",
        "payment_mode": bill.payment_mode or "Cash",
        "total_nag": float(bill.total_quantity or 0),
        "total_quantity": float(bill.total_quantity or 0),
        "total_weight": float(bill.total_weight or 0),
        "ice_amount": float(bill.ice_amount or 0),
        "total_amount": float(bill.total_amount or 0),
        "paid_amount": float(bill.paid_amount or 0),
        "outstanding_amount": float(bill.outstanding_amount or 0),
        "party_balance": None,
        "notes": bill.notes or "",
        "items": [
            {
                "line_order": item.line_order,
                "item_name": item.item_name,
                "line_type": (item.line_type or "STANDARD").upper(),
                "source_item_type": item.source_item_type or "",
                "nag": float(item.quantity or 0),
                "quantity": float(item.quantity or 0),
                "unit": item.unit or "KGS",
                "weight": float(item.weight or 0),
                "rate": float(item.rate or 0),
                "amount": float(item.amount or 0)
            }
            for item in items
        ]
    }


def retail_party_balance_after(db: Session, party_id, outlet_id=None):
    if not party_id:
        return None

    txns_query = db.query(models.Transaction).filter(
        models.Transaction.party_id == party_id
    )
    if outlet_id:
        txns_query = txns_query.filter(models.Transaction.outlet_id == outlet_id)
    txns = txns_query.order_by(models.Transaction.date, models.Transaction.created_at, models.Transaction.id).all()
    balances, _ = build_ledger(db, txns)
    return float(balances["balance"])


def filter_party_ledger_transactions(db: Session, txns):
    return txns


def serialize_payment_receipt(receipt, balance_after=None):
    created_at = receipt.created_at or datetime.utcnow()
    return {
        "id": str(receipt.id),
        "outlet_id": str(receipt.outlet_id) if getattr(receipt, "outlet_id", None) else "",
        "receipt_number": receipt.receipt_number,
        "date": str(receipt.date),
        "time": created_at.strftime("%H:%M:%S"),
        "party_name": receipt.party_name or "",
        "party_phone": receipt.party_phone or "",
        "party_address": receipt.party_address or "",
        "cashier_name": receipt.cashier_name or "admin",
        "direction": receipt.direction or "RECEIVED",
        "payment_mode": receipt.payment_mode or "Cash",
        "amount": float(receipt.amount or 0),
        "balance_after": float(balance_after or 0),
        "notes": receipt.notes or ""
    }


def recompute_dressed_stock_remaining(db: Session, target_date, outlet_id=None):
    if not target_date:
        return

    entries_query = db.query(models.DressedStockEntry).filter(
        models.DressedStockEntry.date == target_date
    )
    if outlet_id:
        entries_query = entries_query.filter(models.DressedStockEntry.outlet_id == outlet_id)

    entries = entries_query.order_by(
        models.DressedStockEntry.created_at.asc()
    ).all()

    if not entries:
        return

    for entry in entries:
        base_weight = Decimal(entry.dressed_weight or 0)
        entry.remaining_dressed_weight = base_weight if base_weight > 0 else Decimal("0")

    total_dressed_sold_query = db.query(
        func.coalesce(func.sum(models.RetailBillItem.weight), 0)
    ).join(
        models.RetailBill,
        models.RetailBill.id == models.RetailBillItem.bill_id
    ).filter(
        models.RetailBill.date == target_date,
        models.RetailBillItem.line_type == "DRESSED"
    )
    if outlet_id:
        total_dressed_sold_query = total_dressed_sold_query.filter(models.RetailBill.outlet_id == outlet_id)
    total_dressed_sold = total_dressed_sold_query.scalar()

    remaining_required = Decimal(total_dressed_sold or 0)
    for entry in entries:
        if remaining_required <= 0:
            break
        available_weight = Decimal(entry.remaining_dressed_weight or 0)
        used_weight = min(available_weight, remaining_required)
        entry.remaining_dressed_weight = available_weight - used_weight
        remaining_required -= used_weight


def merge_party_type(existing_type: str | None, new_type: str | None):
    existing = str(existing_type or "").strip().upper()
    incoming = str(new_type or "").strip().upper()

    if not existing:
        return incoming or None
    if not incoming or incoming == existing:
        return existing
    if "BOTH" in [existing, incoming]:
        return "BOTH"
    return "BOTH"


def update_party_details(
    party,
    party_type: str | None = None,
    phone: str | None = None,
    address: str | None = None,
    overwrite_contact: bool = False
):
    if party is None:
        return

    merged_type = merge_party_type(party.type, party_type)
    if merged_type and merged_type != party.type:
        party.type = merged_type

    cleaned_phone = str(phone or "").strip()
    cleaned_address = str(address or "").strip()
    if cleaned_phone and (overwrite_contact or not (party.phone or "").strip()):
        party.phone = cleaned_phone
    if cleaned_address and (overwrite_contact or not (party.address or "").strip()):
        party.address = cleaned_address


def get_or_create_party(
    db: Session,
    party_name: str,
    party_type: str,
    seen_aliases: dict,
    phone: str | None = None,
    address: str | None = None
):
    normalized = normalize_party_name(party_name)

    if normalized in seen_aliases:
        party_id = seen_aliases[normalized]
        party = db.query(models.Party).filter_by(id=party_id).first()
        update_party_details(party, party_type, phone, address)
        return party_id

    alias = db.query(models.PartyAlias).filter_by(
        normalized_alias=normalized
    ).first()

    if alias:
        party_id = alias.party_id
        party = db.query(models.Party).filter_by(id=party_id).first()
        update_party_details(party, party_type, phone, address)
    else:
        party = models.Party(
            name=party_name,
            normalized_name=normalized,
            type=party_type,
            phone=str(phone or "").strip() or None,
            address=str(address or "").strip() or None
        )
        db.add(party)
        db.flush()

        alias = models.PartyAlias(
            alias=party_name,
            normalized_alias=normalized,
            party_id=party.id
        )
        db.add(alias)

        party_id = party.id

    seen_aliases[normalized] = party_id
    return party_id


def get_party_by_name(db: Session, party_name: str):
    normalized = normalize_party_name(party_name)
    if not normalized:
        return None

    alias = db.query(models.PartyAlias).filter(
        models.PartyAlias.normalized_alias == normalized
    ).first()
    if alias:
        return db.query(models.Party).filter_by(id=alias.party_id).first()

    alias = db.query(models.PartyAlias).filter(
        models.PartyAlias.normalized_alias.contains(normalized)
    ).first()
    if not alias:
        return None

    return db.query(models.Party).filter_by(id=alias.party_id).first()


def normalize_payment_direction(value: str):
    normalized = str(value).strip().upper()

    if normalized in ["RECEIVED", "RECEIVE", "FROM", "IN", "INCOMING", "CREDIT"]:
        return "RECEIVED"

    if normalized in ["PAID", "PAY", "TO", "OUT", "OUTGOING", "DEBIT"]:
        return "PAID"

    return None


def get_first_existing_column(df, candidates):
    for col in candidates:
        if col in df.columns:
            return col

    return None


def require_column(df, candidates, label):
    col = get_first_existing_column(df, candidates)
    if not col:
        return None, {"error": f"Missing column: {label}"}

    return col, None


def get_optional_row_value(row, candidates):
    for col in candidates:
        if col in row.index and not pd.isna(row[col]):
            value = str(row[col]).strip()
            if value:
                return value

    return None


def resolve_upload_date(row, date_col, fallback_date):
    if date_col and date_col in row.index and not pd.isna(row[date_col]):
        return pd.to_datetime(row[date_col], dayfirst=True).date()

    return fallback_date


@app.post("/entries/vendor")
def create_vendor_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one vendor row"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_transactions = set()

    for index, row in enumerate(rows, start=1):
        try:
            party_name = str(row.get("vendor") or row.get("party") or "").strip()
            item_type = normalize_source(row.get("hen_type") or row.get("item_type"))
            category = str(row.get("category") or "").strip() or None
            quantity = parse_decimal(row.get("nag", row.get("quantity")))
            weight = parse_decimal(row.get("kgs", row.get("weight")))
            rate = parse_decimal(row.get("rate_per_kg", row.get("rate")))

            if not party_name or not item_type or weight <= 0 or rate <= 0 or quantity < 0:
                skipped += 1
                row_error(errors, index, "Enter vendor, hen type, valid NAG, kg, and rate")
                continue

            party_id = get_or_create_party(db, party_name, "VENDOR", seen_aliases)
            txn_key = (target_date, party_id, float(weight), float(rate), "SALE", category, item_type)
            existing_txn = db.query(models.Transaction).filter_by(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                weight=float(weight),
                rate=float(rate),
                type="SALE",
                category=category,
                item_type=item_type
            ).first()

            if existing_txn or txn_key in seen_transactions:
                skipped += 1
                continue

            db.add(models.Transaction(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                type="SALE",
                category=category,
                item_type=item_type,
                quantity=quantity if quantity > 0 else None,
                weight=weight,
                rate=rate,
                amount=weight * rate,
                payment_mode="NA"
            ))
            seen_transactions.add(txn_key)
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving vendor entries failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/entries/dealer")
def create_dealer_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one dealer row"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_transactions = set()

    for index, row in enumerate(rows, start=1):
        try:
            party_name = str(row.get("dealer") or row.get("party") or "").strip()
            bill_number = str(row.get("bill_no") or row.get("bill_number") or "").strip() or None
            item_type = normalize_source(row.get("hen_type") or row.get("item_type"))
            quantity = parse_decimal(row.get("nag", row.get("quantity")))
            weight = parse_decimal(row.get("kgs", row.get("weight")))
            rate = parse_decimal(row.get("rate_per_kg", row.get("rate")))
            transport_mortality_nag = parse_decimal(row.get("transport_mortality_nag"))
            transport_mortality_weight = parse_decimal(row.get("transport_mortality_weight"))

            if not party_name or not item_type or weight <= 0 or rate <= 0 or quantity < 0:
                skipped += 1
                row_error(errors, index, "Enter dealer, hen type, valid NAG, kg, and rate")
                continue

            if transport_mortality_nag < 0 or transport_mortality_weight < 0:
                skipped += 1
                row_error(errors, index, "Transport mortality NAG and kg cannot be negative")
                continue

            party_id = get_or_create_party(db, party_name, "DEALER", seen_aliases)
            txn_key = (target_date, party_id, float(weight), float(rate), "PURCHASE", bill_number or "", item_type)
            existing_txn = db.query(models.Transaction).filter_by(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                weight=float(weight),
                rate=float(rate),
                type="PURCHASE",
                category=None,
                item_type=item_type,
                bill_number=bill_number
            ).first()

            if existing_txn or txn_key in seen_transactions:
                skipped += 1
                continue

            db.add(models.Transaction(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                type="PURCHASE",
                category=None,
                item_type=item_type,
                quantity=quantity if quantity > 0 else None,
                weight=weight,
                rate=rate,
                amount=weight * rate,
                payment_mode="NA",
                bill_number=bill_number
            ))
            seen_transactions.add(txn_key)

            if transport_mortality_nag > 0 or transport_mortality_weight > 0:
                mortality_key = (
                    target_date,
                    party_id,
                    float(transport_mortality_nag),
                    float(transport_mortality_weight),
                    "TRANSPORTATION MORTALITY",
                    bill_number or "",
                    item_type
                )
                existing_mortality = db.query(models.Transaction).filter_by(
                    date=target_date,
                    outlet_id=current_outlet.id,
                    party_id=party_id,
                    type="MORTALITY",
                    category="TRANSPORTATION MORTALITY",
                    item_type=item_type,
                    quantity=transport_mortality_nag if transport_mortality_nag > 0 else None,
                    weight=transport_mortality_weight if transport_mortality_weight > 0 else 0,
                    bill_number=bill_number
                ).first()

                if not existing_mortality and mortality_key not in seen_transactions:
                    db.add(models.Transaction(
                        date=target_date,
                        outlet_id=current_outlet.id,
                        party_id=party_id,
                        type="MORTALITY",
                        category="TRANSPORTATION MORTALITY",
                        item_type=item_type,
                        quantity=transport_mortality_nag if transport_mortality_nag > 0 else None,
                        weight=transport_mortality_weight if transport_mortality_weight > 0 else 0,
                        rate=0,
                        amount=0,
                        payment_mode="NA",
                        bill_number=bill_number
                    ))
                    seen_transactions.add(mortality_key)
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving dealer entries failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/entries/payment")
def create_payment_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one payment row"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_payments = set()

    for index, row in enumerate(rows, start=1):
        try:
            party_name = str(row.get("party") or "").strip()
            amount = parse_decimal(row.get("amount"))
            payment_mode = str(row.get("payment_mode") or "NA").strip()
            direction = normalize_payment_direction(row.get("direction"))

            if not party_name or amount <= 0 or not direction:
                skipped += 1
                row_error(errors, index, "Enter party, amount, payment mode, and valid direction")
                continue

            party_id = get_or_create_party(db, party_name, "BOTH", seen_aliases)
            payment_key = (target_date, party_id, float(amount), payment_mode, direction)
            existing_payment = db.query(models.Transaction).filter_by(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                type="PAYMENT",
                amount=amount,
                payment_mode=payment_mode,
                category=direction
            ).first()

            if existing_payment or payment_key in seen_payments:
                skipped += 1
                continue

            db.add(models.Transaction(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                type="PAYMENT",
                category=direction,
                amount=amount,
                payment_mode=payment_mode
            ))
            seen_payments.add(payment_key)
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving payment entries failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/entries/mortality")
def create_mortality_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one mortality row"}

    inserted = 0
    skipped = 0
    errors = []
    seen_transactions = set()

    for index, row in enumerate(rows, start=1):
        try:
            item_type = normalize_source(row.get("hen_type") or row.get("item_type"))
            quantity = parse_decimal(row.get("nag", row.get("quantity")))
            weight = parse_decimal(row.get("weight", row.get("kgs")))

            if not item_type:
                skipped += 1
                row_error(errors, index, "Enter hen type")
                continue

            if quantity < 0 or weight < 0:
                skipped += 1
                row_error(errors, index, "NAG and weight cannot be negative")
                continue

            if quantity <= 0 and weight <= 0:
                skipped += 1
                row_error(errors, index, "Enter NAG or weight")
                continue

            txn_key = (target_date, item_type, float(quantity), float(weight), "SHOP MORTALITY")
            existing_txn = db.query(models.Transaction).filter_by(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=None,
                type="MORTALITY",
                category="SHOP MORTALITY",
                item_type=item_type,
                quantity=quantity if quantity > 0 else None,
                weight=weight if weight > 0 else 0
            ).first()

            if existing_txn or txn_key in seen_transactions:
                skipped += 1
                continue

            db.add(models.Transaction(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=None,
                type="MORTALITY",
                category="SHOP MORTALITY",
                item_type=item_type,
                quantity=quantity if quantity > 0 else None,
                weight=weight if weight > 0 else 0,
                rate=0,
                amount=0,
                payment_mode="NA"
            ))
            seen_transactions.add(txn_key)
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving mortality entries failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/entries/opening-balance")
def create_opening_balance_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one opening balance row"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_opening = set()

    for index, row in enumerate(rows, start=1):
        try:
            party_name = str(row.get("party") or "").strip()
            amount = parse_decimal(row.get("opening_balance", row.get("amount")))
            balance_type = str(row.get("balance_type") or "").strip().upper()

            if balance_type in ["RECEIVABLE", "RECEIVE", "CUSTOMER", "VENDOR"]:
                balance_type = "RECEIVABLE"
                party_type = "VENDOR"
            elif balance_type in ["PAYABLE", "PAY", "SUPPLIER", "DEALER"]:
                balance_type = "PAYABLE"
                party_type = "DEALER"
            else:
                skipped += 1
                row_error(errors, index, "Balance type must be RECEIVABLE or PAYABLE")
                continue

            if not party_name or amount < 0:
                skipped += 1
                row_error(errors, index, "Enter party and valid opening balance")
                continue

            party_id = get_or_create_party(db, party_name, party_type, seen_aliases)
            opening_key = (target_date, party_id, balance_type)
            existing_opening = db.query(models.Transaction).filter_by(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                type="OPENING",
                category=balance_type
            ).first()
            if existing_opening or opening_key in seen_opening:
                skipped += 1
                continue

            db.add(models.Transaction(
                date=target_date,
                outlet_id=current_outlet.id,
                party_id=party_id,
                type="OPENING",
                category=balance_type,
                amount=amount,
                payment_mode="NA"
            ))
            seen_opening.add(opening_key)
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving opening balances failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/entries/opening-stock")
def create_opening_stock_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one opening stock row"}

    inserted = 0
    skipped = 0
    errors = []
    seen_stock = set()

    for index, row in enumerate(rows, start=1):
        try:
            item_type = normalize_source(row.get("hen_type") or row.get("item_type"))
            opening_quantity = parse_decimal(row.get("opening_nag", row.get("nag", row.get("quantity"))))
            opening_weight = parse_decimal(row.get("opening_kgs", row.get("kgs", row.get("weight"))))

            if not item_type or opening_weight < 0 or opening_quantity < 0:
                skipped += 1
                row_error(errors, index, "Enter hen type, valid opening NAG, and opening kg")
                continue

            stock_key = (target_date, item_type)
            existing_stock = db.query(models.ItemOpeningStock).filter_by(
                date=target_date,
                outlet_id=current_outlet.id,
                item_type=item_type
            ).first()
            if existing_stock or stock_key in seen_stock:
                skipped += 1
                continue

            db.add(models.ItemOpeningStock(
                date=target_date,
                outlet_id=current_outlet.id,
                item_type=item_type,
                opening_quantity=opening_quantity if opening_quantity > 0 else None,
                opening_weight=opening_weight
            ))
            seen_stock.add(stock_key)
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving opening stock failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/upload/vendor")
def upload_vendor(file: UploadFile = File(...), preview: bool = False, input_date: str = None, db: Session = Depends(get_db)):

    import io
    import hashlib

    filename = (file.filename or "").lower()
    contents = file.file.read()

    # --- File hash (duplicate file protection) ---
    file_hash = hashlib.sha256(contents).hexdigest()

    existing_file = db.query(models.UploadedFile).filter_by(file_hash=file_hash).first()
    if existing_file:
        return {"error": "File already uploaded"}

    # --- Read file ---
    try:
        if filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")

        elif filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")

        elif filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8")

        else:
            return {"error": "Unsupported file format"}

    except Exception as e:
        return {"error": f"File read failed: {str(e)}"}

    # --- Validation ---
    if df.empty:
        return {"error": "File is empty"}

    df.columns = df.columns.str.strip().str.upper()

    date_col = get_first_existing_column(df, ["DATE"])
    fallback_date = parse_input_date(input_date) if input_date else None
    if not date_col and not fallback_date:
        return {"error": "Provide DATE column in file or select the upload date in the app"}

    party_col, error = require_column(df, ["VENDOR", "PARTY", "NAME"], "VENDOR")
    if error:
        return error

    weight_col, error = require_column(df, ["KGS", "KG", "WEIGHT"], "KGS")
    if error:
        return error

    rate_col, error = require_column(df, ["RATE_PER_KG", "RATE PER KG", "RATE/KG", "RATE"], "RATE PER KG")
    if error:
        return error

    item_col, error = require_column(df, ["HEN_TYPE", "HEN TYPE", "ITEM", "TYPE"], "HEN TYPE")
    if error:
        return error

    quantity_col = get_first_existing_column(df, ["NAG", "QTY", "QUANTITY", "PCS", "PIECES"])

    # --- Numeric validation ---
    try:
        df[weight_col] = df[weight_col].astype(float)
        df[rate_col] = df[rate_col].astype(float)
        if quantity_col:
            df[quantity_col] = df[quantity_col].astype(float)
    except:
        return {"error": "Invalid numeric values"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_transactions = set()

    # --- Process rows ---
    for index, row in df.iterrows():
        row_number = int(index) + 2
        try:
            if pd.isna(row[party_col]) or pd.isna(row[weight_col]) or pd.isna(row[rate_col]) or pd.isna(row[item_col]):
                skipped += 1
                row_error(errors, row_number, "Missing vendor, hen type, kg, or rate")
                continue

            party_name = str(row[party_col]).strip()
            weight = float(row[weight_col])
            rate = float(row[rate_col])
            quantity = Decimal(str(float(row[quantity_col]))) if quantity_col and not pd.isna(row[quantity_col]) else None
            date = resolve_upload_date(row, date_col, fallback_date)
            if not date:
                skipped += 1
                row_error(errors, row_number, "Invalid or missing date")
                continue
            item_type = normalize_source(row[item_col])
            category = get_optional_row_value(row, ["CATEGORY"])
            if weight <= 0 or rate <= 0 or (quantity is not None and quantity < 0):
                skipped += 1
                row_error(errors, row_number, "NAG cannot be negative. KG and rate must be greater than zero")
                continue

            # --- Party mapping ---
            party_id = get_or_create_party(db, party_name, "VENDOR", seen_aliases)

            # --- Duplicate check ---
            existing_txn = db.query(models.Transaction).filter_by(
                date=date,
                party_id=party_id,
                weight=weight,
                rate=rate,
                type="SALE",
                category=category,
                item_type=item_type
            ).first()

            txn_key = (date, party_id, weight, rate, "SALE", category, item_type)

            if existing_txn or txn_key in seen_transactions:
                skipped += 1
                continue

            # --- Create sale transaction ---
            txn = models.Transaction(
                date=date,
                party_id=party_id,
                type="SALE",
                category=category,
                item_type=item_type,
                quantity=quantity,
                weight=weight,
                rate=rate,
                amount=weight * rate,
                payment_mode="NA"
            )

            db.add(txn)
            seen_transactions.add(txn_key)
            inserted += 1

        except Exception as e:
            skipped += 1
            row_error(errors, row_number, str(e))
            continue

    # --- Final commit (ONLY ONCE) ---
    try:
        if preview:
            db.rollback()
            return upload_result(inserted, skipped, errors, {"preview_mode": True})

        file_record = models.UploadedFile(
            file_hash=file_hash,
            file_type="vendor"
        )
        db.add(file_record)
        db.commit()

    except Exception as e:
        db.rollback()
        return {"error": "Transaction failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)

@app.post("/upload/dealer")
def upload_dealer(file: UploadFile = File(...), preview: bool = False, input_date: str = None, db: Session = Depends(get_db)):

    import io
    import hashlib

    filename = (file.filename or "").lower()
    contents = file.file.read()

    # --- File hash protection ---
    file_hash = hashlib.sha256(contents).hexdigest()

    existing_file = db.query(models.UploadedFile).filter_by(file_hash=file_hash).first()
    if existing_file:
        return {"error": "File already uploaded"}

    # --- Read file ---
    try:
        if filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")

        elif filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")

        elif filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8")

        else:
            return {"error": "Unsupported file format"}

    except Exception as e:
        return {"error": f"File read failed: {str(e)}"}

    # --- Validation ---
    if df.empty:
        return {"error": "File is empty"}

    df.columns = df.columns.str.strip().str.upper()

    date_col = get_first_existing_column(df, ["DATE"])
    fallback_date = parse_input_date(input_date) if input_date else None
    if not date_col and not fallback_date:
        return {"error": "Provide DATE column in file or select the upload date in the app"}

    party_col, error = require_column(df, ["DEALER", "PARTY", "NAME"], "DEALER")
    if error:
        return error

    bill_number_col = get_first_existing_column(df, ["BILL_NO", "BILL NO", "BILL_NUMBER", "BILL NUMBER", "INVOICE_NO", "INVOICE NO"])

    item_col, error = require_column(df, ["HEN_TYPE", "HEN TYPE", "ITEM", "TYPE"], "HEN TYPE")
    if error:
        return error

    weight_col, error = require_column(df, ["KGS", "KG", "WEIGHT"], "KGS")
    if error:
        return error

    rate_col, error = require_column(df, ["RATE_PER_KG", "RATE PER KG", "RATE/KG", "RATE"], "RATE PER KG")
    if error:
        return error

    quantity_col = get_first_existing_column(df, ["NAG", "QTY", "QUANTITY", "PCS", "PIECES"])

    # --- Numeric validation ---
    try:
        df[weight_col] = df[weight_col].astype(float)
        df[rate_col] = df[rate_col].astype(float)
        if quantity_col:
            df[quantity_col] = df[quantity_col].astype(float)
    except:
        return {"error": "Invalid numeric values"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_transactions = set()

    # --- Process rows ---
    for index, row in df.iterrows():
        row_number = int(index) + 2
        try:
            if (
                pd.isna(row[party_col]) or
                pd.isna(row[item_col]) or
                pd.isna(row[weight_col]) or
                pd.isna(row[rate_col])
            ):
                skipped += 1
                row_error(errors, row_number, "Missing dealer, hen type, kg, or rate")
                continue

            party_name = str(row[party_col]).strip()
            bill_number = get_optional_row_value(row, ["BILL_NO", "BILL NO", "BILL_NUMBER", "BILL NUMBER", "INVOICE_NO", "INVOICE NO"])
            item_type = normalize_source(row[item_col])
            weight = float(row[weight_col])
            rate = float(row[rate_col])
            quantity = Decimal(str(float(row[quantity_col]))) if quantity_col and not pd.isna(row[quantity_col]) else None
            date = resolve_upload_date(row, date_col, fallback_date)
            if not date:
                skipped += 1
                row_error(errors, row_number, "Invalid or missing date")
                continue

            if weight <= 0 or rate <= 0 or (quantity is not None and quantity < 0):
                skipped += 1
                row_error(errors, row_number, "NAG cannot be negative. KG and rate must be greater than zero")
                continue

            # --- Party mapping ---
            party_id = get_or_create_party(db, party_name, "DEALER", seen_aliases)

            # --- Duplicate check (CORRECT TYPE) ---
            existing_txn = db.query(models.Transaction).filter_by(
                date=date,
                party_id=party_id,
                weight=weight,
                rate=rate,
                type="PURCHASE",
                category=None,
                item_type=item_type,
                bill_number=bill_number
            ).first()

            txn_key = (date, party_id, weight, rate, "PURCHASE", bill_number or "", item_type)

            if existing_txn or txn_key in seen_transactions:
                skipped += 1
                continue

            # --- Create purchase transaction ---
            txn = models.Transaction(
                date=date,
                party_id=party_id,
                type="PURCHASE",
                category=None,
                item_type=item_type,
                quantity=quantity,
                weight=weight,
                rate=rate,
                amount=weight * rate,
                payment_mode="NA",
                bill_number=bill_number
            )

            db.add(txn)
            seen_transactions.add(txn_key)
            inserted += 1

        except Exception as e:
            skipped += 1
            row_error(errors, row_number, str(e))
            continue

    # --- Final commit ---
    try:
        if preview:
            db.rollback()
            return upload_result(inserted, skipped, errors, {"preview_mode": True})

        file_record = models.UploadedFile(
            file_hash=file_hash,
            file_type="dealer"
        )
        db.add(file_record)
        db.commit()

    except Exception as e:
        db.rollback()
        return {"error": "Transaction failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/upload/payment")
def upload_payment(file: UploadFile = File(...), preview: bool = False, input_date: str = None, db: Session = Depends(get_db)):

    import io
    import hashlib

    filename = (file.filename or "").lower()
    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()

    existing_file = db.query(models.UploadedFile).filter_by(file_hash=file_hash).first()
    if existing_file:
        return {"error": "File already uploaded"}

    try:
        if filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        elif filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")
        elif filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8")
        else:
            return {"error": "Unsupported file format"}
    except Exception as e:
        return {"error": f"File read failed: {str(e)}"}

    if df.empty:
        return {"error": "File is empty"}

    df.columns = df.columns.str.strip().str.upper()

    date_col = "DATE" if "DATE" in df.columns else None
    required_cols = ["PARTY", "AMOUNT", "PAYMENT_MODE", "DIRECTION"]
    for col in required_cols:
        if col not in df.columns:
            return {"error": f"Missing column: {col}"}

    try:
        df["AMOUNT"] = df["AMOUNT"].astype(float)
    except Exception:
        return {"error": "Invalid amount values"}

    fallback_date = parse_input_date(input_date) if input_date else None
    if not date_col and not fallback_date:
        return {"error": "Provide DATE column in file or select the upload date in the app"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_payments = set()

    for index, row in df.iterrows():
        row_number = int(index) + 2
        try:
            if pd.isna(row["PARTY"]) or pd.isna(row["AMOUNT"]) or pd.isna(row["DIRECTION"]):
                skipped += 1
                row_error(errors, row_number, "Missing party, amount, or direction")
                continue

            party_name = str(row["PARTY"]).strip()
            amount = Decimal(str(float(row["AMOUNT"])))
            target_date = resolve_upload_date(row, date_col, fallback_date)
            if not target_date:
                skipped += 1
                row_error(errors, row_number, "Invalid or missing date")
                continue
            payment_mode = str(row["PAYMENT_MODE"]).strip() if not pd.isna(row["PAYMENT_MODE"]) else "NA"
            direction = normalize_payment_direction(row["DIRECTION"])

            if not party_name or amount <= 0 or not direction:
                skipped += 1
                row_error(errors, row_number, "Invalid party, amount, or payment direction")
                continue

            party_id = get_or_create_party(db, party_name, "BOTH", seen_aliases)

            existing_payment = db.query(models.Transaction).filter_by(
                date=target_date,
                party_id=party_id,
                type="PAYMENT",
                amount=amount,
                payment_mode=payment_mode,
                category=direction
            ).first()

            payment_key = (target_date, party_id, amount, payment_mode, direction)
            if existing_payment or payment_key in seen_payments:
                skipped += 1
                continue

            txn = models.Transaction(
                date=target_date,
                party_id=party_id,
                type="PAYMENT",
                category=direction,
                amount=amount,
                payment_mode=payment_mode
            )

            db.add(txn)
            seen_payments.add(payment_key)
            inserted += 1

        except Exception as e:
            skipped += 1
            row_error(errors, row_number, str(e))
            continue

    try:
        if preview:
            db.rollback()
            return upload_result(inserted, skipped, errors, {"preview_mode": True})

        file_record = models.UploadedFile(
            file_hash=file_hash,
            file_type="payment"
        )
        db.add(file_record)
        db.commit()

    except Exception as e:
        db.rollback()
        return {"error": "Transaction failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/upload/opening-balance")
def upload_opening_balance(file: UploadFile = File(...), preview: bool = False, db: Session = Depends(get_db)):

    import io
    import hashlib

    filename = (file.filename or "").lower()
    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()

    existing_file = db.query(models.UploadedFile).filter_by(file_hash=file_hash).first()
    if existing_file:
        return {"error": "File already uploaded"}

    try:
        if filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        elif filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")
        elif filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8")
        else:
            return {"error": "Unsupported file format"}
    except Exception as e:
        return {"error": f"File read failed: {str(e)}"}

    if df.empty:
        return {"error": "File is empty"}

    df.columns = df.columns.str.strip().str.upper()

    date_col, error = require_column(df, ["DATE"], "DATE")
    if error:
        return error

    party_col, error = require_column(df, ["PARTY", "NAME", "VENDOR", "DEALER"], "PARTY")
    if error:
        return error

    amount_col, error = require_column(df, ["OPENING_BALANCE", "OPENING BALANCE", "AMOUNT", "BALANCE"], "OPENING BALANCE")
    if error:
        return error

    balance_type_col, error = require_column(df, ["BALANCE_TYPE", "BALANCE TYPE", "TYPE"], "BALANCE TYPE")
    if error:
        return error

    try:
        df[amount_col] = df[amount_col].astype(float)
    except Exception:
        return {"error": "Invalid opening balance values"}

    inserted = 0
    skipped = 0
    errors = []
    seen_aliases = {}
    seen_opening = set()

    for index, row in df.iterrows():
        row_number = int(index) + 2
        try:
            if pd.isna(row[party_col]) or pd.isna(row[amount_col]) or pd.isna(row[balance_type_col]):
                skipped += 1
                row_error(errors, row_number, "Missing party, opening balance, or balance type")
                continue

            party_name = str(row[party_col]).strip()
            amount = Decimal(str(float(row[amount_col])))
            target_date = pd.to_datetime(row[date_col], dayfirst=True).date()
            balance_type = str(row[balance_type_col]).strip().upper()

            if balance_type in ["RECEIVABLE", "RECEIVE", "CUSTOMER", "VENDOR"]:
                balance_type = "RECEIVABLE"
                party_type = "VENDOR"
            elif balance_type in ["PAYABLE", "PAY", "SUPPLIER", "DEALER"]:
                balance_type = "PAYABLE"
                party_type = "DEALER"
            else:
                skipped += 1
                row_error(errors, row_number, "Balance type must be RECEIVABLE or PAYABLE")
                continue

            if not party_name or amount < 0:
                skipped += 1
                row_error(errors, row_number, "Invalid party or opening balance")
                continue

            party_id = get_or_create_party(db, party_name, party_type, seen_aliases)

            existing_opening = db.query(models.Transaction).filter_by(
                date=target_date,
                party_id=party_id,
                type="OPENING",
                category=balance_type
            ).first()

            opening_key = (target_date, party_id, balance_type)
            if existing_opening or opening_key in seen_opening:
                skipped += 1
                continue

            txn = models.Transaction(
                date=target_date,
                party_id=party_id,
                type="OPENING",
                category=balance_type,
                amount=amount,
                payment_mode="NA"
            )

            db.add(txn)
            seen_opening.add(opening_key)
            inserted += 1

        except Exception as e:
            skipped += 1
            row_error(errors, row_number, str(e))
            continue

    try:
        if preview:
            db.rollback()
            return upload_result(inserted, skipped, errors, {"preview_mode": True})

        file_record = models.UploadedFile(
            file_hash=file_hash,
            file_type="opening_balance"
        )
        db.add(file_record)
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Transaction failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/upload/opening-stock")
def upload_opening_stock(file: UploadFile = File(...), preview: bool = False, db: Session = Depends(get_db)):

    import io
    import hashlib

    filename = (file.filename or "").lower()
    contents = file.file.read()
    file_hash = hashlib.sha256(contents).hexdigest()

    existing_file = db.query(models.UploadedFile).filter_by(file_hash=file_hash).first()
    if existing_file:
        return {"error": "File already uploaded"}

    try:
        if filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        elif filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")
        elif filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8")
        else:
            return {"error": "Unsupported file format"}
    except Exception as e:
        return {"error": f"File read failed: {str(e)}"}

    if df.empty:
        return {"error": "File is empty"}

    df.columns = df.columns.str.strip().str.upper()

    date_col, error = require_column(df, ["DATE"], "DATE")
    if error:
        return error

    item_col, error = require_column(df, ["HEN_TYPE", "HEN TYPE", "ITEM", "TYPE"], "HEN TYPE")
    if error:
        return error

    weight_col, error = require_column(df, ["OPENING_KGS", "OPENING KGS", "KGS", "KG", "WEIGHT"], "OPENING KGS")
    if error:
        return error

    quantity_col = get_first_existing_column(df, ["OPENING_NAG", "OPENING QTY", "NAG", "QTY", "QUANTITY", "PCS", "PIECES"])

    try:
        df[weight_col] = df[weight_col].astype(float)
        if quantity_col:
            df[quantity_col] = df[quantity_col].astype(float)
    except Exception:
        return {"error": "Invalid opening stock values"}

    inserted = 0
    skipped = 0
    errors = []
    seen_stock = set()

    for index, row in df.iterrows():
        row_number = int(index) + 2
        try:
            if pd.isna(row[item_col]) or pd.isna(row[weight_col]):
                skipped += 1
                row_error(errors, row_number, "Missing hen type or opening kg")
                continue

            target_date = pd.to_datetime(row[date_col], dayfirst=True).date()
            item_type = normalize_source(row[item_col])
            opening_weight = Decimal(str(float(row[weight_col])))
            opening_quantity = Decimal(str(float(row[quantity_col]))) if quantity_col and not pd.isna(row[quantity_col]) else None

            if not item_type or opening_weight < 0 or (opening_quantity is not None and opening_quantity < 0):
                skipped += 1
                row_error(errors, row_number, "Invalid hen type, opening kg, or opening NAG")
                continue

            existing_stock = db.query(models.ItemOpeningStock).filter_by(
                date=target_date,
                item_type=item_type
            ).first()

            stock_key = (target_date, item_type)
            if existing_stock or stock_key in seen_stock:
                skipped += 1
                continue

            db.add(models.ItemOpeningStock(
                date=target_date,
                item_type=item_type,
                opening_quantity=opening_quantity,
                opening_weight=opening_weight
            ))
            seen_stock.add(stock_key)
            inserted += 1

        except Exception as e:
            skipped += 1
            row_error(errors, row_number, str(e))
            continue

    try:
        if preview:
            db.rollback()
            return upload_result(inserted, skipped, errors, {"preview_mode": True})

        file_record = models.UploadedFile(
            file_hash=file_hash,
            file_type="opening_stock"
        )
        db.add(file_record)
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Transaction failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.post("/process-day")
def process_day(input_date: str, actual_stock: float, db: Session = Depends(get_db)):
    return {"error": "Use Process Day hen-type counts; a combined count cannot identify live stock sources."}


@app.post("/process-day/items")
def process_day_items(input_date: str, actual_stock: list[dict], db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Invalid date format"}
    if not actual_stock:
        return {"error": "Enter actual stock for at least one hen type"}

    actuals = {}
    for row in actual_stock:
        item = normalize_source(row.get("item_type"))
        if item not in SOURCE_ITEMS:
            return {"error": f"Choose a live hen type: {', '.join(SOURCE_ITEMS)}"}
        if item in actuals:
            return {"error": f"Duplicate live count for {item}"}
        try:
            weight = Decimal(str(row.get("actual_weight", 0) or 0))
            quantity = Decimal(str(row.get("actual_quantity", 0) or 0))
            if not weight.is_finite() or not quantity.is_finite() or weight < 0 or quantity < 0 or quantity != quantity.to_integral_value():
                raise ValueError()
        except Exception:
            return {"error": f"Enter non-negative kg and whole NAG for {item}"}
        actuals[item] = (weight, quantity)
    for item in SOURCE_ITEMS:
        actuals.setdefault(item, (Decimal("0"), Decimal("0")))

    try:
        # Serialize same-outlet counts and downstream snapshot refreshes.
        db.query(models.Outlet).filter_by(id=current_outlet.id).with_for_update().one()
        existing = db.query(models.DailyItemStock).filter_by(
            outlet_id=current_outlet.id, date=target_date
        )
        replaced_existing = existing.count() > 0
        existing.delete(synchronize_session=False)
        for item, (weight, quantity) in actuals.items():
            db.add(models.DailyItemStock(
                outlet_id=current_outlet.id, date=target_date, item_type=item,
                actual_closing_weight=weight, actual_closing_quantity=quantity,
            ))
        history = refresh_stock_snapshots(db, current_outlet.id, target_date)
        snapshot = history.snapshot(target_date)
        db.commit()
    except Exception as exc:
        db.rollback()
        return {"error": "Processing failed", "details": str(exc)}

    total = snapshot["totals"]
    return {
        "status": "success", "date": str(target_date),
        "replaced_existing": replaced_existing, "warning": snapshot["warning"],
        "items": [{
            "item": row["item"],
            "opening_stock": optional_float(row["opening_weight"]),
            "expected_stock": optional_float(row["expected_closing_weight"]),
            "actual_stock": optional_float(row["actual_closing_weight"]),
            "actual_nag": optional_float(row["actual_closing_quantity"]),
            "leakage": optional_float(row["leakage"]),
        } for row in snapshot["rows"]],
        "total_expected_nag": optional_float(total["expected_closing_quantity"]),
        "total_expected_stock": optional_float(total["expected_closing_weight"]),
        "total_actual_nag": optional_float(total["actual_closing_quantity"]),
        "total_actual_stock": optional_float(total["actual_closing_weight"]),
        "total_quantity_leakage": optional_float(total["quantity_leakage"]),
        "total_leakage": optional_float(total["leakage"]),
    }


@app.get("/party/{party_id}/ledger")
def get_party_ledger(
    party_id: UUID,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
    scope=Depends(get_outlet_scope)
):

    # --- Validate party ---
    party = db.query(models.Party).filter_by(id=party_id).first()
    if not party:
        return {"error": "Party not found"}

    start = None
    end = None

    query = apply_outlet_scope(
        db.query(models.Transaction).filter_by(party_id=party_id),
        models.Transaction,
        scope
    )

    if start_date:
        start = parse_input_date(start_date)
        if not start:
            return {"error": "Invalid start date"}

    if end_date:
        end = parse_input_date(end_date)
        if not end:
            return {"error": "Invalid end date"}
    if error := ledger_range_error(start, end):
        return {"error": error}
    opening_balances, balances, txns, ledger, mode = build_party_ledger_window(db, query, start=start, end=end)

    if not txns:
        return {
            "party_id": party_id,
            "party_name": party.name,
            "total_balance": float(opening_balances["balance"]),
            "balances": serialize_ledger_balances(opening_balances),
            "summary": build_party_summary_for_period([], opening_balances, opening_balances),
            "ledger_mode": mode,
            "cutover_date": str(LEDGER_CUTOVER_DATE),
            "ledger": []
        }

    return {
        "party_id": party_id,
        "party_name": party.name,
        "total_balance": float(balances["balance"]),
        "balances": serialize_ledger_balances(balances),
        "summary": build_party_summary_for_period(ledger, opening_balances, balances),
        "ledger_mode": mode,
        "cutover_date": str(LEDGER_CUTOVER_DATE),
        "ledger": ledger
    }

@app.get("/party/search")
def search_party(name: str, db: Session = Depends(get_db)):

    if not name or len(name.strip()) < 2:
        return {"results": []}

    normalized = normalize_party_name(name)

    # --- Search aliases (limited for performance) ---
    aliases = db.query(models.PartyAlias).filter(
        models.PartyAlias.normalized_alias.contains(normalized)
    ).limit(10).all()

    if not aliases:
        return {"results": []}

    party_ids = list(set([a.party_id for a in aliases]))

    parties = db.query(models.Party).filter(
        models.Party.id.in_(party_ids)
    ).all()

    # --- Sort by name (better UX) ---
    parties = sorted(parties, key=lambda x: x.name)

    return {
        "results": [
            {
                "id": str(p.id),
                "name": p.name,
                "type": p.type,
                "phone": p.phone or "",
                "address": p.address or ""
            }
            for p in parties
        ]
    }


@app.get("/party/profile")
def get_party_profile(name: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    if not name or len(name.strip()) < 2:
        return {"error": "Invalid party name"}

    party = get_party_by_name(db, name)
    if not party:
        return {"error": "Party not found"}

    txns = apply_outlet_scope(
        db.query(models.Transaction).filter(models.Transaction.party_id == party.id),
        models.Transaction,
        scope
    ).order_by(
        models.Transaction.date.asc(),
        models.Transaction.created_at.asc()
    ).all()
    balances, _ = build_ledger(db, txns)
    serialized_balances = serialize_ledger_balances(balances)
    primary_balance = balances["balance"]

    return {
        "party": {
            "id": str(party.id),
            "name": party.name,
            "type": party.type or "",
            "phone": party.phone or "",
            "address": party.address or "",
            "balance_after": float(primary_balance),
            "receivable_balance": serialized_balances["receivable"],
            "payable_balance": serialized_balances["payable"],
            "net_balance": serialized_balances["net"],
        },
        "balances": serialized_balances,
    }


@app.get("/party-directory")
def list_party_directory(name: str | None = None, db: Session = Depends(get_db)):
    query = db.query(models.Party)
    if name and len(name.strip()) >= 2:
        normalized = normalize_party_name(name)
        query = query.filter(
            or_(
                models.Party.normalized_name.contains(normalized),
                func.lower(models.Party.name).contains(name.strip().lower())
            )
        )

    parties = query.order_by(models.Party.name.asc()).limit(500).all()
    return {
        "results": [
            {
                "id": str(p.id),
                "name": p.name,
                "type": p.type or "",
                "phone": p.phone or "",
                "address": p.address or ""
            }
            for p in parties
        ]
    }


@app.post("/party-directory")
def save_party_directory(payload: dict = Body(...), db: Session = Depends(get_db)):
    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one party"}

    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    for index, row in enumerate(rows, start=1):
        try:
            party_id_raw = str(row.get("party_id") or "").strip()
            party_name = str(row.get("name") or row.get("party") or "").strip()
            party_type = str(row.get("type") or "BOTH").strip().upper() or "BOTH"
            phone = str(row.get("phone") or "").strip()
            address = str(row.get("address") or "").strip()

            if not party_name:
                skipped += 1
                row_error(errors, index, "Enter party name")
                continue

            party = None
            if party_id_raw:
                try:
                    party = db.query(models.Party).filter_by(id=UUID(party_id_raw)).first()
                except ValueError:
                    party = None
            if party is None:
                party = get_party_by_name(db, party_name)
            if party:
                had_name = party.name
                had_type = party.type or ""
                had_phone = (party.phone or "").strip()
                had_address = (party.address or "").strip()
                update_party_details(party, party_type, phone, address, overwrite_contact=True)
                if party.name != party_name:
                    party.name = party_name
                    party.normalized_name = normalize_party_name(party_name)
                    normalized_alias = party.normalized_name
                    alias = db.query(models.PartyAlias).filter_by(
                        party_id=party.id,
                        normalized_alias=normalized_alias
                    ).first()
                    if not alias:
                        db.add(models.PartyAlias(
                            alias=party_name,
                            normalized_alias=normalized_alias,
                            party_id=party.id
                        ))
                changed = any([
                    had_name != party.name,
                    had_type != (party.type or ""),
                    had_phone != (party.phone or "").strip(),
                    had_address != (party.address or "").strip()
                ])
                updated += 1 if changed else 0
            else:
                get_or_create_party(db, party_name, party_type, {}, phone=phone, address=address)
                inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving party directory failed", "details": str(e)}

    return {
        "status": "success",
        "rows_inserted": inserted,
        "rows_updated": updated,
        "rows_skipped": skipped,
        "errors": errors
    }


@app.get("/party/ledger")
def get_ledger_by_name(
    name: str,
    start_date: str | None = None,
    end_date: str | None = None,
    db: Session = Depends(get_db),
    scope=Depends(get_outlet_scope)
):
    if not name or len(name.strip()) < 2:
        return {"error": "Invalid party name"}

    normalized = normalize_party_name(name)

    # --- Step 1: Try EXACT match first (important) ---
    alias = db.query(models.PartyAlias).filter(
        models.PartyAlias.normalized_alias == normalized
    ).first()

    if alias:
        party_ids = [alias.party_id]
    else:
        # --- Step 2: Fuzzy match (limited results) ---
        aliases = db.query(models.PartyAlias).filter(
            models.PartyAlias.normalized_alias.contains(normalized)
        ).limit(10).all()

        if not aliases:
            return {"error": "Party not found"}

        party_ids = list(set([a.party_id for a in aliases]))

    # --- Step 3: Multiple matches ---
    if len(party_ids) > 1:
        parties = db.query(models.Party).filter(
            models.Party.id.in_(party_ids)
        ).all()

        return {
            "multiple_matches": True,
            "results": [
                {
                    "id": str(p.id),
                    "name": p.name,
                    "type": p.type
                }
                for p in parties
            ]
        }

    # --- Step 4: Single match → ledger ---
    party_id = party_ids[0]

    start = None
    end = None

    query = apply_outlet_scope(
        db.query(models.Transaction).filter(models.Transaction.party_id == party_id),
        models.Transaction,
        scope
    )

    if start_date:
        start = parse_input_date(start_date)
        if not start:
            return {"error": "Invalid start date"}

    if end_date:
        end = parse_input_date(end_date)
        if not end:
            return {"error": "Invalid end date"}

    if error := ledger_range_error(start, end):
        return {"error": error}

    party = db.query(models.Party).filter_by(id=party_id).first()
    opening_balances, balances, txns, ledger, mode = build_party_ledger_window(db, query, start=start, end=end)

    if not txns:
        return {
            "party_name": party.name if party else name,
            "total_balance": float(opening_balances["balance"]),
            "balances": serialize_ledger_balances(opening_balances),
            "summary": build_party_summary_for_period([], opening_balances, opening_balances),
            "ledger_mode": mode,
            "cutover_date": str(LEDGER_CUTOVER_DATE),
            "ledger": []
        }

    return {
        "party_name": party.name if party else name,
        "total_balance": float(balances["balance"]),
        "balances": serialize_ledger_balances(balances),
        "summary": build_party_summary_for_period(ledger, opening_balances, balances),
        "ledger_mode": mode,
        "cutover_date": str(LEDGER_CUTOVER_DATE),
        "ledger": ledger
    }


@app.get("/party/detail")
def get_party_detail(name: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    if not name or len(name.strip()) < 2:
        return {"error": "Invalid party name"}

    normalized = normalize_party_name(name)
    alias = db.query(models.PartyAlias).filter(
        models.PartyAlias.normalized_alias == normalized
    ).first()

    if not alias:
        alias = db.query(models.PartyAlias).filter(
            models.PartyAlias.normalized_alias.contains(normalized)
        ).first()

    if not alias:
        return {"error": "Party not found"}

    party = db.query(models.Party).filter_by(id=alias.party_id).first()
    if not party:
        return {"error": "Party not found"}

    txns = apply_outlet_scope(
        db.query(models.Transaction).filter_by(party_id=party.id),
        models.Transaction,
        scope
    ).order_by(models.Transaction.date.asc()).all()
    txns = filter_party_ledger_transactions(db, txns)

    balances, ledger = build_ledger(db, txns)

    return {
        "party": {
            "id": str(party.id),
            "name": party.name,
            "type": party.type,
            "phone": party.phone or "",
            "address": party.address or ""
        },
        "balances": serialize_ledger_balances(balances),
        "summary": build_party_summary(db, txns, balances),
        "ledger": ledger
    }


@app.get("/reports/export")
def export_report(
    report_type: str,
    file_format: str = "excel",
    party: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    date: str | None = None,
    db: Session = Depends(get_db),
    scope=Depends(get_outlet_scope)
):
    report_type = report_type.lower().strip()
    file_format = file_format.lower().strip()
    start, end = parse_report_dates(start_date, end_date)

    if (start_date and not start) or (end_date and not end):
        return {"error": "Invalid date format"}

    if report_type == "inventory" and scope["mode"] == "all":
        return {"error": "Select one outlet for inventory report"}

    if report_type == "ledger":
        if not party or len(party.strip()) < 2:
            return {"error": "Party name is required for ledger report"}
        if error := ledger_range_error(start, end):
            return {"error": error}

        normalized = normalize_party_name(party)
        alias = db.query(models.PartyAlias).filter(
            models.PartyAlias.normalized_alias == normalized
        ).first()

        if not alias:
            alias = db.query(models.PartyAlias).filter(
                models.PartyAlias.normalized_alias.contains(normalized)
            ).first()

        if not alias:
            return {"error": "Party not found"}

        party_row = db.query(models.Party).filter_by(id=alias.party_id).first()
        query = apply_outlet_scope(
            db.query(models.Transaction).filter(models.Transaction.party_id == alias.party_id),
            models.Transaction,
            scope
        )

        opening_balances, balances, txns, ledger, mode = build_party_ledger_window(db, query, start=start, end=end)
        if mode == ACCOUNT_LEDGER:
            rows = [{
                "Date": format_export_date(row["date"]),
                "Account": row.get("account", ""),
                "Type": row["type"],
                "Bill No": row.get("bill_number", ""),
                "Details": ledger_row_details(row),
                "Debit": row.get("debit", 0),
                "Credit": row.get("credit", 0),
                "Account Balance": row.get("account_balance", 0),
                "Net": row.get("net_balance", 0),
            } for row in ledger]
            rows.append({
                "Date": "", "Account": "", "Type": "Closing Balance", "Bill No": "",
                "Details": "", "Debit": "", "Credit": "",
                "Account Balance": "", "Net": float(balances["net"]),
            })
            columns = ["Date", "Account", "Type", "Bill No", "Details", "Debit", "Credit", "Account Balance", "Net"]
        else:
            rows = [{
                "Date": format_export_date(row["date"]),
                "Type": row["type"],
                "Bill No": row.get("bill_number", ""),
                "Category": row["category"],
                "Item": row["item"],
                "NAG": row.get("quantity", 0),
                "KGS": row.get("weight", 0),
                "Rate": row.get("rate", 0),
                "Mode": row["payment_mode"],
                "Amount": row["amount"],
                "Balance": row["balance"]
            } for row in ledger]
            rows.append({
                "Date": "", "Type": "Closing Balance", "Bill No": "", "Category": "",
                "Item": "", "NAG": "", "KGS": "", "Rate": "", "Mode": "", "Amount": "",
                "Balance": float(balances["balance"]),
            })
            columns = ["Date", "Type", "Bill No", "Category", "Item", "NAG", "KGS", "Rate", "Mode", "Amount", "Balance"]
        period_label = "All Dates"
        if start and end:
            period_label = f"{format_export_date(start)} to {format_export_date(end)}"
        elif start:
            period_label = f"From {format_export_date(start)}"
        elif end:
            period_label = f"Up to {format_export_date(end)}"
        filename = safe_filename(f"ledger_{party}")
        balance_meta = [
            f"Opening Receivable: {float(opening_balances['receivable']):,.2f}",
            f"Opening Payable: {float(opening_balances['payable']):,.2f}",
            f"Closing Receivable: {float(balances['receivable']):,.2f}",
            f"Closing Payable: {float(balances['payable']):,.2f}",
            f"Closing Net: {float(balances['net']):,.2f}",
        ] if mode == ACCOUNT_LEDGER else [
            f"Opening Balance: {float(opening_balances['balance']):,.2f}",
            f"Closing Balance: {float(balances['balance']):,.2f}",
        ]
        meta_rows = [
            f"Party: {party_row.name if party_row else party}",
            f"Period: {period_label}",
            f"Ledger method: {'Receivable / Payable' if mode == ACCOUNT_LEDGER else 'Restored historical balance'}",
            *balance_meta,
        ]
        return report_response(rows, columns, filename, file_format, f"Ledger Report - {party}", meta_rows=meta_rows)

    if report_type == "summary":
        events = financial_events(db, scope, start, end)
        by_date = {}
        for event in events:
            day = event["date"]
            row = by_date.setdefault(day, {
                "Sales": Decimal("0"), "Purchase": Decimal("0"),
                "Payment Received": Decimal("0"), "Payment Paid": Decimal("0"), "Opening": Decimal("0"),
            })
            key = {"SALE": "Sales", "PURCHASE": "Purchase", "PAYMENT RECEIVED": "Payment Received",
                   "PAYMENT PAID": "Payment Paid"}.get(event["type"])
            if event["entry_type"] == "OPENING":
                key = "Opening"
            if key:
                row[key] += Decimal(event["amount"])
        histories = scoped_stock_histories(db, scope, end or max(by_date, default=pd.Timestamp.today().date()))
        rows = [{
            "Date": format_export_date(day),
            **{key: float(value) for key, value in row.items()},
            "Profit": optional_float(stock_total_for_day(histories, day, "gross_profit")),
        } for day, row in sorted(by_date.items())]
        columns = ["Date", "Sales", "Purchase", "Profit", "Payment Received", "Payment Paid", "Opening"]
        return report_response(rows, columns, "financial_summary", file_format, "Financial Summary",
                               meta_rows=["Profit is gross margin using purchase costs; blank means cost data is incomplete."])

    if report_type == "outstanding":
        as_of = end or ledger_today()
        mode = ledger_mode_for_date(as_of)
        query = apply_outlet_scope(db.query(models.Transaction).filter(
            models.Transaction.party_id.isnot(None)
        ), models.Transaction, scope)
        query = query.filter(models.Transaction.date <= as_of)
        query = query.filter(
            models.Transaction.date >= LEDGER_CUTOVER_DATE
            if mode == ACCOUNT_LEDGER
            else models.Transaction.date < LEDGER_CUTOVER_DATE
        )
        txns = query.order_by(models.Transaction.date, models.Transaction.created_at, models.Transaction.id).all()
        grouped = {}
        for txn in txns:
            grouped.setdefault(txn.party_id, []).append(txn)
        parties = db.query(models.Party).filter(models.Party.id.in_(list(grouped))).order_by(models.Party.name).all()
        rows = []
        for party_row in parties:
            balances, _ = build_ledger(db, grouped[party_row.id], mode=mode, as_of=as_of)
            if mode == ACCOUNT_LEDGER:
                if not balances["receivable"] and not balances["payable"]:
                    continue
                rows.append({
                    "Party": party_row.name,
                    "Type": party_row.type or "",
                    "Receivable": float(balances["receivable"]),
                    "Payable": float(balances["payable"]),
                    "Net Outstanding": float(balances["net"]),
                })
            else:
                if not balances["balance"]:
                    continue
                rows.append({
                    "Party": party_row.name,
                    "Type": party_row.type or "",
                    "Balance": float(balances["balance"]),
                })

        columns = ["Party", "Type", "Receivable", "Payable", "Net Outstanding"] if mode == ACCOUNT_LEDGER else ["Party", "Type", "Balance"]
        return report_response(rows, columns, "outstanding_balances", file_format, "Outstanding Balances")

    if report_type == "inventory":
        target = parse_input_date(date) if date else pd.Timestamp.today().date()
        if not target:
            return {"error": "Invalid date format"}
        snapshot = StockHistory(db, resolve_scope_outlet_id(scope), target).snapshot(target)
        rows = [{
            "Date": format_export_date(target), "Item": row["item"],
            "Opening NAG": optional_float(row["opening_quantity"]),
            "Opening Kg": optional_float(row["opening_weight"]),
            "Purchase Kg": optional_float(row["purchase_weight"]),
            "Sales Kg": optional_float(row["sales_weight"]),
            "Live Cut Kg": optional_float(row["cut_weight"]),
            "Mortality Kg": optional_float(row["mortality_weight"]),
            "Expected Kg": optional_float(row["expected_closing_weight"]),
            "Actual NAG": optional_float(row["actual_closing_quantity"]),
            "Actual Kg": optional_float(row["actual_closing_weight"]),
            "Leakage Kg": optional_float(row["leakage"]),
        } for row in snapshot["rows"]]
        columns = ["Date", "Item", "Opening NAG", "Opening Kg", "Purchase Kg", "Sales Kg", "Live Cut Kg", "Mortality Kg", "Expected Kg", "Actual NAG", "Actual Kg", "Leakage Kg"]
        return report_response(rows, columns, f"inventory_{target}", file_format,
                               f"Live Bird Inventory - {format_export_date(target)}",
                               meta_rows=[snapshot["warning"]] if snapshot["warning"] else [])

    if report_type == "transactions":
        query = db.query(models.Transaction, models.Party).outerjoin(
            models.Party,
            models.Transaction.party_id == models.Party.id
        )
        query = apply_outlet_scope(query, models.Transaction, scope)
        if start:
            query = query.filter(models.Transaction.date >= start)
        if end:
            query = query.filter(models.Transaction.date <= end)
        if party and len(party.strip()) >= 2:
            normalized = normalize_party_name(party)
            aliases = db.query(models.PartyAlias).filter(
                models.PartyAlias.normalized_alias.contains(normalized)
            ).all()
            party_ids = [alias.party_id for alias in aliases]
            query = query.filter(models.Transaction.party_id.in_(party_ids))

        rows = []
        for txn, party_row in query.order_by(models.Transaction.date.asc()).all():
            rows.append({
                "Date": str(txn.date),
                "Party": party_row.name if party_row else (txn.category.title() if txn.category else "Walk-in Customer"),
                "Party Type": party_row.type if party_row else ("RETAIL" if txn.type == "SALE" else ""),
                "Type": txn.type or "",
                "Category": txn.category or "",
                "Item": txn.item_type or "",
                "Kg": float(txn.weight or 0),
                "Rate": float(txn.rate or 0),
                "Amount": float(txn.amount or 0),
                "Mode": txn.payment_mode or ""
            })

        columns = ["Date", "Party", "Party Type", "Type", "Category", "Item", "Kg", "Rate", "Amount", "Mode"]
        return report_response(rows, columns, "transactions", file_format, "Transaction Report")

    return {"error": "Unknown report type"}


@app.get("/dashboard")
def get_dashboard(date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):

    target_date = parse_input_date(date)
    if not target_date:
        return {"error": "Invalid date format"}

    try:
        totals = db.query(
            func.sum(
                case(
                    (models.Transaction.type == "SALE", models.Transaction.amount),
                    else_=0
                )
            ).label("sales"),
            func.sum(
                case(
                    (models.Transaction.type == "PURCHASE", models.Transaction.amount),
                    else_=0
                )
            ).label("purchase"),
            func.sum(receivable_case()).label("receivable"),
            func.sum(payable_case()).label("payable")
        ).filter(
            models.Transaction.date <= target_date,
            outlet_scope_filter(models.Transaction, scope)
        ).first()

        daily_totals = db.query(
            func.sum(
                case(
                    (models.Transaction.type == "SALE", models.Transaction.amount),
                    else_=0
                )
            ).label("sales"),
            func.sum(
                case(
                    (models.Transaction.type == "PURCHASE", models.Transaction.amount),
                    else_=0
                )
            ).label("purchase")
        ).filter(
            models.Transaction.date == target_date,
            outlet_scope_filter(models.Transaction, scope)
        ).first()

        daily_operational = db.query(
            func.sum(
                case(
                    (
                        (models.Transaction.type == "SALE") &
                        (models.Transaction.category.in_(["RETAIL", "RETAIL DRESSED"])),
                        models.Transaction.amount
                    ),
                    else_=0
                )
            ).label("retail_sales"),
            func.sum(
                case(
                    (
                        (models.Transaction.type == "SALE") &
                        (models.Transaction.category == "RETAIL DRESSED"),
                        models.Transaction.amount
                    ),
                    else_=0
                )
            ).label("dressed_sales_amount"),
            func.sum(
                case(
                    (
                        (models.Transaction.type == "PAYMENT") &
                        (models.Transaction.category == "RECEIVED"),
                        models.Transaction.amount
                    ),
                    else_=0
                )
            ).label("payments_received"),
            func.sum(
                case(
                    (
                        (models.Transaction.type == "PAYMENT") &
                        (models.Transaction.category == "PAID"),
                        models.Transaction.amount
                    ),
                    else_=0
                )
            ).label("payments_paid"),
            func.sum(
                case(
                    (models.Transaction.type == "MORTALITY", models.Transaction.weight),
                    else_=0
                )
            ).label("mortality_weight"),
            func.sum(
                case(
                    (models.Transaction.type == "MORTALITY", models.Transaction.quantity),
                    else_=0
                )
            ).label("mortality_quantity")
        ).filter(
            models.Transaction.date == target_date,
            outlet_scope_filter(models.Transaction, scope)
        ).first()

        sales = daily_totals.sales or 0
        purchase = daily_totals.purchase or 0
        if ledger_mode_for_date(target_date) == ACCOUNT_LEDGER:
            party_balances = party_account_balances_as_of(db, scope, target_date)
            receivable = sum((balance["receivable"] for balance in party_balances.values()), Decimal("0"))
            payable = sum((balance["payable"] for balance in party_balances.values()), Decimal("0"))
        else:
            receivable = totals.receivable or 0
            payable = totals.payable or 0
        retail_sales = daily_operational.retail_sales or 0
        dressed_sales_amount = daily_operational.dressed_sales_amount or 0
        payments_received = daily_operational.payments_received or 0
        payments_paid = daily_operational.payments_paid or 0
        mortality_weight = daily_operational.mortality_weight or 0
        mortality_quantity = daily_operational.mortality_quantity or 0

        histories = scoped_stock_histories(db, scope, target_date)
        stock_totals = {
            key: stock_total_for_day(histories, target_date, key)
            for key in ("expected_closing_weight", "actual_closing_weight", "leakage", "gross_profit")
        }
        processed_rows = sum(
            sum(1 for row in history.snapshot(target_date)["rows"] if row["actual_closing_weight"] is not None)
            for history in histories
        )
        is_processed = stock_totals["actual_closing_weight"] is not None
        process_meta = f"{processed_rows} hen types counted" if processed_rows else "Actual live count not entered"
        profit = optional_float(stock_totals["gross_profit"])
        events = financial_events(db, scope, target_date, target_date)
        payments_received = sum(event["amount"] for event in events if event["type"] == "PAYMENT RECEIVED")
        payments_paid = sum(event["amount"] for event in events if event["type"] == "PAYMENT PAID")

    except Exception as e:
        return {"error": "Dashboard calculation failed", "details": str(e)}

    return {
        "date": str(target_date),

        "purchase": float(purchase or 0),
        "sales": float(sales or 0),
        "profit": profit,

        "expected_stock": optional_float(stock_totals["expected_closing_weight"]),
        "actual_stock": optional_float(stock_totals["actual_closing_weight"]),
        "leakage": optional_float(stock_totals["leakage"]),

        "receivable": float(receivable),
        "payable": float(payable),
        "total_outstanding": float(receivable - payable),
        "retail_sales": float(retail_sales or 0),
        "dressed_sales_amount": float(dressed_sales_amount or 0),
        "payments_received": float(payments_received or 0),
        "payments_paid": float(payments_paid or 0),
        "mortality_weight": float(mortality_weight or 0),
        "mortality_quantity": float(mortality_quantity or 0),
        "processed_items_count": int(processed_rows or 0),
        "process_status": "Processed" if is_processed else "Pending",
        "process_meta": process_meta
    }


@app.get("/top-debtors")
def top_debtors(start_date: str | None = None, end_date: str | None = None, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    start = parse_input_date(start_date) if start_date else None
    end = parse_input_date(end_date) if end_date else None
    if (start_date and not start) or (end_date and not end):
        return {"error": "Invalid date format"}

    as_of = end or ledger_today()
    if ledger_mode_for_date(as_of) == ACCOUNT_LEDGER:
        balances = party_account_balances_as_of(db, scope, as_of)
        party_ids = list(balances)
        parties = {
            party.id: party.name
            for party in db.query(models.Party).filter(models.Party.id.in_(party_ids)).all()
        } if party_ids else {}
        rows = sorted(
            ((parties.get(party_id, str(party_id)), values["receivable"])
             for party_id, values in balances.items() if values["receivable"] > 0),
            key=lambda row: row[1], reverse=True,
        )[:5]
        return {"top_debtors": [{"party_name": name, "balance": float(balance)} for name, balance in rows]}

    balance_expr = func.sum(receivable_case())
    query = db.query(
        models.Party.name,
        balance_expr.label("balance")
    ).join(
        models.Transaction,
        models.Transaction.party_id == models.Party.id
    )
    query = apply_outlet_scope(query, models.Transaction, scope)

    # Outstanding is a balance as at the selected day. Future-dated cutover
    # openings must not appear before 04/09/2026.
    query = query.filter(models.Transaction.date <= as_of)

    rows = query.group_by(
        models.Party.id,
        models.Party.name
    ).having(
        balance_expr > 0
    ).order_by(
        balance_expr.desc()
    ).limit(5).all()

    return {
        "top_debtors": [
            {"party_name": row.name, "balance": float(row.balance or 0)}
            for row in rows
        ]
    }


@app.get("/top-payables")
def top_payables(db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    as_of = ledger_today()
    if ledger_mode_for_date(as_of) == ACCOUNT_LEDGER:
        balances = party_account_balances_as_of(db, scope, as_of)
        party_ids = list(balances)
        parties = {
            party.id: party.name
            for party in db.query(models.Party).filter(models.Party.id.in_(party_ids)).all()
        } if party_ids else {}
        rows = sorted(
            ((parties.get(party_id, str(party_id)), values["payable"])
             for party_id, values in balances.items() if values["payable"] > 0),
            key=lambda row: row[1], reverse=True,
        )[:5]
        return {"top_payables": [{"party_name": name, "balance": float(balance)} for name, balance in rows]}

    balance_expr = func.sum(payable_case())
    rows = db.query(
        models.Party.name,
        balance_expr.label("balance")
    ).join(
        models.Transaction,
        models.Transaction.party_id == models.Party.id
    )
    rows = apply_outlet_scope(rows, models.Transaction, scope).filter(
        models.Transaction.date <= as_of
    ).group_by(
        models.Party.id,
        models.Party.name
    ).having(
        balance_expr > 0
    ).order_by(
        balance_expr.desc()
    ).limit(5).all()

    return {
        "top_payables": [
            {"party_name": row.name, "balance": float(row.balance or 0)}
            for row in rows
        ]
    }


@app.get("/analytics/trend")
def get_trend(start_date: str, end_date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):

    start = parse_input_date(start_date)
    end = parse_input_date(end_date)
    if not start or not end:
        return {"error": "Invalid date format"}

    results = db.query(
        models.Transaction.date,
        func.sum(
            case(
                (models.Transaction.type == "SALE", models.Transaction.amount),
                else_=0
            )
        ).label("sales"),
        func.sum(
            case(
                (models.Transaction.type == "PURCHASE", models.Transaction.amount),
                else_=0
            )
        ).label("purchase"),
        func.sum(
            case(
                (
                    (models.Transaction.type == "SALE") &
                    (models.Transaction.category == "RETAIL"),
                    models.Transaction.amount
                ),
                else_=0
            )
        ).label("regular_billing"),
        func.sum(
            case(
                (
                    (models.Transaction.type == "SALE") &
                    (models.Transaction.category == "RETAIL DRESSED"),
                    models.Transaction.amount
                ),
                else_=0
            )
        ).label("dressed_billing")
    ).filter(
        models.Transaction.date.between(start, end),
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(models.Transaction.date).order_by(models.Transaction.date).all()

    by_date = {
        r.date: {
            "sales": float(r.sales or 0),
            "purchase": float(r.purchase or 0),
            "regular_billing": float(r.regular_billing or 0),
            "dressed_billing": float(r.dressed_billing or 0)
        }
        for r in results
    }

    histories = scoped_stock_histories(db, scope, end)
    trend = []
    for day in pd.date_range(start=start, end=end):
        current_date = day.date()
        row = by_date.get(current_date, {"sales": 0, "purchase": 0, "regular_billing": 0, "dressed_billing": 0})
        trend.append({
            "date": str(current_date),
            "sales": row["sales"],
            "purchase": row["purchase"],
            "profit": optional_float(stock_total_for_day(histories, current_date, "gross_profit")),
            "regular_billing": row["regular_billing"],
            "dressed_billing": row["dressed_billing"]
        })

    return trend

@app.get("/analytics/leakage")
def leakage_trend(start_date: str, end_date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):

    start = parse_input_date(start_date)
    end = parse_input_date(end_date)
    if not start or not end:
        return {"error": "Invalid date format"}

    histories = scoped_stock_histories(db, scope, end)
    days = sorted({row.date for history in histories for row in history.actual_rows if start <= row.date <= end})
    return [{"date": str(day), "leakage": optional_float(stock_total_for_day(histories, day, "leakage"))}
            for day in days]


@app.get("/analytics/summary")
def analytics_summary(start_date: str, end_date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    start = parse_input_date(start_date)
    end = parse_input_date(end_date)
    if not start or not end:
        return {"error": "Invalid date format"}

    totals = db.query(
        func.sum(
            case(
                (models.Transaction.type == "SALE", models.Transaction.amount),
                else_=0
            )
        ).label("sales"),
        func.sum(
            case(
                (models.Transaction.type == "PURCHASE", models.Transaction.amount),
                else_=0
            )
        ).label("purchase"),
        func.sum(
            case(
                (
                    (models.Transaction.type == "PAYMENT") & (models.Transaction.category == "RECEIVED"),
                    models.Transaction.amount
                ),
                else_=0
            )
        ).label("received"),
        func.sum(
            case(
                (
                    (models.Transaction.type == "PAYMENT") & (models.Transaction.category == "PAID"),
                    models.Transaction.amount
                ),
                else_=0
            )
        ).label("paid")
    ).filter(
        models.Transaction.date.between(start, end),
        outlet_scope_filter(models.Transaction, scope)
    ).first()

    sales = Decimal(totals.sales or 0)
    purchase = Decimal(totals.purchase or 0)
    events = financial_events(db, scope, start, end)
    received = sum((event["amount"] for event in events if event["type"] == "PAYMENT RECEIVED"), Decimal("0"))
    paid = sum((event["amount"] for event in events if event["type"] == "PAYMENT PAID"), Decimal("0"))
    histories = scoped_stock_histories(db, scope, end)
    days = [day.date() for day in pd.date_range(start, end)]
    leakage = complete_sum(stock_total_for_day(histories, day, "leakage") for day in days)
    profit = complete_sum(stock_total_for_day(histories, day, "gross_profit") for day in days)

    return {
        "sales": float(sales),
        "purchase": float(purchase),
        "profit": optional_float(profit),
        "received": float(received),
        "paid": float(paid),
        "net_cash": float(received - paid),
        "leakage": optional_float(leakage)
    }


@app.get("/analytics/item-volume")
def item_volume(start_date: str, end_date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    start = parse_input_date(start_date)
    end = parse_input_date(end_date)
    if not start or not end:
        return {"error": "Invalid date format"}

    rows = db.query(
        models.Transaction.item_type.label("item"),
        func.sum(
            case(
                (models.Transaction.type == "PURCHASE", models.Transaction.weight),
                else_=0
            )
        ).label("purchase_kg"),
        func.sum(
            case(
                (models.Transaction.type == "SALE", models.Transaction.weight),
                else_=0
            )
        ).label("sales_kg")
    ).filter(
        models.Transaction.date.between(start, end),
        models.Transaction.item_type.isnot(None),
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        models.Transaction.item_type
    ).order_by(
        models.Transaction.item_type.asc()
    ).all()

    return [
        {
            "item": row.item,
            "purchase_kg": float(row.purchase_kg or 0),
            "sales_kg": float(row.sales_kg or 0)
        }
        for row in rows
    ]


@app.get("/analytics/payment-modes")
def payment_modes(start_date: str, end_date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    start = parse_input_date(start_date)
    end = parse_input_date(end_date)
    if not start or not end:
        return {"error": "Invalid date format"}

    mode = func.coalesce(models.Transaction.payment_mode, "NA").label("mode")
    rows = db.query(
        mode,
        func.sum(
            case(
                (models.Transaction.category == "RECEIVED", models.Transaction.amount),
                else_=0
            )
        ).label("received"),
        func.sum(
            case(
                (models.Transaction.category == "PAID", models.Transaction.amount),
                else_=0
            )
        ).label("paid")
    ).filter(
        models.Transaction.date.between(start, end),
        models.Transaction.type == "PAYMENT",
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        mode
    ).all()

    return [
        {
            "mode": row.mode,
            "received": float(row.received or 0),
            "paid": float(row.paid or 0),
            "total": float((row.received or 0) + (row.paid or 0))
        }
        for row in sorted(rows, key=lambda value: (value.received or 0) + (value.paid or 0), reverse=True)
    ]


@app.get("/retail-bills/next-number")
def next_retail_bill_number(date: str, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(date)
    if not target_date:
        return {"error": "Invalid date format"}
    return {"bill_number": current_shared_document_number(target_date, db, current_outlet.id)}


@app.get("/retail-bills")
def list_retail_bills(
    date: str = None,
    bill_number: str = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_outlet: models.Outlet = Depends(get_current_outlet)
):
    page_limit = min(max(int(limit or 50), 1), 100)
    page_offset = max(int(offset or 0), 0)
    query = db.query(models.RetailBill).order_by(
        models.RetailBill.date.desc(),
        models.RetailBill.created_at.desc()
    ).filter(models.RetailBill.outlet_id == current_outlet.id)

    if date:
        target_date = parse_input_date(date)
        if not target_date:
            return {"error": "Invalid date format"}
        query = query.filter(models.RetailBill.date == target_date)

    if bill_number:
        query = query.filter(models.RetailBill.bill_number == str(bill_number).strip())

    page = query.offset(page_offset).limit(page_limit + 1).all()
    has_more = len(page) > page_limit
    bills = page[:page_limit]
    bill_ids = [bill.id for bill in bills]
    mode_map = {}
    if bill_ids:
        item_rows = db.query(
            models.RetailBillItem.bill_id,
            models.RetailBillItem.line_type
        ).filter(
            models.RetailBillItem.bill_id.in_(bill_ids)
        ).all()
        for row in item_rows:
            current_mode = mode_map.get(row.bill_id)
            row_mode = "dressed" if str(row.line_type or "STANDARD").upper() == "DRESSED" else "regular"
            if current_mode and current_mode != row_mode:
                mode_map[row.bill_id] = "both"
            else:
                mode_map.setdefault(row.bill_id, row_mode)

    return {
        "results": [
            {
                "id": str(bill.id),
                "bill_number": bill.bill_number,
                "bill_mode": mode_map.get(bill.id, "regular"),
                "date": str(bill.date),
                "customer_name": bill.customer_name or "Walk-in Customer",
                "payment_mode": bill.payment_mode or "Cash",
                "total_amount": float(bill.total_amount or 0),
                "paid_amount": float(bill.paid_amount or 0),
                "outstanding_amount": float(bill.outstanding_amount or 0)
            }
            for bill in bills
        ],
        "has_more": has_more,
        "next_offset": page_offset + len(bills)
    }


@app.get("/payment-receipts/next-number")
def next_payment_receipt_number(date: str, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(date)
    if not target_date:
        return {"error": "Invalid date format"}
    return {"receipt_number": current_shared_document_number(target_date, db, current_outlet.id)}


@app.get("/payment-receipts")
def list_payment_receipts(date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    query = db.query(models.PaymentReceipt).order_by(
        models.PaymentReceipt.date.desc(),
        models.PaymentReceipt.created_at.desc()
    ).filter(models.PaymentReceipt.outlet_id == current_outlet.id)

    if date:
        target_date = parse_input_date(date)
        if not target_date:
            return {"error": "Invalid date format"}
        query = query.filter(models.PaymentReceipt.date == target_date)
        receipts = query.all()
    else:
        receipts = query.limit(50).all()

    return {
        "results": [
            {
                "id": str(receipt.id),
                "receipt_number": receipt.receipt_number,
                "date": str(receipt.date),
                "party_name": receipt.party_name or "",
                "direction": receipt.direction or "RECEIVED",
                "payment_mode": receipt.payment_mode or "Cash",
                "amount": float(receipt.amount or 0)
            }
            for receipt in receipts
        ]
    }


@app.get("/dressed-stock")
def list_dressed_stock(date: str | None = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    query = db.query(models.DressedStockEntry).order_by(
        models.DressedStockEntry.date.desc(),
        models.DressedStockEntry.created_at.desc()
    ).filter(models.DressedStockEntry.outlet_id == current_outlet.id)

    target_date = parse_input_date(date) if date else None
    if date and not target_date:
        return {"error": "Invalid date format"}
    if target_date:
        query = query.filter(models.DressedStockEntry.date == target_date)

    entries = query.all()
    grouped = {}
    for entry in entries:
        grouped.setdefault(entry.item_name, {
            "item_name": entry.item_name,
            "live_quantity": Decimal("0"),
            "live_weight": Decimal("0"),
            "available_dressed_weight": Decimal("0"),
            "default_rate": Decimal("0")
        })
        grouped[entry.item_name]["live_quantity"] += Decimal(entry.live_quantity or 0)
        grouped[entry.item_name]["live_weight"] += Decimal(entry.live_weight or 0)
        grouped[entry.item_name]["available_dressed_weight"] += Decimal(entry.remaining_dressed_weight or 0)
        if Decimal(grouped[entry.item_name]["default_rate"] or 0) <= 0 and Decimal(entry.default_rate or 0) > 0:
            grouped[entry.item_name]["default_rate"] = Decimal(entry.default_rate or 0)

    return {
        "entries": [serialize_dressed_stock_entry(entry) for entry in entries],
        "available_items": [
            {
                "item_name": item["item_name"],
                "live_quantity": float(item["live_quantity"]),
                "live_weight": float(item["live_weight"]),
                "available_dressed_weight": float(item["available_dressed_weight"]),
                "default_rate": float(item["default_rate"] or 0)
            }
            for item in grouped.values()
            if item["available_dressed_weight"] > 0
        ]
    }


@app.post("/dressed-stock")
def create_dressed_stock_entries(payload: dict = Body(...), input_date: str = None, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(input_date)
    if not target_date:
        return {"error": "Select working date"}

    rows = payload.get("rows") or []
    if not rows:
        return {"error": "Add at least one dressed stock row"}

    inserted = 0
    skipped = 0
    errors = []

    for index, row in enumerate(rows, start=1):
        try:
            item_name = normalize_source(row.get("source_item_type") or row.get("item_name") or row.get("hen_type"))
            live_quantity = parse_decimal(row.get("live_quantity", row.get("nag")))
            live_weight = parse_decimal(row.get("live_weight"))
            dressed_weight = parse_decimal(row.get("dressed_weight"))
            default_rate = parse_decimal(row.get("default_rate"))
            notes = str(row.get("notes") or "").strip() or None

            if item_name not in SOURCE_ITEMS or not all(value.is_finite() for value in (live_quantity, live_weight, dressed_weight, default_rate)) or live_quantity <= 0 or live_quantity != live_quantity.to_integral_value() or live_weight <= 0 or dressed_weight < 0 or dressed_weight > live_weight or default_rate < 0:
                skipped += 1
                row_error(errors, index, "Select BB, CB, COCREL, LEGOAN or DP and enter live NAG and kg; dressed kg cannot exceed live kg")
                continue

            db.add(models.DressedStockEntry(
                date=target_date,
                outlet_id=current_outlet.id,
                item_name=item_name,
                live_quantity=live_quantity if live_quantity > 0 else None,
                live_weight=live_weight if live_weight > 0 else None,
                dressed_weight=dressed_weight if dressed_weight > 0 else None,
                remaining_dressed_weight=dressed_weight if dressed_weight > 0 else None,
                default_rate=default_rate if default_rate > 0 else None,
                notes=notes
            ))
            inserted += 1
        except Exception as e:
            skipped += 1
            row_error(errors, index, str(e))

    if errors:
        db.rollback()
        return {"error": "Correct the dressed cutting rows before saving", "details": errors}

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving dressed stock failed", "details": str(e)}

    return upload_result(inserted, skipped, errors)


@app.get("/payment-receipts/{receipt_id}")
def get_payment_receipt(receipt_id: UUID, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    receipt = db.query(models.PaymentReceipt).filter(
        models.PaymentReceipt.id == receipt_id,
        models.PaymentReceipt.outlet_id == current_outlet.id
    ).first()
    if not receipt:
        return {"error": "Payment receipt not found"}
    txns = db.query(models.Transaction).filter(models.Transaction.party_id == receipt.party_id).order_by(
        models.Transaction.date.asc(),
        models.Transaction.created_at.asc()
    ).all()
    txns = [txn for txn in txns if txn.outlet_id == current_outlet.id]
    balances, _ = build_ledger(db, txns)
    return serialize_payment_receipt(receipt, payment_direction_balance(balances, receipt.direction))


@app.post("/payment-receipts")
def create_payment_receipt(payload: dict = Body(...), db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(payload.get("date"))
    if not target_date:
        return {"error": "Invalid receipt date"}

    party_name = str(payload.get("party_name") or "").strip()
    if not party_name:
        return {"error": "Party name is required"}

    amount = parse_decimal(payload.get("amount"))
    if amount <= 0:
        return {"error": "Amount must be greater than 0"}

    direction = normalize_payment_direction(payload.get("direction"))
    if not direction:
        return {"error": "Direction must be RECEIVED or PAID"}

    receipt_number = reserve_shared_document_number(target_date, db, current_outlet.id)

    party_phone = str(payload.get("party_phone") or "").strip()
    party_address = str(payload.get("party_address") or "").strip()
    cashier_name = str(payload.get("cashier_name") or "admin").strip()
    payment_mode = str(payload.get("payment_mode") or "Cash").strip()
    notes = str(payload.get("notes") or "").strip()

    party_id = get_or_create_party(db, party_name, "BOTH", {}, phone=party_phone, address=party_address)

    receipt = models.PaymentReceipt(
        id=uuid4(),
        receipt_number=receipt_number,
        date=target_date,
        outlet_id=current_outlet.id,
        party_id=party_id,
        party_name=party_name,
        party_phone=party_phone or None,
        party_address=party_address or None,
        cashier_name=cashier_name or "admin",
        direction=direction,
        payment_mode=payment_mode,
        amount=amount,
        notes=notes or None
    )
    db.add(receipt)
    db.flush()

    db.add(models.Transaction(
        date=target_date,
        outlet_id=current_outlet.id,
        party_id=party_id,
        type="PAYMENT",
        category=direction,
        amount=amount,
        payment_mode=payment_mode,
        bill_number=receipt_number,
        source_ref=f"payment-receipt:{receipt.id}"
    ))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Saving payment receipt failed", "details": str(e)}

    db.refresh(receipt)
    txns = db.query(models.Transaction).filter(
        models.Transaction.party_id == party_id,
        models.Transaction.outlet_id == current_outlet.id
    ).order_by(
        models.Transaction.date.asc(),
        models.Transaction.created_at.asc()
    ).all()
    balances, _ = build_ledger(db, txns)
    return {"receipt": serialize_payment_receipt(receipt, payment_direction_balance(balances, receipt.direction))}


@app.put("/payment-receipts/{receipt_id}")
def update_payment_receipt(receipt_id: UUID, payload: dict = Body(...), db: Session = Depends(get_db), user: models.User = Depends(require_owner), current_outlet: models.Outlet = Depends(get_current_outlet)):
    receipt = db.query(models.PaymentReceipt).filter(
        models.PaymentReceipt.id == receipt_id,
        models.PaymentReceipt.outlet_id == current_outlet.id
    ).first()
    if not receipt:
        return {"error": "Payment receipt not found"}

    target_date = parse_input_date(payload.get("date"))
    if not target_date:
        return {"error": "Invalid receipt date"}

    party_name = str(payload.get("party_name") or "").strip()
    if not party_name:
        return {"error": "Party name is required"}

    amount = parse_decimal(payload.get("amount"))
    if amount <= 0:
        return {"error": "Amount must be greater than 0"}

    direction = normalize_payment_direction(payload.get("direction"))
    if not direction:
        return {"error": "Direction must be RECEIVED or PAID"}

    receipt_number = str(payload.get("receipt_number") or "").strip()
    if not receipt_number:
        return {"error": "Receipt number is required"}

    existing = db.query(models.PaymentReceipt).filter(
        models.PaymentReceipt.date == target_date,
        models.PaymentReceipt.outlet_id == current_outlet.id,
        models.PaymentReceipt.receipt_number == receipt_number,
        models.PaymentReceipt.id != receipt.id
    ).first()
    existing_bill = db.query(models.RetailBill).filter(
        models.RetailBill.date == target_date,
        models.RetailBill.outlet_id == current_outlet.id,
        models.RetailBill.bill_number == receipt_number
    ).first()
    if existing or existing_bill:
        return {"error": "Receipt number already exists"}

    party_phone = str(payload.get("party_phone") or "").strip()
    party_address = str(payload.get("party_address") or "").strip()
    cashier_name = str(payload.get("cashier_name") or "admin").strip()
    payment_mode = str(payload.get("payment_mode") or "Cash").strip()
    notes = str(payload.get("notes") or "").strip()

    party_id = get_or_create_party(db, party_name, "BOTH", {}, phone=party_phone, address=party_address)

    receipt.receipt_number = receipt_number
    receipt.date = target_date
    receipt.outlet_id = current_outlet.id
    receipt.party_id = party_id
    receipt.party_name = party_name
    receipt.party_phone = party_phone or None
    receipt.party_address = party_address or None
    receipt.cashier_name = cashier_name or "admin"
    receipt.direction = direction
    receipt.payment_mode = payment_mode
    receipt.amount = amount
    receipt.notes = notes or None

    db.query(models.Transaction).filter(
        models.Transaction.source_ref == f"payment-receipt:{receipt.id}"
    ).delete(synchronize_session=False)

    db.add(models.Transaction(
        date=target_date,
        outlet_id=current_outlet.id,
        party_id=party_id,
        type="PAYMENT",
        category=direction,
        amount=amount,
        payment_mode=payment_mode,
        bill_number=receipt_number,
        source_ref=f"payment-receipt:{receipt.id}"
    ))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Updating payment receipt failed", "details": str(e)}

    db.refresh(receipt)
    txns = db.query(models.Transaction).filter(
        models.Transaction.party_id == party_id,
        models.Transaction.outlet_id == current_outlet.id
    ).order_by(
        models.Transaction.date.asc(),
        models.Transaction.created_at.asc()
    ).all()
    balances, _ = build_ledger(db, txns)
    return {"receipt": serialize_payment_receipt(receipt, payment_direction_balance(balances, receipt.direction))}


@app.get("/retail-bills/{bill_id}")
def get_retail_bill(bill_id: UUID, db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    bill = db.query(models.RetailBill).filter(
        models.RetailBill.id == bill_id,
        models.RetailBill.outlet_id == current_outlet.id
    ).first()
    if not bill:
        return {"error": "Retail bill not found"}

    items = db.query(models.RetailBillItem).filter(
        models.RetailBillItem.bill_id == bill.id
    ).order_by(models.RetailBillItem.line_order.asc()).all()

    data = serialize_retail_bill(bill, items)
    data["party_balance"] = retail_party_balance_after(db, bill.party_id, current_outlet.id)
    return data


@app.post("/retail-bills")
def create_retail_bill(payload: dict = Body(...), db: Session = Depends(get_db), current_outlet: models.Outlet = Depends(get_current_outlet)):
    target_date = parse_input_date(payload.get("date"))
    if not target_date:
        return {"error": "Invalid bill date"}

    raw_items = payload.get("items") or []
    if not raw_items:
        return {"error": "Add at least one retail item"}

    bill_number = reserve_shared_document_number(target_date, db, current_outlet.id)

    customer_name = str(payload.get("customer_name") or "").strip()
    customer_phone = str(payload.get("customer_phone") or "").strip()
    customer_address = str(payload.get("customer_address") or "").strip()
    cashier_name = str(payload.get("cashier_name") or "admin").strip()
    payment_mode = str(payload.get("payment_mode") or "Cash").strip()
    notes = str(payload.get("notes") or "").strip()
    ice_amount = parse_decimal(payload.get("ice_amount"))
    raw_paid_amount = payload.get("paid_amount")
    paid_amount = parse_decimal(raw_paid_amount)

    party_id = None
    if customer_name:
        party_id = get_or_create_party(db, customer_name, "VENDOR", {}, phone=customer_phone, address=customer_address)

    normalized_items = []
    total_quantity = Decimal("0")
    total_weight = Decimal("0")
    total_amount = Decimal("0")

    for index, raw_item in enumerate(raw_items, start=1):
        item_name = str(raw_item.get("item_name") or "").strip()
        if not item_name:
            return {"error": f"Item name missing on row {index}"}

        line_type = str(raw_item.get("line_type") or "STANDARD").strip().upper()
        if line_type not in ["STANDARD", "DRESSED"]:
            line_type = "STANDARD"
        source_item_type = resolve_stock_source_type(item_name, raw_item.get("source_item_type") or "")
        quantity = parse_decimal(raw_item.get("nag", raw_item.get("quantity")))
        rate = parse_decimal(raw_item.get("rate"))
        unit = str(raw_item.get("unit") or "KGS").strip().upper()
        weight = parse_decimal(raw_item.get("weight"))

        if line_type != "DRESSED" and quantity <= 0:
            return {"error": f"Quantity must be greater than 0 on row {index}"}

        if line_type != "DRESSED" and unit == "KGS" and weight <= 0:
            weight = quantity

        amount = parse_decimal(raw_item.get("amount"))
        if line_type == "DRESSED":
            unit = "KGS"
            if quantity < 0:
                return {"error": f"Quantity cannot be negative on row {index}"}
            if weight <= 0:
                return {"error": f"Kgs must be greater than 0 for dressed chicken on row {index}"}
            if amount <= 0 and rate > 0:
                amount = weight * rate
            if amount <= 0:
                return {"error": f"Amount must be greater than 0 for dressed chicken on row {index}"}
            rate = decimal_ratio(amount, weight)
            quantity = Decimal("0")
        elif amount <= 0:
            amount_base = weight if weight > 0 else quantity
            amount = amount_base * rate

        normalized_items.append({
            "line_order": index,
            "item_name": item_name,
            "line_type": line_type,
            "source_item_type": source_item_type,
            "quantity": quantity,
            "unit": unit,
            "weight": weight,
            "rate": rate,
            "amount": amount
        })

        total_quantity += quantity
        total_weight += weight
        total_amount += amount

    if ice_amount < 0:
        return {"error": "Ice amount cannot be negative"}

    total_amount += ice_amount

    if raw_paid_amount in [None, ""] and payment_mode.strip().upper() != "CREDIT":
        paid_amount = total_amount

    if paid_amount < 0:
        return {"error": "Paid amount cannot be negative"}

    if paid_amount > total_amount:
        paid_amount = total_amount

    outstanding_amount = total_amount - paid_amount
    if outstanding_amount > 0 and not customer_name:
        return {"error": "Customer name is required for credit retail bills"}

    if outstanding_amount > 0 and payment_mode.strip().upper() == "CASH":
        payment_mode = "Credit"

    transaction_party_id = party_id if party_id else None

    bill = models.RetailBill(
        id=uuid4(),
        bill_number=bill_number,
        date=target_date,
        outlet_id=current_outlet.id,
        party_id=party_id,
        customer_name=customer_name or None,
        customer_phone=customer_phone or None,
        customer_address=customer_address or None,
        cashier_name=cashier_name or "admin",
        payment_mode=payment_mode,
        total_quantity=total_quantity,
        total_weight=total_weight,
        ice_amount=ice_amount,
        total_amount=total_amount,
        paid_amount=paid_amount,
        outstanding_amount=outstanding_amount,
        notes=notes or None
    )
    db.add(bill)
    db.flush()

    for item in normalized_items:
        db.add(models.RetailBillItem(
            bill_id=bill.id,
            line_order=item["line_order"],
            item_name=item["item_name"],
            line_type=item["line_type"],
            source_item_type=item["source_item_type"] or None,
            quantity=item["quantity"],
            unit=item["unit"],
            weight=item["weight"],
            rate=item["rate"],
            amount=item["amount"]
        ))
        transaction_category = "RETAIL DRESSED" if item["line_type"] == "DRESSED" else "RETAIL"
        db.add(models.Transaction(
            date=target_date,
            outlet_id=current_outlet.id,
            party_id=transaction_party_id,
            type="SALE",
            category=transaction_category,
            item_type=item["source_item_type"] or item["item_name"],
            quantity=item["quantity"],
            weight=item["weight"],
            rate=item["rate"],
            amount=item["amount"],
            payment_mode=payment_mode,
            bill_number=bill_number,
            source_ref=f"retail-bill:{bill.id}:{item['line_order']}"
        ))

    if ice_amount > 0:
        db.add(models.Transaction(
            date=target_date,
            outlet_id=current_outlet.id,
            party_id=transaction_party_id,
            type="SALE",
            category="RETAIL DRESSED" if any(item["line_type"] == "DRESSED" for item in normalized_items) else "RETAIL",
            item_type="ICE",
            quantity=Decimal("0"),
            weight=Decimal("0"),
            rate=ice_amount,
            amount=ice_amount,
            payment_mode=payment_mode,
            bill_number=bill_number,
            source_ref=f"retail-bill:{bill.id}:ice"
        ))

    if paid_amount > 0 and outstanding_amount > 0:
        db.add(models.Transaction(
            date=target_date,
            outlet_id=current_outlet.id,
            party_id=party_id,
            type="PAYMENT",
            category="RECEIVED",
            item_type="Retail Bill Payment",
            quantity=Decimal("0"),
            weight=0,
            rate=0,
            amount=paid_amount,
            payment_mode=payment_mode,
            bill_number=bill_number,
            source_ref=f"retail-payment:{bill.id}"
        ))

    recompute_dressed_stock_remaining(db, target_date, current_outlet.id)
    db.commit()
    db.refresh(bill)

    saved_items = db.query(models.RetailBillItem).filter(
        models.RetailBillItem.bill_id == bill.id
    ).order_by(models.RetailBillItem.line_order.asc()).all()

    bill_data = serialize_retail_bill(bill, saved_items)
    bill_data["party_balance"] = retail_party_balance_after(db, bill.party_id, current_outlet.id)

    return {
        "status": "success",
        "message": "Retail bill created",
        "bill": bill_data
    }


@app.put("/retail-bills/{bill_id}")
def update_retail_bill(bill_id: UUID, payload: dict = Body(...), db: Session = Depends(get_db), user: models.User = Depends(require_owner), current_outlet: models.Outlet = Depends(get_current_outlet)):
    bill = db.query(models.RetailBill).filter(
        models.RetailBill.id == bill_id,
        models.RetailBill.outlet_id == current_outlet.id
    ).first()
    if not bill:
        return {"error": "Retail bill not found"}

    target_date = parse_input_date(payload.get("date"))
    if not target_date:
        return {"error": "Invalid bill date"}

    raw_items = payload.get("items") or []
    if not raw_items:
        return {"error": "Add at least one retail item"}

    bill_number = str(payload.get("bill_number") or "").strip()
    if not bill_number:
        return {"error": "Bill number is required"}

    existing = db.query(models.RetailBill).filter(
        models.RetailBill.date == target_date,
        models.RetailBill.outlet_id == current_outlet.id,
        models.RetailBill.bill_number == bill_number,
        models.RetailBill.id != bill.id
    ).first()
    existing_receipt = db.query(models.PaymentReceipt).filter(
        models.PaymentReceipt.date == target_date,
        models.PaymentReceipt.outlet_id == current_outlet.id,
        models.PaymentReceipt.receipt_number == bill_number
    ).first()
    if existing or existing_receipt:
        return {"error": "Bill number already exists"}

    customer_name = str(payload.get("customer_name") or "").strip()
    customer_phone = str(payload.get("customer_phone") or "").strip()
    customer_address = str(payload.get("customer_address") or "").strip()
    cashier_name = str(payload.get("cashier_name") or "admin").strip()
    payment_mode = str(payload.get("payment_mode") or "Cash").strip()
    notes = str(payload.get("notes") or "").strip()
    ice_amount = parse_decimal(payload.get("ice_amount"))
    raw_paid_amount = payload.get("paid_amount")
    paid_amount = parse_decimal(raw_paid_amount)

    party_id = None
    if customer_name:
        party_id = get_or_create_party(db, customer_name, "VENDOR", {}, phone=customer_phone, address=customer_address)

    normalized_items = []
    total_quantity = Decimal("0")
    total_weight = Decimal("0")
    total_amount = Decimal("0")

    for index, raw_item in enumerate(raw_items, start=1):
        item_name = str(raw_item.get("item_name") or "").strip()
        if not item_name:
            return {"error": f"Item name missing on row {index}"}

        line_type = str(raw_item.get("line_type") or "STANDARD").strip().upper()
        if line_type not in ["STANDARD", "DRESSED"]:
            line_type = "STANDARD"
        source_item_type = resolve_stock_source_type(item_name, raw_item.get("source_item_type") or "")
        quantity = parse_decimal(raw_item.get("nag", raw_item.get("quantity")))
        rate = parse_decimal(raw_item.get("rate"))
        unit = str(raw_item.get("unit") or "KGS").strip().upper()
        weight = parse_decimal(raw_item.get("weight"))

        if line_type != "DRESSED" and quantity <= 0:
            return {"error": f"Quantity must be greater than 0 on row {index}"}

        if line_type != "DRESSED" and unit == "KGS" and weight <= 0:
            weight = quantity

        amount = parse_decimal(raw_item.get("amount"))
        if line_type == "DRESSED":
            unit = "KGS"
            if quantity < 0:
                return {"error": f"Quantity cannot be negative on row {index}"}
            if weight <= 0:
                return {"error": f"Kgs must be greater than 0 for dressed chicken on row {index}"}
            if amount <= 0 and rate > 0:
                amount = weight * rate
            if amount <= 0:
                return {"error": f"Amount must be greater than 0 for dressed chicken on row {index}"}
            rate = decimal_ratio(amount, weight)
            quantity = Decimal("0")
        elif amount <= 0:
            amount_base = weight if weight > 0 else quantity
            amount = amount_base * rate

        normalized_items.append({
            "line_order": index,
            "item_name": item_name,
            "line_type": line_type,
            "source_item_type": source_item_type,
            "quantity": quantity,
            "unit": unit,
            "weight": weight,
            "rate": rate,
            "amount": amount
        })

        total_quantity += quantity
        total_weight += weight
        total_amount += amount

    if ice_amount < 0:
        return {"error": "Ice amount cannot be negative"}

    total_amount += ice_amount

    if raw_paid_amount in [None, ""] and payment_mode.strip().upper() != "CREDIT":
        paid_amount = total_amount

    if paid_amount < 0:
        return {"error": "Paid amount cannot be negative"}

    if paid_amount > total_amount:
        paid_amount = total_amount

    outstanding_amount = total_amount - paid_amount
    if outstanding_amount > 0 and not customer_name:
        return {"error": "Customer name is required for credit retail bills"}

    if outstanding_amount > 0 and payment_mode.strip().upper() == "CASH":
        payment_mode = "Credit"

    transaction_party_id = party_id if party_id else None

    previous_date = bill.date

    bill.bill_number = bill_number
    bill.date = target_date
    bill.outlet_id = current_outlet.id
    bill.party_id = party_id
    bill.customer_name = customer_name or None
    bill.customer_phone = customer_phone or None
    bill.customer_address = customer_address or None
    bill.cashier_name = cashier_name or "admin"
    bill.payment_mode = payment_mode
    bill.total_quantity = total_quantity
    bill.total_weight = total_weight
    bill.ice_amount = ice_amount
    bill.total_amount = total_amount
    bill.paid_amount = paid_amount
    bill.outstanding_amount = outstanding_amount
    bill.notes = notes or None

    db.query(models.RetailBillItem).filter(
        models.RetailBillItem.bill_id == bill.id
    ).delete(synchronize_session=False)

    db.query(models.Transaction).filter(
        or_(
            models.Transaction.source_ref.like(f"retail-bill:{bill.id}:%"),
            models.Transaction.source_ref == f"retail-payment:{bill.id}"
        )
    ).delete(synchronize_session=False)

    for item in normalized_items:
        db.add(models.RetailBillItem(
            bill_id=bill.id,
            line_order=item["line_order"],
            item_name=item["item_name"],
            line_type=item["line_type"],
            source_item_type=item["source_item_type"] or None,
            quantity=item["quantity"],
            unit=item["unit"],
            weight=item["weight"],
            rate=item["rate"],
            amount=item["amount"]
        ))
        transaction_category = "RETAIL DRESSED" if item["line_type"] == "DRESSED" else "RETAIL"
        db.add(models.Transaction(
            date=target_date,
            outlet_id=current_outlet.id,
            party_id=transaction_party_id,
            type="SALE",
            category=transaction_category,
            item_type=item["source_item_type"] or item["item_name"],
            quantity=item["quantity"],
            weight=item["weight"],
            rate=item["rate"],
            amount=item["amount"],
            payment_mode=payment_mode,
            bill_number=bill_number,
            source_ref=f"retail-bill:{bill.id}:{item['line_order']}"
        ))

    if ice_amount > 0:
        db.add(models.Transaction(
            date=target_date,
            outlet_id=current_outlet.id,
            party_id=transaction_party_id,
            type="SALE",
            category="RETAIL DRESSED" if any(item["line_type"] == "DRESSED" for item in normalized_items) else "RETAIL",
            item_type="ICE",
            quantity=Decimal("0"),
            weight=Decimal("0"),
            rate=ice_amount,
            amount=ice_amount,
            payment_mode=payment_mode,
            bill_number=bill_number,
            source_ref=f"retail-bill:{bill.id}:ice"
        ))

    if paid_amount > 0 and outstanding_amount > 0:
        db.add(models.Transaction(
            date=target_date,
            outlet_id=current_outlet.id,
            party_id=party_id,
            type="PAYMENT",
            category="RECEIVED",
            item_type="Retail Bill Payment",
            quantity=Decimal("0"),
            weight=0,
            rate=0,
            amount=paid_amount,
            payment_mode=payment_mode,
            bill_number=bill_number,
            source_ref=f"retail-payment:{bill.id}"
        ))

    recompute_dressed_stock_remaining(db, previous_date, current_outlet.id)
    if previous_date != target_date:
        recompute_dressed_stock_remaining(db, target_date, current_outlet.id)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": "Updating retail bill failed", "details": str(e)}

    db.refresh(bill)
    saved_items = db.query(models.RetailBillItem).filter(
        models.RetailBillItem.bill_id == bill.id
    ).order_by(models.RetailBillItem.line_order.asc()).all()

    bill_data = serialize_retail_bill(bill, saved_items)
    bill_data["party_balance"] = retail_party_balance_after(db, bill.party_id, current_outlet.id)

    return {
        "status": "success",
        "message": "Retail bill updated",
        "bill": bill_data
    }


@app.get("/daily-sheet")
def daily_sheet(
    date: str,
    sheet_type: str = "stock",
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
    scope=Depends(get_outlet_scope)
):
    if scope["mode"] == "all":
        return {"error": "Select one outlet for Daily Sheet"}

    target_date = parse_input_date(date)
    if not target_date:
        return {"error": "Invalid date format"}

    sheet_type = sheet_type.strip().lower()

    if sheet_type in ["vendor", "dealer"]:
        if sheet_type == "vendor":
            party_type_filter = or_(
                models.Party.type == "VENDOR",
                models.Party.type == "BOTH"
            )
            received_payment = (
                (models.Transaction.type == "PAYMENT") &
                (models.Transaction.category == "RECEIVED")
            )
            relevant_filter = (
                (models.Transaction.type == "SALE") |
                ((models.Transaction.type == "OPENING") & (models.Transaction.category == "RECEIVABLE")) |
                received_payment
            )

            def include_txn(txn):
                return (
                    (txn.type == "SALE") or
                    (txn.type == "OPENING" and (txn.category or "").upper() == "RECEIVABLE") or
                    (txn.type == "PAYMENT" and (txn.category or "").upper() == "RECEIVED")
                )

            def opening_belongs_to_old(txn, current_date):
                return txn.type == "OPENING" and (txn.category or "").upper() == "RECEIVABLE" and txn.date <= current_date

            def running_delta(txn, settled_keys):
                return receivable_delta(txn, settled_keys)

            def day_purchase_amount(txn, current_date, settled_keys):
                if txn.date != current_date or txn.type != "SALE":
                    return Decimal("0")
                return max(receivable_delta(txn, settled_keys), Decimal("0"))

            def day_payment_amount(txn, current_date, settled_keys):
                return Decimal(txn.amount or 0) if txn.date == current_date and (txn.type == "PAYMENT") and ((txn.category or "").upper() == "RECEIVED") else Decimal("0")
        else:
            party_type_filter = or_(
                models.Party.type == "DEALER",
                models.Party.type == "BOTH"
            )
            relevant_filter = (
                (models.Transaction.type == "PURCHASE") |
                ((models.Transaction.type == "OPENING") & (models.Transaction.category == "PAYABLE")) |
                ((models.Transaction.type == "PAYMENT") & (models.Transaction.category == "PAID"))
            )
            def include_txn(txn):
                return (
                    txn.type == "PURCHASE" or
                    (txn.type == "OPENING" and (txn.category or "").upper() == "PAYABLE") or
                    (txn.type == "PAYMENT" and (txn.category or "").upper() == "PAID")
                )

            def opening_belongs_to_old(txn, current_date):
                return txn.type == "OPENING" and (txn.category or "").upper() == "PAYABLE" and txn.date <= current_date

            def running_delta(txn, settled_keys):
                return payable_delta(txn)

            def day_purchase_amount(txn, current_date, settled_keys):
                return Decimal(txn.amount or 0) if txn.date == current_date and txn.type == "PURCHASE" else Decimal("0")

            def day_payment_amount(txn, current_date, settled_keys):
                return Decimal(txn.amount or 0) if txn.date == current_date and txn.type == "PAYMENT" and (txn.category or "").upper() == "PAID" else Decimal("0")

        party_txn_rows = db.query(
            models.Party.id.label("party_id"),
            models.Party.name.label("party_name"),
            models.Transaction
        ).join(
            models.Transaction,
            models.Transaction.party_id == models.Party.id
        ).filter(
            relevant_filter,
            party_type_filter,
            models.Transaction.date <= target_date,
            outlet_scope_filter(models.Transaction, scope)
        ).order_by(
            models.Party.name.asc(),
            models.Transaction.date.asc(),
            models.Transaction.created_at.asc()
        ).all()

        grouped_parties = {}
        for party_id, party_name, txn in party_txn_rows:
            bucket = grouped_parties.setdefault(party_id, {"party_name": party_name, "txns": []})
            bucket["txns"].append(txn)

        sheet_data = build_balance_sheet_rows_from_ledger(
            db=db,
            grouped_parties=grouped_parties,
            target_date=target_date,
            include_txn=include_txn,
            running_delta=running_delta,
            day_purchase_amount=day_purchase_amount,
            day_payment_amount=day_payment_amount,
            opening_belongs_to_old=opening_belongs_to_old
        )

        return {
            "date": str(target_date),
            "sheet_type": sheet_type,
            "title": "Vendor Balance Sheet" if sheet_type == "vendor" else "Dealer Balance Sheet",
            "rows": sheet_data["rows"],
            "totals": sheet_data["totals"]
        }

    stock = StockHistory(db, resolve_scope_outlet_id(scope), target_date).snapshot(target_date)
    stock_totals = stock["totals"]
    stock_warning = stock["warning"]
    opening_total_quantity = stock_totals["opening_quantity"]
    opening_total_weight = stock_totals["opening_weight"]
    opening_total_amount = stock_totals["opening_amount"]
    opening_rows = [
        format_sheet_row(row["item"], row["opening_weight"], row["opening_rate"],
                         row["opening_amount"], row["opening_quantity"])
        for row in stock["rows"]
    ]

    purchase_rows_raw = apply_outlet_scope(
        db.query(models.Transaction).filter(
            models.Transaction.date == target_date,
            models.Transaction.type == "PURCHASE"
        ),
        models.Transaction,
        scope
    ).order_by(models.Transaction.item_type.asc(), models.Transaction.party_id.asc()).all()

    purchase_rows = []
    purchase_total_quantity = None
    purchase_total_weight = Decimal("0")
    purchase_total_amount = Decimal("0")
    purchase_total_rate_weight = Decimal("0")
    for txn in purchase_rows_raw:
        nag = Decimal(txn.quantity or 0) if txn.quantity is not None else None
        weight = Decimal(txn.weight or 0)
        rate = Decimal(txn.rate or 0)
        amount = Decimal(txn.amount or 0)
        label = txn.item_type or "Unknown"
        purchase_rows.append(format_sheet_row(label, weight, rate, amount, nag))
        if nag is not None:
            purchase_total_quantity = Decimal(purchase_total_quantity or 0) + nag
        purchase_total_weight += weight
        purchase_total_amount += amount
        purchase_total_rate_weight += weight * rate

    transport_mortality_grouped = db.query(
        models.Transaction.item_type.label("item_type"),
        func.sum(models.Transaction.quantity).label("quantity"),
        func.sum(models.Transaction.weight).label("weight")
    ).filter(
        models.Transaction.date == target_date,
        models.Transaction.type == "MORTALITY",
        models.Transaction.category == "TRANSPORTATION MORTALITY",
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        models.Transaction.item_type
    ).order_by(
        models.Transaction.item_type.asc()
    ).all()

    transport_mortality_rows = []
    transport_mortality_total_quantity = None
    transport_mortality_total_weight = Decimal("0")
    for row in transport_mortality_grouped:
        quantity = Decimal(row.quantity or 0) if row.quantity is not None else None
        weight = Decimal(row.weight or 0)
        transport_mortality_rows.append(format_sheet_row(row.item_type or "Unknown", weight, Decimal("0"), Decimal("0"), quantity))
        if quantity is not None:
            transport_mortality_total_quantity = Decimal(transport_mortality_total_quantity or 0) + quantity
        transport_mortality_total_weight += weight

    shop_mortality_grouped = db.query(
        models.Transaction.item_type.label("item_type"),
        func.sum(models.Transaction.quantity).label("quantity"),
        func.sum(models.Transaction.weight).label("weight")
    ).filter(
        models.Transaction.date == target_date,
        models.Transaction.type == "MORTALITY",
        outlet_scope_filter(models.Transaction, scope),
        or_(
            models.Transaction.category == "SHOP MORTALITY",
            models.Transaction.category == "MORTALITY"
        )
    ).group_by(
        models.Transaction.item_type
    ).order_by(
        models.Transaction.item_type.asc()
    ).all()

    shop_mortality_rows = []
    shop_mortality_total_quantity = None
    shop_mortality_total_weight = Decimal("0")
    for row in shop_mortality_grouped:
        quantity = Decimal(row.quantity or 0) if row.quantity is not None else None
        weight = Decimal(row.weight or 0)
        shop_mortality_rows.append(format_sheet_row(row.item_type or "Unknown", weight, Decimal("0"), Decimal("0"), quantity))
        if quantity is not None:
            shop_mortality_total_quantity = Decimal(shop_mortality_total_quantity or 0) + quantity
        shop_mortality_total_weight += weight

    mortality_total_quantity = optional_decimal_sum(
        value for value in [transport_mortality_total_quantity, shop_mortality_total_quantity] if value is not None
    )
    mortality_total_weight = transport_mortality_total_weight + shop_mortality_total_weight

    sales_raw = apply_outlet_scope(
        db.query(models.Transaction).filter(
            models.Transaction.date == target_date,
            models.Transaction.type == "SALE"
        ),
        models.Transaction,
        scope
    ).order_by(models.Transaction.category.asc().nulls_last(), models.Transaction.item_type.asc()).all()

    sales_grouped = db.query(
        models.Transaction.category.label("category"),
        models.Transaction.item_type.label("item_type"),
        func.sum(models.Transaction.quantity).label("quantity"),
        func.sum(models.Transaction.weight).label("weight"),
        func.sum(models.Transaction.amount).label("amount")
    ).filter(
        models.Transaction.date == target_date,
        models.Transaction.type == "SALE",
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        models.Transaction.category,
        models.Transaction.item_type
    ).order_by(
        models.Transaction.category.asc().nulls_last(),
        models.Transaction.item_type.asc()
    ).all()

    sales_sections = {}
    for row in sales_grouped:
        section = (row.category or "OTHER").strip().upper()
        sales_sections.setdefault(section, [])
        quantity = Decimal(row.quantity or 0) if row.quantity is not None else None
        weight = Decimal(row.weight or 0)
        amount = Decimal(row.amount or 0)
        avg_rate = decimal_ratio(amount, weight)
        sales_sections[section].append(format_sheet_row(row.item_type or "Unknown", weight, avg_rate, amount, quantity))

    section_order = ["WHOLESALE", "HOTEL", "RETAIL", "RETAIL DRESSED", "RETAILS", "SHOP", "CUSTOMER", "OTHER"]
    ordered_sales_sections = []
    for section in section_order + [s for s in sales_sections.keys() if s not in section_order]:
        rows = sales_sections.get(section)
        if not rows:
            continue
        total_quantity = optional_decimal_sum(Decimal(str(row["nag"])) for row in rows if row["nag"] not in ["", None])
        total_weight = sum(Decimal(str(row["weight"])) for row in rows)
        total_amount = sum(Decimal(str(row["total"])) for row in rows)
        avg_rate = (total_amount / total_weight) if total_weight > 0 else Decimal("0")
        ordered_sales_sections.append({
            "title": section.title(),
            "rows": rows,
            "total": format_sheet_row("TOTAL", total_weight, avg_rate, total_amount, total_quantity)
        })

    total_sales_quantity = optional_decimal_sum(
        Decimal(t.quantity or 0) for t in sales_raw if t.quantity is not None
    )
    total_sales_weight = sum(Decimal(t.weight or 0) for t in sales_raw)
    total_sales_amount = sum(Decimal(t.amount or 0) for t in sales_raw)
    total_sales_rate = (total_sales_amount / total_sales_weight) if total_sales_weight > 0 else Decimal("0")

    total_purchase_rate = (purchase_total_amount / purchase_total_weight) if purchase_total_weight > 0 else Decimal("0")

    closing_quantity = stock_totals["expected_closing_quantity"]
    closing_weight = stock_totals["expected_closing_weight"]
    closing_amount = stock_totals["closing_amount"]
    closing_rate = (closing_amount / closing_weight) if closing_weight and closing_amount is not None else None
    actual_quantity = stock_totals["actual_closing_quantity"]
    actual_weight = stock_totals["actual_closing_weight"]
    actual_amount = stock_totals["actual_amount"]
    short_quantity = stock_totals["quantity_leakage"]
    short_weight = stock_totals["leakage"]
    short_amount = stock_totals["short_amount"]
    gross_profit = stock_totals["gross_profit"]

    retail_credit_bills = apply_outlet_scope(
        db.query(models.RetailBill).filter(
            models.RetailBill.date == target_date,
            models.RetailBill.outstanding_amount > 0
        ),
        models.RetailBill,
        scope
    ).order_by(
        models.RetailBill.customer_name.asc().nulls_last(),
        models.RetailBill.bill_number.asc()
    ).all()

    retail_credit_rows = []
    retail_credit_total = Decimal("0")
    retail_credit_paid = Decimal("0")
    retail_credit_outstanding = Decimal("0")
    for bill in retail_credit_bills:
        bill_total = Decimal(bill.total_amount or 0)
        bill_paid = Decimal(bill.paid_amount or 0)
        bill_outstanding = Decimal(bill.outstanding_amount or 0)
        retail_credit_rows.append(format_retail_credit_row(
            bill.customer_name,
            bill.bill_number,
            bill_total,
            bill_paid,
            bill_outstanding,
            bill.payment_mode
        ))
        retail_credit_total += bill_total
        retail_credit_paid += bill_paid
        retail_credit_outstanding += bill_outstanding

    purchase_rate_rows = []
    purchase_rate_query = db.query(
        models.Transaction.item_type.label("item_type"),
        func.sum(models.Transaction.weight).label("weight"),
        func.sum(models.Transaction.amount).label("amount")
    ).filter(
        models.Transaction.date == target_date,
        models.Transaction.type == "PURCHASE",
        models.Transaction.item_type.isnot(None),
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        models.Transaction.item_type
    ).order_by(
        models.Transaction.item_type.asc()
    ).all()
    for row in purchase_rate_query:
        weight = Decimal(row.weight or 0)
        amount = Decimal(row.amount or 0)
        avg_rate = decimal_ratio(amount, weight)
        purchase_rate_rows.append(
            format_rate_analysis_row(row.item_type, avg_rate, weight, amount)
        )

    category_rate_rows = []
    category_mix_rows = []
    category_rate_query = db.query(
        models.Transaction.category.label("category"),
        func.sum(models.Transaction.weight).label("weight"),
        func.sum(models.Transaction.amount).label("amount")
    ).filter(
        models.Transaction.date == target_date,
        models.Transaction.type == "SALE",
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        models.Transaction.category
    ).order_by(
        models.Transaction.category.asc().nulls_last()
    ).all()
    for row in category_rate_query:
        weight = Decimal(row.weight or 0)
        amount = Decimal(row.amount or 0)
        avg_rate = decimal_ratio(amount, weight)
        weight_share = (weight / total_sales_weight * Decimal("100")) if total_sales_weight > 0 else Decimal("0")
        amount_share = (amount / total_sales_amount * Decimal("100")) if total_sales_amount > 0 else Decimal("0")
        category_rate_rows.append(
            format_rate_analysis_row(row.category or "OTHER", avg_rate, weight, amount, category=row.category or "OTHER")
        )
        category_mix_rows.append({
            "category": row.category or "OTHER",
            "weight": float(weight),
            "amount": float(amount),
            "avg_rate": float(avg_rate),
            "weight_share": float(weight_share),
            "amount_share": float(amount_share)
        })

    category_item_rate_rows = []
    category_item_rate_query = db.query(
        models.Transaction.category.label("category"),
        models.Transaction.item_type.label("item_type"),
        func.sum(models.Transaction.weight).label("weight"),
        func.sum(models.Transaction.amount).label("amount")
    ).filter(
        models.Transaction.date == target_date,
        models.Transaction.type == "SALE",
        models.Transaction.item_type.isnot(None),
        outlet_scope_filter(models.Transaction, scope)
    ).group_by(
        models.Transaction.category,
        models.Transaction.item_type
    ).order_by(
        models.Transaction.category.asc().nulls_last(),
        models.Transaction.item_type.asc()
    ).all()
    for row in category_item_rate_query:
        weight = Decimal(row.weight or 0)
        amount = Decimal(row.amount or 0)
        category_item_rate_rows.append(
            format_rate_analysis_row(
                f"{row.category or 'OTHER'} - {row.item_type}",
                decimal_ratio(amount, weight),
                weight,
                amount,
                category=row.category or "OTHER",
                goods=row.item_type
            )
        )

    item_performance_rows = []
    for row in stock["rows"]:
        sold = row["sales_weight"] + row["dressed_sales_weight"]
        revenue = row["sales_amount"] + row["dressed_sales_amount"]
        sale_rate = revenue / sold if sold > 0 else None
        spread = sale_rate - row["rate"] if sale_rate is not None and row["rate"] is not None and not row["dressed_sales_weight"] else None
        if row["purchase_weight"] or sold:
            item_performance_rows.append({
                "item": row["item"], "purchase_kg": float(row["purchase_weight"]),
                "sales_kg": float(sold), "buy_rate": optional_float(row["rate"]),
                "sell_rate": optional_float(sale_rate), "spread": optional_float(spread),
                "gross_profit": optional_float(row["gross_profit"]),
            })

    total_purchase_rate_value = decimal_ratio(purchase_total_amount, purchase_total_weight) if purchase_total_weight > 0 else None
    total_sales_rate_value = decimal_ratio(total_sales_amount, total_sales_weight) if total_sales_weight > 0 else None
    dressed_sales_weight = sum(Decimal(str(row["total"]["weight"])) for row in ordered_sales_sections if row["title"].upper() == "RETAIL DRESSED")
    dressed_sales_amount = sum(Decimal(str(row["total"]["total"])) for row in ordered_sales_sections if row["title"].upper() == "RETAIL DRESSED")
    dressed_live_cut_weight = Decimal(
        apply_outlet_scope(
            db.query(func.coalesce(func.sum(models.DressedStockEntry.live_weight), 0)).filter(
                models.DressedStockEntry.date == target_date
            ),
            models.DressedStockEntry,
            scope
        ).scalar() or 0
    )
    dressed_yield_weight = Decimal(
        apply_outlet_scope(
            db.query(func.coalesce(func.sum(models.DressedStockEntry.dressed_weight), 0)).filter(
                models.DressedStockEntry.date == target_date
            ),
            models.DressedStockEntry,
            scope
        ).scalar() or 0
    )
    dressed_avg_on_live_weight = decimal_ratio(dressed_sales_amount, dressed_live_cut_weight) if dressed_live_cut_weight > 0 else None
    dressed_yield_percent = (dressed_yield_weight / dressed_live_cut_weight * Decimal("100")) if dressed_live_cut_weight > 0 else None
    available_live = complete_sum([opening_total_weight, stock_totals["purchase_weight"]])
    live_sales_weight = stock_totals["sales_weight"]
    live_sales_quantity = stock_totals["sales_quantity"]
    live_sales_amount = stock_totals["sales_amount"]
    live_cut_weight = stock_totals["cut_weight"]
    live_cut_quantity = stock_totals["cut_quantity"]
    sell_through = live_sales_weight / available_live * Decimal("100") if available_live and available_live > 0 else None
    leakage_percent = short_weight / closing_weight * Decimal("100") if short_weight is not None and closing_weight is not None and closing_weight > 0 else None
    realized_spread = total_sales_rate_value - total_purchase_rate_value if total_sales_rate_value is not None and total_purchase_rate_value is not None and not dressed_sales_weight else None
    retail_total_amount = sum(Decimal(str(section["total"]["total"])) for section in ordered_sales_sections if section["title"].upper() in ["RETAIL", "RETAIL DRESSED"])
    retail_mix_percent = (retail_total_amount / total_sales_amount * Decimal("100")) if total_sales_amount > 0 else Decimal("0")
    stock_coverage_days = closing_weight / live_sales_weight if live_sales_weight > 0 and closing_weight is not None and closing_weight >= 0 else None
    stock_math_invalid = closing_weight is None or closing_weight < 0
    missing_dressed_live_cut = dressed_sales_weight > 0 and dressed_live_cut_weight <= 0
    profit_percent = gross_profit / total_sales_amount * Decimal("100") if gross_profit is not None and total_sales_amount > 0 else (Decimal("0") if gross_profit == 0 else None)

    return {
        "date": str(target_date),
        "stock_warning": stock_warning,
        "opening_stock": {
            "rows": opening_rows,
            "total": format_sheet_row("TOTAL", opening_total_weight, (opening_total_amount / opening_total_weight) if opening_total_weight and opening_total_amount is not None else None, opening_total_amount, opening_total_quantity)
        },
        "purchase_stock": {
            "rows": purchase_rows,
            "total": format_sheet_row("TOTAL", purchase_total_weight, total_purchase_rate, purchase_total_amount, purchase_total_quantity)
        },
        "transport_mortality_stock": {
            "rows": transport_mortality_rows,
            "total": format_sheet_row("TOTAL", transport_mortality_total_weight, Decimal("0"), Decimal("0"), transport_mortality_total_quantity)
        },
        "shop_mortality_stock": {
            "rows": shop_mortality_rows,
            "total": format_sheet_row("TOTAL", shop_mortality_total_weight, Decimal("0"), Decimal("0"), shop_mortality_total_quantity)
        },
        "sales_sections": ordered_sales_sections,
        "final_stock": {
            "total_purchases": format_sheet_row("TOTAL PURCHASES", purchase_total_weight, total_purchase_rate, purchase_total_amount, purchase_total_quantity),
            "transport_mortality": format_sheet_row("TRANSPORTATION MORTALITY", transport_mortality_total_weight, Decimal("0"), Decimal("0"), transport_mortality_total_quantity),
            "shop_mortality": format_sheet_row("SHOP MORTALITY", shop_mortality_total_weight, Decimal("0"), Decimal("0"), shop_mortality_total_quantity),
            "sales": format_sheet_row("LIVE BIRD SALES", live_sales_weight, decimal_ratio(live_sales_amount, live_sales_weight), live_sales_amount, live_sales_quantity),
            "live_cut": format_sheet_row("LIVE BIRDS TAKEN FOR DRESSING", live_cut_weight, None, None, live_cut_quantity),
            "closing_stock": format_sheet_row("CLOSING STOCK", closing_weight, closing_rate, closing_amount, closing_quantity),
            "actual_stock": format_sheet_row("ACTUAL STOCK", actual_weight, closing_rate, actual_amount, actual_quantity),
            "short_by": format_sheet_row("SHORT BY", short_weight, closing_rate, short_amount, short_quantity),
            "gross_profit": {
                "rate": optional_float(profit_percent),
                "total": optional_float(gross_profit)
            }
        },
        "retail_credit_sheet": {
            "rows": retail_credit_rows,
            "total": {
                "label": "TOTAL CREDIT",
                "total_amount": float(retail_credit_total),
                "paid_amount": float(retail_credit_paid),
                "outstanding_amount": float(retail_credit_outstanding)
            }
        },
        "meta": {
            "nag_available": any(value is not None for value in [opening_total_quantity, purchase_total_quantity, total_sales_quantity, actual_quantity]),
            "actual_entered": actual_weight is not None,
            "stock_basis": "LIVE BIRDS"
        },
        "rate_analysis": {
            "purchase_by_hen_type": purchase_rate_rows,
            "sales_by_category": category_rate_rows,
            "sales_by_hen_type_category": category_item_rate_rows
        },
        "business_controls": {
            "category_mix": category_mix_rows,
            "item_performance": item_performance_rows
        },
        "metric_cards": [
            {
                "label": "Opening Stock",
                "value": optional_float(opening_total_weight),
                "display_value": "Not entered" if opening_total_weight is None else None,
                "suffix": " kg",
                "subvalue": opening_total_quantity is not None and f"{float(opening_total_quantity):.0f} NAG" or None
            },
            {
                "label": "Purchases",
                "value": float(purchase_total_weight),
                "suffix": " kg",
                "subvalue": purchase_total_quantity is not None and f"{float(purchase_total_quantity):.0f} NAG" or None
            },
            {
                "label": "Live Sales",
                "value": float(live_sales_weight),
                "suffix": " kg",
                "subvalue": live_sales_quantity is not None and f"{float(live_sales_quantity):.0f} NAG" or None
            },
            {
                "label": "Live Birds Taken for Dressing",
                "value": optional_float(live_cut_weight),
                "display_value": "Not entered" if live_cut_weight is None else None,
                "suffix": " kg",
                "subvalue": f"{float(live_cut_quantity):.0f} NAG" if live_cut_quantity is not None else "NAG not entered"
            },
            {
                "label": "Transport Mortality",
                "value": float(transport_mortality_total_weight),
                "suffix": " kg",
                "subvalue": transport_mortality_total_quantity is not None and f"{float(transport_mortality_total_quantity):.0f} NAG" or None
            },
            {
                "label": "Shop Mortality",
                "value": float(shop_mortality_total_weight),
                "suffix": " kg",
                "subvalue": shop_mortality_total_quantity is not None and f"{float(shop_mortality_total_quantity):.0f} NAG" or None
            },
            {
                "label": "Expected Closing",
                "value": optional_float(closing_weight),
                "display_value": "N/A" if closing_weight is None else None,
                "suffix": " kg",
                "subvalue": "Live stock cannot be reconciled" if stock_math_invalid else (closing_quantity is not None and f"{float(closing_quantity):.0f} NAG" or None)
            },
            {
                "label": "Actual Stock",
                "value": optional_float(actual_weight),
                "display_value": "Not entered" if actual_weight is None else None,
                "suffix": " kg",
                "subvalue": actual_quantity is not None and f"{float(actual_quantity):.0f} NAG" or None
            },
            {
                "label": "Short By",
                "value": optional_float(short_weight),
                "display_value": "N/A" if short_weight is None else None,
                "suffix": " kg",
                "subvalue": "Invalid until stock source is corrected" if stock_math_invalid else (short_quantity is not None and f"{float(short_quantity):.0f} NAG" or None)
            },
            {
                "label": "Avg Buy Rate",
                "value": optional_float(total_purchase_rate_value),
                "display_value": "N/A" if total_purchase_rate_value is None else None,
                "suffix": "/kg"
            },
            {
                "label": "Avg Sale Rate",
                "value": optional_float(total_sales_rate_value),
                "display_value": "N/A" if total_sales_rate_value is None else None,
                "suffix": "/kg"
            },
            {
                "label": "Leakage %",
                "value": float(leakage_percent) if leakage_percent is not None else None,
                "suffix": "%",
                "display_value": "N/A" if leakage_percent is None else None,
                "subvalue": "Live stock cannot be reconciled" if stock_math_invalid else None
            },
            {
                "label": "Sell Through",
                "value": optional_float(sell_through),
                "display_value": "N/A" if sell_through is None else None,
                "suffix": "%"
            },
            {
                "label": "Sale-Buy Spread",
                "value": optional_float(realized_spread),
                "display_value": "N/A" if realized_spread is None else None,
                "suffix": "/kg"
            },
            {
                "label": "Retail Mix",
                "value": float(retail_mix_percent),
                "suffix": "%"
            },
            {
                "label": "Retail Credit",
                "value": float(retail_credit_outstanding),
                "prefix": "Rs "
            },
            {
                "label": "Dressed Sale",
                "value": float(dressed_sales_weight),
                "suffix": " kg",
                "subvalue": f"Rs {float(dressed_sales_amount):,.2f}" if dressed_sales_amount > 0 else None
            },
            {
                "label": "Dressed Avg",
                "value": float(dressed_avg_on_live_weight) if dressed_avg_on_live_weight is not None else None,
                "suffix": "/live kg",
                "display_value": "N/A" if dressed_avg_on_live_weight is None else None,
                "subvalue": missing_dressed_live_cut and "Live cut weight missing" or (f"Live cut {float(dressed_live_cut_weight):.3f} kg" if dressed_live_cut_weight > 0 else None)
            },
            {
                "label": "Stock Cover",
                "value": float(stock_coverage_days) if stock_coverage_days is not None else None,
                "suffix": " days",
                "display_value": "N/A" if stock_coverage_days is None else None,
                "subvalue": "Live stock cannot be reconciled" if stock_math_invalid else None
            },
            {
                "label": "Gross Profit",
                "value": optional_float(profit_percent),
                "display_value": "N/A" if gross_profit is None else None,
                "suffix": "%",
                "subvalue": f"Rs {float(gross_profit):,.2f}" if gross_profit is not None else "Cost unavailable"
            }
        ],
        "special_sections": {
            "dressed_retail": next((section for section in ordered_sales_sections if section["title"].upper() == "RETAIL DRESSED"), None),
            "dressed_cutting_summary": {
                "live_weight_cut": float(dressed_live_cut_weight),
                "dressed_weight_prepared": float(dressed_yield_weight),
                "dressed_weight_sold": float(dressed_sales_weight),
                "dressed_sales_amount": float(dressed_sales_amount),
                "avg_amount_per_live_kg": float(dressed_avg_on_live_weight) if dressed_avg_on_live_weight is not None else None,
                "yield_percent": float(dressed_yield_percent) if dressed_yield_percent is not None else None,
                "warning": "Live cut weight missing for dressed entries." if missing_dressed_live_cut else None
            }
        }
    }


@app.get("/daily-sheet/export")
def export_daily_sheet(
    date: str,
    sheet_type: str = "stock",
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
    scope=Depends(get_outlet_scope)
):
    sheet_payload = daily_sheet(date=date, sheet_type=sheet_type, db=db, user=user, scope=scope)
    if isinstance(sheet_payload, dict) and sheet_payload.get("error"):
        return sheet_payload

    target_date = parse_input_date(date)
    if not target_date:
        return {"error": "Invalid date format"}

    return build_daily_sheet_export_report(sheet_payload, sheet_type.strip().lower(), target_date)


@app.get("/inventory/by-item")
def inventory_by_item(date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    if scope["mode"] == "all":
        return {"error": "Select one outlet for inventory details", "inventory": []}

    target_date = parse_input_date(date)
    if not target_date:
        return {"error": "Invalid date format"}

    snapshot = StockHistory(db, resolve_scope_outlet_id(scope), target_date).snapshot(target_date)
    return {
        "warning": snapshot["warning"],
        "inventory": [{
            "item": row["item"], "opening_date": str(target_date),
            **{key: optional_float(row[key]) for key in (
                "opening_weight", "purchase_weight", "sales_weight",
                "expected_closing_weight", "actual_closing_weight", "leakage",
                "cut_weight", "mortality_weight",
            )},
            "closing_weight": optional_float(row["actual_closing_weight"]),
        } for row in snapshot["rows"]],
    }


@app.get("/items/search")
def search_items(q: str = "", db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    normalized_query = q.strip().lower()
    items = stock_item_names_query(db, scope)

    results = sorted(
        item for item in items
        if not normalized_query or normalized_query in item.lower()
    )[:20]

    return {"results": results}


@app.get("/items/tracked")
def tracked_items(db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    if scope["mode"] == "all":
        return {"error": "Select one outlet for tracked items", "items": []}

    return {"items": PROCESS_DAY_SOURCE_ITEMS}


@app.get("/analytics/profit-by-item")
def profit_by_item(start_date: str, end_date: str, db: Session = Depends(get_db), scope=Depends(get_outlet_scope)):
    start = parse_input_date(start_date)
    end = parse_input_date(end_date)
    if not start or not end:
        return {"error": "Invalid date format"}

    histories = scoped_stock_histories(db, scope, end)
    grouped = {item: [] for item in SOURCE_ITEMS}
    for day in pd.date_range(start, end):
        for history in histories:
            for row in history.snapshot(day.date())["rows"]:
                grouped[row["item"]].append(row)
    return [{
        "item": item,
        "sales": float(sum((row["sales_amount"] + row["dressed_sales_amount"] for row in rows), Decimal("0"))),
        "purchase": float(sum((row["purchase_amount"] for row in rows), Decimal("0"))),
        "profit": optional_float(complete_sum(row["gross_profit"] for row in rows)),
    } for item, rows in grouped.items() if any(row["sales_amount"] or row["dressed_sales_amount"] or row["purchase_amount"] for row in rows)]

if __name__ == "__main__":
    uvicorn.run("app.main:app", host="0.0.0.0", port=10000)
